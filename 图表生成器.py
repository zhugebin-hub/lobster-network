#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
高三二模考试分析图表生成器
生成 Excel 格式的图表模板，包含数据分析所需的各种图表
"""

import json
from openpyxl import Workbook
from openpyxl.chart import (
    BarChart, LineChart, PieChart, ScatterChart, RadarChart,
    Reference, Series
)
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_analysis_workbook():
    """创建分析工作簿"""
    wb = Workbook()
    wb.remove(wb.active)  # 移除默认 sheet
    
    # 创建各个工作表
    create_summary_sheet(wb)
    create_class_comparison_sheet(wb)
    create_subject_analysis_sheet(wb)
    create_student_analysis_sheet(wb)
    create_progress_sheet(wb)
    create_knowledge_sheet(wb)
    
    return wb

def style_cell(cell, bold=False, fill_color=None, align='center'):
    """设置单元格样式"""
    if bold:
        cell.font = Font(bold=True)
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    cell.alignment = Alignment(horizontal=align, vertical='center')

def create_summary_sheet(wb):
    """创建总览分析表"""
    ws = wb.create_sheet("总览分析")
    
    # 标题
    ws['A1'] = "高三二模考试总览分析"
    ws['A1'].font = Font(bold=True, size=16)
    
    # 分数段分布数据
    ws['A3'] = "分数段分布"
    ws['A3'].font = Font(bold=True)
    
    headers = ['分数段', '人数', '占比']
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
        style_cell(ws.cell(row=4, column=col), bold=True, fill_color='DDDDDD')
    
    # 示例数据
    ranges = ['680+', '650-679', '620-649', '600-619', '580-599', '560-579', 
              '540-559', '520-539', '500-519', '480-499', '460-479', '440-459', '400-439', '400 以下']
    for i, r in enumerate(ranges, 5):
        ws.cell(row=i, column=1, value=r)
        ws.cell(row=i, column=2, value=0)  # 人数待填
        ws.cell(row=i, column=3, value=0)  # 占比待填
    
    # 创建柱状图
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "分数段分布图"
    chart.y_axis.title = "人数"
    chart.x_axis.title = "分数段"
    
    data = Reference(ws, min_col=2, min_row=4, max_row=4+len(ranges), max_col=2)
    cats = Reference(ws, min_col=1, min_row=5, max_row=4+len(ranges))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, "E3")
    
    # 核心指标
    ws['A20'] = "核心指标"
    ws['A20'].font = Font(bold=True)
    
    indicators = ['参考人数', '平均分', '最高分', '最低分', '一本线', '一本人数', '一本率', 
                  '二本线', '二本人数', '二本率', '特控线', '特控率']
    for i, ind in enumerate(indicators, 21):
        ws.cell(row=i, column=1, value=ind)
        ws.cell(row=i, column=2, value=0)  # 数值待填
        ws.cell(row=i, column=3, value="")  # 对比待填
    
    style_cell(ws['A20'], bold=True)
    for row in range(21, 21+len(indicators)):
        style_cell(ws.cell(row=row, column=1), align='left')

def create_class_comparison_sheet(wb):
    """创建班级对比表"""
    ws = wb.create_sheet("班级对比")
    
    ws['A1'] = "班级成绩对比分析"
    ws['A1'].font = Font(bold=True, size=16)
    
    # 表头
    headers = ['班级', '人数', '平均分', '最高分', '最低分', '一本率%', '二本率%', '年级排名']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
        style_cell(ws.cell(row=3, column=col), bold=True, fill_color='DDDDDD')
    
    # 班级数据行
    for i in range(1, 13):
        ws.cell(row=4+i-1, column=1, value=f"高三 ({i}) 班")
        for col in range(2, 9):
            ws.cell(row=4+i-1, column=col, value=0)
    
    # 创建平均分对比图
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "班级平均分对比"
    chart.y_axis.title = "平均分"
    chart.x_axis.title = "班级"
    
    data = Reference(ws, min_col=3, min_row=3, max_row=15, max_col=3)
    cats = Reference(ws, min_col=1, min_row=4, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "I3")
    
    # 一本率对比图
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "班级一本率对比"
    chart2.y_axis.title = "一本率%"
    
    data2 = Reference(ws, min_col=6, min_row=3, max_row=15, max_col=6)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, "I20")

def create_subject_analysis_sheet(wb):
    """创建学科分析表"""
    ws = wb.create_sheet("学科分析")
    
    ws['A1'] = "学科质量分析"
    ws['A1'].font = Font(bold=True, size=16)
    
    # 学科三率一分表
    ws['A3'] = "各学科三率一分"
    ws['A3'].font = Font(bold=True)
    
    headers = ['学科', '参考人数', '平均分', '及格率%', '优秀率%', '低分率%', '标准差']
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
        style_cell(ws.cell(row=4, column=col), bold=True, fill_color='DDDDDD')
    
    subjects = ['语文', '数学', '英语', '物理', '化学', '生物', '政治', '历史', '地理']
    for i, sub in enumerate(subjects, 5):
        ws.cell(row=i, column=1, value=sub)
        for col in range(2, 8):
            ws.cell(row=i, column=col, value=0)
    
    # 学科贡献率雷达图
    chart = RadarChart()
    chart.title = "学科贡献率雷达图"
    
    ws['H4'] = "学科"
    ws['I4'] = "贡献率"
    for i, sub in enumerate(subjects, 5):
        ws.cell(row=i, column=8, value=sub)
        ws.cell(row=i, column=9, value=0)
    
    data = Reference(ws, min_col=9, min_row=4, max_row=4+len(subjects), max_col=9)
    cats = Reference(ws, min_col=8, min_row=5, max_row=4+len(subjects))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "H15")
    
    # 学科对比折线图
    ws['A20'] = "学科平均分对比"
    ws['A20'].font = Font(bold=True)
    
    chart2 = LineChart()
    chart2.title = "学科平均分趋势"
    chart2.y_axis.title = "平均分"
    
    data2 = Reference(ws, min_col=3, min_row=4, max_row=4+len(subjects), max_col=3)
    cats2 = Reference(ws, min_col=1, min_row=5, max_row=4+len(subjects))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, "E20")

def create_student_analysis_sheet(wb):
    """创建学生分析表"""
    ws = wb.create_sheet("学生分析")
    
    ws['A1'] = "学生分层分析"
    ws['A1'].font = Font(bold=True, size=16)
    
    # 尖子生分析
    ws['A3'] = "尖子生分析 (年级前 50 名)"
    ws['A3'].font = Font(bold=True, size=12)
    
    headers = ['姓名', '班级', '总分', '年级排名', '语文', '数学', '英语', '综合', '优势学科', '薄弱学科']
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
        style_cell(ws.cell(row=5, column=col), bold=True, fill_color='DDDDDD')
    
    for i in range(6, 16):
        for col in range(1, 11):
            ws.cell(row=i, column=col, value="")
    
    # 临界生分析
    ws['A18'] = "一本临界生分析 (一本线±20 分)"
    ws['A18'].font = Font(bold=True, size=12)
    
    headers2 = ['姓名', '班级', '总分', '与线差距', '语文', '数学', '英语', '综合', '增分点']
    for col, h in enumerate(headers2, 1):
        ws.cell(row=20, column=col, value=h)
        style_cell(ws.cell(row=20, column=col), bold=True, fill_color='DDDDDD')
    
    for i in range(21, 31):
        for col in range(1, 10):
            ws.cell(row=i, column=col, value="")

def create_progress_sheet(wb):
    """创建进退步分析表"""
    ws = wb.create_sheet("进退步分析")
    
    ws['A1'] = "学生进退步分析"
    ws['A1'].font = Font(bold=True, size=16)
    
    # 进步学生
    ws['A3'] = "进步显著学生 (进步 50 名以上)"
    ws['A3'].font = Font(bold=True, size=12)
    
    headers = ['姓名', '班级', '一模排名', '二模排名', '进步名次', '进步原因']
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
        style_cell(ws.cell(row=5, column=col), bold=True, fill_color='90EE90')
    
    for i in range(6, 16):
        for col in range(1, 7):
            ws.cell(row=i, column=col, value="")
    
    # 退步学生
    ws['A18'] = "退步明显学生 (退步 50 名以上)"
    ws['A18'].font = Font(bold=True, size=12)
    
    headers2 = ['姓名', '班级', '一模排名', '二模排名', '退步名次', '可能原因', '对策']
    for col, h in enumerate(headers2, 1):
        ws.cell(row=20, column=col, value=h)
        style_cell(ws.cell(row=20, column=col), bold=True, fill_color='FFB6C1')
    
    for i in range(21, 31):
        for col in range(1, 8):
            ws.cell(row=i, column=col, value="")
    
    # 进退步散点图
    ws['H3'] = "进退步散点图"
    ws['H3'].font = Font(bold=True)
    
    ws['H5'] = "学生"
    ws['I5'] = "一模排名"
    ws['J5'] = "二模排名"
    
    for i in range(6, 16):
        ws.cell(row=i, column=7, value=f"学生{i-5}")
        ws.cell(row=i, column=8, value=0)
        ws.cell(row=i, column=9, value=0)
    
    chart = ScatterChart()
    chart.title = "一模 vs 二模排名散点图"
    chart.x_axis.title = "一模排名"
    chart.y_axis.title = "二模排名"
    
    xdata = Reference(ws, min_col=8, min_row=5, max_row=15, max_col=8)
    ydata = Reference(ws, min_col=9, min_row=5, max_row=15, max_col=9)
    chart.add_data(ydata, titles_from_data=True)
    chart.add_data(xdata, titles_from_data=True)
    ws.add_chart(chart, "H20")

def create_knowledge_sheet(wb):
    """创建知识点分析表"""
    ws = wb.create_sheet("知识点分析")
    
    ws['A1'] = "知识点掌握分析"
    ws['A1'].font = Font(bold=True, size=16)
    
    # 数学知识点
    ws['A3'] = "数学知识点得分率"
    ws['A3'].font = Font(bold=True)
    
    headers = ['知识模块', '分值', '平均分', '得分率%', '掌握程度']
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
        style_cell(ws.cell(row=5, column=col), bold=True, fill_color='DDDDDD')
    
    math_topics = ['集合逻辑', '函数导数', '三角函数', '数列', '立体几何', 
                   '解析几何', '概率统计', '向量', '不等式']
    for i, topic in enumerate(math_topics, 6):
        ws.cell(row=i, column=1, value=topic)
        for col in range(2, 5):
            ws.cell(row=i, column=col, value=0)
        ws.cell(row=i, column=5, value="")
    
    # 知识点热力图数据
    ws['H3'] = "薄弱知识点 TOP10"
    ws['H3'].font = Font(bold=True)
    
    headers2 = ['排名', '学科', '知识点', '得分率%', '需强化']
    for col, h in enumerate(headers2, 1):
        ws.cell(row=5, column=col+7, value=h)
        style_cell(ws.cell(row=5, column=col+7), bold=True, fill_color='FF6B6B')
    
    for i in range(6, 16):
        ws.cell(row=i, column=7, value=i-5)
        for col in range(8, 12):
            ws.cell(row=i, column=col, value="")

def save_workbook(wb, filename):
    """保存工作簿"""
    wb.save(filename)
    print(f"图表模板已生成：{filename}")

if __name__ == "__main__":
    wb = create_analysis_workbook()
    save_workbook(wb, "高三二模考试分析图表模板.xlsx")
    print("生成完成！")
