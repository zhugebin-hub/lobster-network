#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成九年级课程总表 Excel 文件 - 修复版
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 创建 workbook
wb = Workbook()
ws = wb.active
ws.title = "九年级课程总表"

# 样式定义
title_font = Font(name='Arial', size=18, bold=True, color='1F4E79')
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
normal_font = Font(name='Arial', size=10)
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
title_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
alt_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

def set_center(cell):
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 设置列宽
for col in range(1, 10):
    ws.column_dimensions[chr(64+col)].width = 12

# ========== 标题 ==========
ws.merge_cells('A1:I1')
title = ws['A1']
title.value = "九年级课程总表"
title.font = title_font
title.fill = title_fill
title.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

# ========== 基本信息 ==========
ws.merge_cells('A2:I2')
info = ws['A2']
info.value = "学校：______________    学期：2025-2026 学年第二学期    制表日期：2026 年 4 月"
info.font = Font(name='Arial', size=10, italic=True)
info.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 25

# ========== 课时汇总表 ==========
ws.merge_cells('A3:I3')
summary_title = ws['A3']
summary_title.value = "课时汇总统计"
summary_title.font = Font(name='Arial', size=12, bold=True, color='2E75B6')
summary_title.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[3].height = 25

# 汇总表头
headers = ["班级", "语文", "数学", "英语", "科学", "社会", "周课时", "备注"]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border_thin
    set_center(cell)
ws.row_dimensions[4].height = 28

# 汇总数据
classes = ['901', '902', '903', '904', '905', '906', '907', '908']
for row, cls in enumerate(classes, start=5):
    ws.cell(row=row, column=1, value=cls).font = normal_font
    ws.cell(row=row, column=2, value=5).font = normal_font
    ws.cell(row=row, column=3, value=8).font = normal_font
    ws.cell(row=row, column=4, value=6).font = normal_font
    ws.cell(row=row, column=5, value=6).font = normal_font
    ws.cell(row=row, column=6, value=5).font = normal_font
    ws.cell(row=row, column=7, value=30).font = Font(name='Arial', size=10, bold=True)
    
    for col in range(1, 9):
        cell = ws.cell(row=row, column=col)
        cell.border = border_thin
        set_center(cell)
        if row % 2 == 1:
            cell.fill = alt_fill
    
    ws.row_dimensions[row].height = 25

# 合计行
ws.cell(row=13, column=1, value="合计").font = Font(name='Arial', size=11, bold=True)
ws.cell(row=13, column=3, value=40).font = Font(name='Arial', size=11, bold=True, color='C0392B')
ws.cell(row=13, column=4, value=64).font = Font(name='Arial', size=11, bold=True, color='C0392B')
ws.cell(row=13, column=5, value=48).font = Font(name='Arial', size=11, bold=True, color='C0392B')
ws.cell(row=13, column=6, value=48).font = Font(name='Arial', size=11, bold=True, color='C0392B')
ws.cell(row=13, column=7, value=40).font = Font(name='Arial', size=11, bold=True, color='C0392B')
ws.cell(row=13, column=8, value=240).font = Font(name='Arial', size=11, bold=True, color='C0392B')

for col in range(1, 9):
    cell = ws.cell(row=13, column=col)
    cell.border = border_thin
    set_center(cell)
ws.row_dimensions[13].height = 28

# ========== 详细课程表 - 周一 ==========
ws.merge_cells('A14:I14')
detail_title = ws['A14']
detail_title.value = "详细课程表"
detail_title.font = Font(name='Arial', size=12, bold=True, color='2E75B6')
detail_title.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[14].height = 25

# 课程数据
course_data = {
    "周一": [
        ["第 1 节", "语文", "语文", "英语", "英语", "数学", "数学", "科学", "科学"],
        ["第 2 节", "数学", "数学", "数学", "数学", "英语", "英语", "语文", "语文"],
        ["第 3 节", "英语", "英语", "科学", "科学", "语文", "语文", "数学", "数学"],
        ["第 4 节", "科学", "科学", "社会", "社会", "科学", "科学", "英语", "英语"],
        ["第 5 节", "社会", "社会", "语文", "语文", "社会", "社会", "科学", "科学"],
        ["第 6 节", "自习", "自习", "班会", "班会", "体育", "体育", "自习", "自习"],
    ],
    "周二": [
        ["第 1 节", "数学", "数学", "英语", "英语", "语文", "语文", "科学", "科学"],
        ["第 2 节", "英语", "英语", "语文", "语文", "数学", "数学", "社会", "社会"],
        ["第 3 节", "语文", "语文", "数学", "数学", "英语", "英语", "数学", "数学"],
        ["第 4 节", "科学", "科学", "社会", "社会", "科学", "科学", "语文", "语文"],
        ["第 5 节", "社会", "社会", "科学", "科学", "社会", "社会", "英语", "英语"],
        ["第 6 节", "体育", "体育", "自习", "自习", "班会", "班会", "自习", "自习"],
    ],
    "周三": [
        ["第 1 节", "语文", "语文", "数学", "数学", "科学", "科学", "英语", "英语"],
        ["第 2 节", "数学", "数学", "英语", "英语", "语文", "语文", "科学", "科学"],
        ["第 3 节", "英语", "英语", "科学", "科学", "数学", "数学", "语文", "语文"],
        ["第 4 节", "科学", "科学", "社会", "社会", "英语", "英语", "数学", "数学"],
        ["第 5 节", "社会", "社会", "语文", "语文", "社会", "社会", "科学", "科学"],
        ["第 6 节", "班会", "班会", "体育", "体育", "自习", "自习", "自习", "自习"],
    ],
    "周四": [
        ["第 1 节", "英语", "英语", "语文", "语文", "数学", "数学", "科学", "科学"],
        ["第 2 节", "数学", "数学", "科学", "科学", "英语", "英语", "语文", "语文"],
        ["第 3 节", "语文", "语文", "数学", "数学", "科学", "科学", "数学", "数学"],
        ["第 4 节", "社会", "社会", "英语", "英语", "语文", "语文", "科学", "科学"],
        ["第 5 节", "科学", "科学", "社会", "社会", "社会", "社会", "英语", "英语"],
        ["第 6 节", "自习", "自习", "自习", "自习", "体育", "体育", "班会", "班会"],
    ],
    "周五": [
        ["第 1 节", "数学", "数学", "语文", "语文", "英语", "英语", "社会", "社会"],
        ["第 2 节", "语文", "语文", "英语", "英语", "科学", "科学", "数学", "数学"],
        ["第 3 节", "英语", "英语", "数学", "数学", "语文", "语文", "科学", "科学"],
        ["第 4 节", "科学", "科学", "社会", "社会", "数学", "数学", "语文", "语文"],
        ["第 5 节", "社会", "社会", "科学", "科学", "社会", "社会", "英语", "英语"],
        ["第 6 节", "语文", "语文", "语文", "语文", "班会", "班会", "自习", "自习"],
    ],
}

row = 15
# 表头
ws.cell(row=row, column=1, value="班级/节次").font = header_font
ws.cell(row=row, column=1).fill = header_fill
ws.cell(row=row, column=1).border = border_thin
set_center(ws.cell(row=row, column=1))

classes_header = ["901 班", "902 班", "903 班", "904 班", "905 班", "906 班", "907 班", "908 班"]
for col, cls in enumerate(classes_header, start=2):
    cell = ws.cell(row=row, column=col, value=cls)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border_thin
    set_center(cell)
ws.row_dimensions[row].height = 28

# 填充每天的数据
for day, data in course_data.items():
    row += 1
    ws.cell(row=row, column=1, value=f"  {day}").font = Font(name='Arial', size=11, bold=True, color='2E75B6')
    ws.row_dimensions[row].height = 28
    
    for lesson in data:
        row += 1
        for col, value in enumerate(lesson, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.border = border_thin
            set_center(cell)
            if row % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[row].height = 25

# 保存文件
wb.save('/home/admin/.openclaw/workspace/course-schedule/九年级课程总表.xlsx')
print("✅ Excel 文件已生成：九年级课程总表.xlsx")
