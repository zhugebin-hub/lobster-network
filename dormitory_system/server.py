#!/usr/bin/env python3
"""
新生选寝系统 - 业务小龙虾版
兼容 Python 3.12+，使用 email.parser 替代已废弃的 cgi 模块
"""

import os
import sys
import json
import csv
import re
import io
import uuid
import shutil
import hmac
import hashlib
import time
from pathlib import Path
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from email.parser import BytesParser
from email.policy import default as default_policy

# 第三方依赖
import openpyxl

# ==================== 路径配置 ====================
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
VERSIONS_DIR = BASE_DIR / "versions"
STATIC_DIR = BASE_DIR / "static"

for d in [DATA_DIR, VERSIONS_DIR, STATIC_DIR]:
    d.mkdir(parents=True, exist_ok=True)

SCHOOL_CITY_MAP = DATA_DIR / "school_city_map.csv"

# ==================== 鉴权配置 ====================
API_TOKEN = os.environ.get("DORM_API_TOKEN", "")
ALLOWED_TOKENS = set()

# 从文件加载允许 token 列表
TOKENS_FILE = BASE_DIR / ".api_tokens"
if API_TOKEN:
    ALLOWED_TOKENS.add(API_TOKEN)
if TOKENS_FILE.exists():
    for line in TOKENS_FILE.read_text().strip().split("\n"):
        line = line.strip()
        if line and not line.startswith("#"):
            ALLOWED_TOKENS.add(line)

def check_auth(headers):
    """验证请求鉴权，返回 (is_ok, error_msg)"""
    if not ALLOWED_TOKENS:
        return True, ""  # 未配置 token 时跳过鉴权
    
    auth = headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
        if token in ALLOWED_TOKENS:
            return True, ""
    
    api_key = headers.get("X-API-Key", "")
    if api_key in ALLOWED_TOKENS:
        return True, ""
    
    return False, "鉴权失败：请提供有效的 Authorization: Bearer <token> 或 X-API-Key"

def mask_sensitive(value, visible=4):
    """脱敏：只显示最后 N 位"""
    if not value or len(value) <= visible:
        return "***"
    return "*" * (len(value) - visible) + value[-visible:]

def mask_student_for_output(student, show_sensitive=False):
    """对学生数据进行脱敏处理"""
    if show_sensitive:
        return student
    s = dict(student)
    if "phone" in s and s["phone"]:
        s["phone"] = mask_sensitive(s["phone"])
    if "remark" in s and s["remark"]:
        # 保留非敏感的备注信息
        remark = s["remark"]
        health_keywords = ["过敏", "哮喘", "癫痫", "心脏病", "糖尿病", "高血压", "残疾", "行动不便", "清真", "素食"]
        has_health = any(kw in remark for kw in health_keywords)
        if has_health:
            s["remark"] = "[敏感信息已隐藏]"
    return s

# ==================== 内置城市映射 ====================
SCHOOL_CITY_HINTS = {
    "浙江工商大学": "杭州", "浙商大": "杭州",
    "浙江大学": "杭州", "浙大": "杭州",
    "杭州电子科技大学": "杭州", "杭州师大": "杭州",
    "浙江理工大学": "杭州", "浙江科技大学": "杭州",
    "宁波大学": "宁波", "宁波诺丁汉": "宁波",
    "温州医科大学": "温州", "温州大学": "温州",
    "浙江工业大学": "杭州", "浙江财经大学": "杭州",
    "浙江农林大学": "杭州",
    "中国计量大学": "杭州",
    "浙江中医药大学": "杭州",
    "绍兴文理学院": "绍兴",
    "嘉兴学院": "嘉兴", "嘉兴大学": "嘉兴",
    "湖州师范学院": "湖州",
    "台州学院": "台州",
    "丽水学院": "丽水",
    "衢州学院": "衢州",
    "浙江万里学院": "宁波",
    "浙江大学城市学院": "杭州",
    "同济大学": "上海", "复旦大学": "上海", "上海交通大学": "上海",
    "北京大学": "北京", "清华大学": "北京",
    "南京大学": "南京", "东南大学": "南京",
}

# ==================== 字段别名 ====================
FIELD_ALIASES = {
    "id": ["学号", "学生编号", "考生编号", "录取编号", "id", "studentid", "student_id"],
    "name": ["姓名", "名字", "学生姓名", "name"],
    "gender": ["性别", "男女", "gender"],
    "phone": ["钉钉手机号", "手机", "手机号", "联系电话", "电话"],
    "origin": ["户口地", "户籍地", "生源地", "生源省份", "籍贯", "家乡", "省份"],
    "undergrad_school": ["本科毕业单位", "本科院校", "本科学校"],
    "undergrad_city": ["本科学校所在地"],
    "intent": ["协同强意向", "意向室友", "希望同住", "指定室友"],
    "smoke": ["是否抽烟", "抽烟", "吸烟"],
    "schedule": ["作息时间", "作息"],
    "game": ["打游戏", "游戏耐受度", "游戏"],
    "noise": ["噪音"],
    "hygiene": ["卫生"],
    "ac": ["空调温度", "空调"],
    "nap": ["午睡"],
    "social": ["社交"],
    "diet": ["饮食"],
    "remark": ["备注", "特殊情况", "住宿需求", "补充说明"],
    "is_local": ["是否本地", "本地"],
    "is_home_school": ["是否本校", "本校"],
}

# ==================== 城市解析 ====================
DIRECT_CITIES = {"北京", "上海", "天津", "重庆", "香港", "澳门"}

def load_school_city_map():
    """加载学校-城市映射 CSV"""
    mapping = {}
    if SCHOOL_CITY_MAP.exists():
        with open(SCHOOL_CITY_MAP, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                school = row.get("学校", row.get("school", "")).strip()
                city = row.get("城市", row.get("city", "")).strip()
                if school and city:
                    mapping[school] = city
    mapping.update(SCHOOL_CITY_HINTS)
    return mapping

def extract_city(text):
    """从文本中提取城市名"""
    if not text:
        return ""
    text = str(text).strip()
    if text in DIRECT_CITIES:
        return text
    city_keywords = [
        "北京", "上海", "天津", "重庆", "广州", "深圳", "杭州", "宁波",
        "温州", "绍兴", "嘉兴", "湖州", "金华", "台州", "丽水", "衢州",
        "舟山", "南京", "苏州", "无锡", "常州", "南通", "镇江", "扬州",
        "合肥", "武汉", "成都", "西安", "长沙", "郑州", "济南", "青岛",
        "福州", "厦门", "南昌", "贵阳", "昆明", "兰州", "哈尔滨", "长春",
        "沈阳", "石家庄", "太原", "呼和浩特", "乌鲁木齐", "拉萨", "西宁",
        "银川", "南宁", "海口",
    ]
    for city in city_keywords:
        if city in text:
            return city
    # 去掉省/市/自治区后缀
    for suffix in ["省", "市", "自治区", "地区", "自治州"]:
        text = text.replace(suffix, "")
    return text.strip() if text.strip() else ""

# ==================== Excel/CSV 解析 ====================
def read_excel_file(file_content, file_name=None):
    """读取 Excel 文件，返回列名列表和行数据"""
    wb = openpyxl.load_bookmark(io.BytesIO(file_content)) if isinstance(file_content, bytes) else openpyxl.load_bookmark(file_content)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows

def read_excel_safe(file_content):
    """安全读取 Excel"""
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    return rows

def read_csv_file(file_content, encoding="utf-8-sig"):
    """读取 CSV/TSV 文件"""
    text = file_content.decode(encoding)
    # 自动检测分隔符
    first_line = text.split("\n")[0]
    if "\t" in first_line:
        delimiter = "\t"
    elif ";" in first_line:
        delimiter = ";"
    else:
        delimiter = ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [row for row in reader]
    return rows

def find_header_row(rows, max_search=10):
    """在前 N 行中寻找包含学号或姓名的表头行"""
    for i in range(min(max_search, len(rows))):
        row = rows[i]
        text = " ".join(str(c) for c in row)
        if any(kw in text for kw in ["学号", "学生编号", "id", "姓名", "学生姓名"]):
            return i
    return 0

def normalize_headers(header_row):
    """将表头映射到标准字段名"""
    mapping = {}
    for i, h in enumerate(header_row):
        h_str = str(h).strip()
        for field, aliases in FIELD_ALIASES.items():
            if h_str in aliases or any(a in h_str for a in aliases) or h_str in [a.lower() for a in aliases]:
                if field not in mapping:  # 第一个匹配的优先
                    mapping[field] = i
    return mapping

def parse_students(rows, header_row_idx=0):
    """解析学生数据"""
    if header_row_idx >= len(rows):
        return [], []
    
    header_map = normalize_headers(rows[header_row_idx])
    students = []
    
    for i in range(header_row_idx + 1, len(rows)):
        row = rows[i]
        if not any(str(c).strip() for c in row):
            continue  # 跳过空行
        
        student = {}
        for field, col_idx in header_map.items():
            student[field] = str(row[col_idx]).strip() if col_idx < len(row) else ""
        
        # 至少要有姓名或学号
        if student.get("name") or student.get("id"):
            students.append(student)
    
    return students, header_map

def try_read_file(file_content, file_name=""):
    """尝试用多种方式读取文件"""
    # 尝试 Excel
    try:
        rows = read_excel_safe(file_content)
        if rows and len(rows) > 1:
            return rows
    except Exception:
        pass
    
    # 尝试 CSV/TSV 各种编码
    for encoding in ["utf-8-sig", "utf-8", "gb18030", "gbk"]:
        try:
            rows = read_csv_file(file_content, encoding)
            if rows and len(rows) > 1:
                return rows
        except Exception:
            continue
    
    raise ValueError(f"无法解析文件: {file_name}")

# ==================== 学生画像归一化 ====================
def normalize_gender(text):
    if not text:
        return "未填"
    t = text.lower().strip()
    if "女" in t or "f" in t or "female" in t:
        return "女"
    if "男" in t or "m" in t or "male" in t:
        return "男"
    return text.strip()

def normalize_boolean(text, positive_words=None):
    if not text:
        return "未填"
    t = text.strip()
    if positive_words:
        if any(w in t for w in positive_words):
            return "是"
        if any(w in t for w in ["否", "不", "无"]):
            return "否"
    return t

def extract_remark_tags(remark):
    """从备注中提取语义标签"""
    tags = []
    if not remark:
        return tags
    
    r = remark.lower()
    
    # 健康相关
    health_keywords = ["过敏", "哮喘", "癫痫", "心脏病", "糖尿病", "高血压", "残疾", "行动不便"]
    if any(k in r for k in health_keywords):
        tags.append("健康需求")
    
    # 饮食
    diet_keywords = ["清真", "素食", "不吃", "忌口", "过敏"]
    if any(k in r for k in diet_keywords):
        tags.append("饮食特殊")
    
    # 强意向
    intent_keywords = ["想和", "希望和", "指定", "要求同住", "一起住", "同住"]
    if any(k in r for k in intent_keywords):
        tags.append("强意向")
    
    # 回避
    avoid_keywords = ["不想和", "不希望", "避开", "避免", "不能和", "不同住"]
    if any(k in r for k in avoid_keywords):
        tags.append("回避同住")
    
    # 安静需求
    quiet_keywords = ["安静", "怕吵", "需要安静"]
    if any(k in r for k in quiet_keywords):
        tags.append("需要安静")
    
    # 朝向
    direction_keywords = ["朝南", "朝北", "朝阳", "背阴"]
    if any(k in r for k in direction_keywords):
        tags.append("朝向需求")
    
    return tags

def normalize_student(raw):
    """将原始学生数据归一化为统一画像"""
    s = {}
    
    # 学号
    s["id"] = raw.get("id", "").strip()
    if not s["id"]:
        s["id"] = f"NOID-{raw.get('_row_idx', '?')}"
    
    # 姓名
    s["name"] = raw.get("name", "未知").strip() or "未知"
    
    # 性别
    s["gender"] = normalize_gender(raw.get("gender", ""))
    
    # 手机
    s["phone"] = raw.get("phone", "").strip()
    
    # 生源地
    s["origin"] = raw.get("origin", "").strip()
    s["origin_city"] = extract_city(s["origin"])
    
    # 本科院校/城市
    s["undergrad_school"] = raw.get("undergrad_school", "").strip()
    school_city_map = load_school_city_map()
    undergrad_city = raw.get("undergrad_city", "").strip()
    if undergrad_city:
        s["undergrad_city"] = undergrad_city
    elif s["undergrad_school"] in school_city_map:
        s["undergrad_city"] = school_city_map[s["undergrad_school"]]
    else:
        s["undergrad_city"] = extract_city(s["undergrad_school"])
    
    # 本校/本地判断
    s["is_home_school"] = raw.get("is_home_school", "").strip()
    if not s["is_home_school"]:
        if s["undergrad_school"] and any(k in s["undergrad_school"] for k in ["浙江工商大学", "浙商大"]):
            s["is_home_school"] = "是"
    s["is_local"] = raw.get("is_local", "").strip()
    if not s["is_local"]:
        if s["origin"] and any(k in s["origin"] for k in ["浙江", "杭州"]):
            s["is_local"] = "是"
    
    # 强意向
    s["intent"] = raw.get("intent", "").strip()
    
    # 备注 & 标签
    s["remark"] = raw.get("remark", "").strip()
    s["remark_tags"] = extract_remark_tags(s["remark"])
    
    # 生活偏好
    smoke_text = raw.get("smoke", "").strip()
    if "不抽" in smoke_text or "否" in smoke_text:
        s["smoke"] = "不抽"
    elif "抽" in smoke_text:
        s["smoke"] = "抽烟"
    else:
        s["smoke"] = "不抽"  # 默认
    
    schedule_text = raw.get("schedule", "").strip()
    if any(w in schedule_text for w in ["早睡", "22", "21", "23"]):
        s["schedule"] = "早睡"
    elif any(w in schedule_text for w in ["晚睡", "熬夜", "1", "2"]):
        s["schedule"] = "晚睡"
    else:
        s["schedule"] = "23:00左右"  # 默认
    
    game_text = raw.get("game", "").strip()
    if any(w in game_text for w in ["不打", "偶尔", "少"]):
        s["game_freq"] = "偶尔"
    elif any(w in game_text for w in ["经常", "高频", "每天"]):
        s["game_freq"] = "高频"
    else:
        s["game_freq"] = "偶尔"  # 默认
    
    noise_text = raw.get("noise", "").strip()
    if "敏感" in noise_text or "不能" in noise_text:
        s["noise_sensitive"] = True
    else:
        s["noise_sensitive"] = False
    
    s["hygiene"] = raw.get("hygiene", "一般").strip() or "一般"
    s["ac"] = raw.get("ac", "无特殊").strip() or "无特殊"
    s["nap"] = raw.get("nap", "偶尔").strip() or "偶尔"
    s["social"] = raw.get("social", "一般").strip() or "一般"
    s["diet"] = raw.get("diet", "无特殊").strip() or "无特殊"
    
    # 标记是否漏填问卷
    s["missing_survey"] = raw.get("_missing_survey", False)
    
    return s

# ==================== 合并两份数据 ====================
def merge_students(official_students, survey_students):
    """以学号为主键合并官方名单和问卷数据"""
    merged = {}
    warnings = []
    
    # 先载入官方名单
    for i, s in enumerate(official_students):
        s["_row_idx"] = i
        sid = s.get("id", "").strip()
        name = s.get("name", "").strip()
        key = sid if sid else name
        if key:
            merged[key] = {**s, "_from_official": True}
    
    # 合并问卷数据
    survey_matched = 0
    survey_unmatched = []
    for s in survey_students:
        sid = s.get("id", "").strip()
        name = s.get("name", "").strip()
        key = sid if sid else name
        
        if key and key in merged:
            # 合并问卷数据到官方记录
            for k, v in s.items():
                if k not in merged[key] or not merged[key].get(k):
                    merged[key][k] = v
            survey_matched += 1
        elif key:
            survey_unmatched.append(s)
    
    if survey_unmatched:
        names = [s.get("name", s.get("id", "?")) for s in survey_unmatched[:5]]
        warnings.append(f"问卷中有 {len(survey_unmatched)} 人未在官方名单中匹配：{', '.join(names)}...")
    
    # 漏填问卷的学生使用默认画像
    missing_count = 0
    for key, s in merged.items():
        if not s.get("smoke") and not s.get("schedule"):
            s["_missing_survey"] = True
            # 设置默认值
            s.setdefault("smoke", "不抽")
            s.setdefault("schedule", "23:00左右")
            s.setdefault("game", "偶尔")
            s.setdefault("noise", "一般")
            s.setdefault("hygiene", "一般")
            s.setdefault("ac", "无特殊")
            s.setdefault("nap", "偶尔")
            s.setdefault("social", "一般")
            s.setdefault("diet", "无特殊")
            missing_count += 1
    
    if missing_count:
        warnings.append(f"{missing_count} 人未填写问卷，使用默认画像（卡片将标'漏报'）")
    
    return list(merged.values()), warnings

# ==================== 匹配算法 ====================
class UnionFind:
    """并查集 - 用于强意向绑定"""
    def __init__(self):
        self.parent = {}
    
    def find(self, x):
        if x not in self.parent:
            self.parent[x] = x
        if self.parent[x] != x:
            self.parent[x] = self.find(self.parent[x])
        return self.parent[x]
    
    def union(self, x, y):
        px, py = self.find(x), self.find(y)
        if px != py:
            self.parent[px] = py

def extract_intent_ids(student):
    """从学生的意向字段中提取学号/姓名列表"""
    ids = []
    intent = student.get("intent", "") + " " + student.get("remark", "")
    if not intent:
        return ids
    
    # 查找学号模式
    id_patterns = re.findall(r'\d{8,12}', intent)
    ids.extend(id_patterns)
    
    # 查找姓名（中文，2-4字）
    name_patterns = re.findall(r'[\u4e00-\u9fff]{2,4}', intent)
    ids.extend(name_patterns)
    
    return list(set(ids))

def extract_avoid_ids(student):
    """从备注中提取回避的学号/姓名"""
    ids = []
    remark = student.get("remark", "")
    if not remark:
        return ids
    
    avoid_keywords = ["不想和", "不希望", "避开", "避免", "不能和", "不同住", "不要和"]
    has_avoid = any(kw in remark for kw in avoid_keywords)
    if not has_avoid:
        return ids
    
    id_patterns = re.findall(r'\d{8,12}', remark)
    ids.extend(id_patterns)
    
    # 在回避关键词后面找姓名
    for kw in avoid_keywords:
        idx = remark.find(kw)
        if idx >= 0:
            after = remark[idx + len(kw):]
            names = re.findall(r'[\u4e00-\u9fff]{2,4}', after[:10])
            ids.extend(names)
    
    return list(set(ids))

def calc_preference_score(s1, s2):
    """计算两个学生之间的偏好差异分"""
    score = 0
    
    # 作息差异 * 100
    schedule_order = {"早睡": 0, "23:00左右": 1, "晚睡": 2}
    sch1 = schedule_order.get(s1.get("schedule", ""), 1)
    sch2 = schedule_order.get(s2.get("schedule", ""), 1)
    score += abs(sch1 - sch2) * 100
    
    # 游戏差异 * 18
    game_order = {"不打": 0, "偶尔": 1, "高频": 2}
    g1 = game_order.get(s1.get("game_freq", "偶尔"), 1)
    g2 = game_order.get(s2.get("game_freq", "偶尔"), 1)
    score += abs(g1 - g2) * 18
    
    # 游戏高频 + 噪音敏感 +45
    if g1 >= 2 and s2.get("noise_sensitive") or g2 >= 2 and s1.get("noise_sensitive"):
        score += 45
    
    # 噪音差异 * 16
    score += abs(s1.get("noise_sensitive", False) - s2.get("noise_sensitive", False)) * 16
    
    # 卫生差异 +10
    if s1.get("hygiene", "") != s2.get("hygiene", ""):
        score += 10
    
    # 空调差异 * 8
    if s1.get("ac", "") != s2.get("ac", ""):
        score += 8
    
    # 午睡/社交差异各 +8
    if s1.get("nap", "") != s2.get("nap", ""):
        score += 8
    if s1.get("social", "") != s2.get("social", ""):
        score += 8
    
    return score

def calc_bond_score(s1, s2):
    """计算纽带加分（负分=加分）"""
    score = 0
    
    # 意向同住 -5000
    if s1.get("id") in extract_intent_ids(s2) or s2.get("id") in extract_intent_ids(s1):
        score -= 5000
    
    # 本科同校 -14
    if s1.get("undergrad_school") and s1["undergrad_school"] == s2.get("undergrad_school"):
        score -= 14
    
    # 本科同城 -10
    if s1.get("undergrad_city") and s1["undergrad_city"] == s2.get("undergrad_city"):
        score -= 10
    
    # 生源同乡 -5
    if s1.get("origin_city") and s1["origin_city"] == s2.get("origin_city"):
        score -= 5
    
    # 无本校/本地带动 +20
    if s1.get("is_home_school") != "是" and s2.get("is_home_school") != "是":
        if s1.get("is_local") != "是" and s2.get("is_local") != "是":
            score += 20
    
    return score

def check_hard_conflict(s1, s2):
    """检查硬冲突"""
    conflicts = []
    
    # 性别不一致
    if s1["gender"] != s2["gender"]:
        conflicts.append("性别不一致")
    
    # 抽烟冲突
    smoke1 = s1.get("smoke", "不抽")
    smoke2 = s2.get("smoke", "不抽")
    if (smoke1 == "抽烟" and smoke2 == "抽烟") or \
       (smoke1 == "抽烟" and "极度介意" in str(s2.get("remark", ""))) or \
       (smoke2 == "抽烟" and "极度介意" in str(s1.get("remark", ""))):
        conflicts.append("抽烟底线冲突")
    
    # 极端作息冲突
    if s1.get("schedule") == "早睡" and s2.get("schedule") == "晚睡":
        conflicts.append("极端作息冲突")
    if s1.get("schedule") == "晚睡" and s2.get("schedule") == "早睡":
        conflicts.append("极端作息冲突")
    
    return conflicts

def check_avoid_conflict(s1, s2):
    """检查回避同住"""
    avoid1 = extract_avoid_ids(s1)
    avoid2 = extract_avoid_ids(s2)
    
    if s2.get("id") in avoid1 or s2.get("name") in avoid1:
        return True
    if s1.get("id") in avoid2 or s1.get("name") in avoid2:
        return True
    return False

def match_dormitories(students, room_size=4):
    """核心匹配算法：启发式分组 + 贪心装箱"""
    warnings = []
    advice = []
    
    # 按性别分池
    male_pool = [s for s in students if s["gender"] == "男"]
    female_pool = [s for s in students if s["gender"] == "女"]
    other_pool = [s for s in students if s["gender"] not in ["男", "女"]]
    
    # 处理强意向绑定
    def process_intent_groups(pool):
        uf = UnionFind()
        id_map = {s["id"]: s for s in pool}
        name_map = {s["name"]: s for s in pool}
        
        for s in pool:
            intents = extract_intent_ids(s)
            for target in intents:
                if target in id_map:
                    uf.union(s["id"], id_map[target]["id"])
                elif target in name_map:
                    uf.union(s["id"], name_map[target]["id"])
        
        groups = {}
        for s in pool:
            root = uf.find(s["id"])
            if root not in groups:
                groups[root] = []
            groups[root].append(s)
        
        return groups, uf
    
    # 构建意向组
    male_groups, _ = process_intent_groups(male_pool)
    female_groups, _ = process_intent_groups(female_pool)
    
    def assign_rooms(groups, pool, room_num_start):
        rooms = []
        suspended = []
        room_counter = room_num_start
        
        # 按意向组大小排序（大的优先）
        sorted_groups = sorted(groups.values(), key=len, reverse=True)
        
        # 收集所有待分配学生
        unassigned = []
        for group in sorted_groups:
            if len(group) <= room_size:
                unassigned.append(group)
            else:
                # 超大意向组拆分
                for i in range(0, len(group), room_size):
                    unassigned.append(group[i:i + room_size])
        
        # 贪心装箱
        for group in unassigned:
            if not group:
                continue
            
            # 尝试放入现有未满房间
            best_room_idx = None
            best_score = float("inf")
            
            for ri, room in enumerate(rooms):
                if len(room) >= room_size:
                    continue
                
                # 检查是否有硬冲突
                has_conflict = False
                for existing in room:
                    for new_s in group:
                        if check_hard_conflict(existing, new_s):
                            has_conflict = True
                            break
                        if check_avoid_conflict(existing, new_s):
                            has_conflict = True
                            break
                    if has_conflict:
                        break
                
                if has_conflict:
                    continue
                
                # 计算增量分数
                total_delta = 0
                for new_s in group:
                    for existing in room:
                        total_delta += calc_preference_score(existing, new_s)
                        total_delta += calc_bond_score(existing, new_s)
                
                if total_delta < best_score:
                    best_score = total_delta
                    best_room_idx = ri
            
            if best_room_idx is not None:
                rooms[best_room_idx].extend(group)
            else:
                # 开新房
                if len(group) <= room_size:
                    rooms.append(list(group))
                else:
                    rooms.append(list(group[:room_size]))
                    for s in group[room_size:]:
                        unassigned.append([s])
        
        # 余数处理：所有房间都有效，不强制挂起
        # 挂起池用于人工调整时临时移出
        suspended = []
        
        final_rooms = [r for r in rooms if len(r) > 0]
        
        # 编号
        for i, room in enumerate(final_rooms):
            room_num = room_num_start + i
            for s in room:
                s["_room_id"] = f"{room_num}"
                s["_room_num"] = room_num
        
        return final_rooms, suspended
    
    male_rooms, male_suspended = assign_rooms(male_groups, male_pool, 101)
    female_rooms, female_suspended = assign_rooms(female_groups, female_pool, 201)
    other_rooms, other_suspended = assign_rooms({s["id"]: [s] for s in other_pool}, other_pool, 301)
    
    all_rooms = male_rooms + female_rooms + other_rooms
    all_suspended = male_suspended + female_suspended + other_suspended
    
    # 检查冲突并标记 - 将元数据附加到房间列表上
    room_metadata = {}
    for room in all_rooms:
        room_id = room[0].get("_room_id", f"room_{id(room)}")
        room_conflicts = []
        room_bonds = []
        score = 0
        for i in range(len(room)):
            for j in range(i + 1, len(room)):
                conflicts = check_hard_conflict(room[i], room[j])
                room_conflicts.extend(conflicts)
                
                bond = calc_bond_score(room[i], room[j])
                if bond < 0:
                    room_bonds.append(f"{room[i]['name']}-{room[j]['name']}: 纽带+{abs(bond)}")
                
                if check_avoid_conflict(room[i], room[j]):
                    room_conflicts.append(f"{room[i]['name']}/{room[j]['name']}: 回避同住")
                
                score += calc_preference_score(room[i], room[j]) + calc_bond_score(room[i], room[j])
        
        room_metadata[room_id] = {
            "_conflicts": list(set(room_conflicts)),
            "_bonds": room_bonds,
            "_score": score,
        }
        # 同时把元数据附加到每个学生身上，方便查找
        for s in room:
            s["_room_conflicts"] = list(set(room_conflicts))
            s["_room_bonds"] = room_bonds
            s["_room_score"] = score
    
    # 统计
    total = len(students)
    room_count = len(all_rooms)
    suspended_count = len(all_suspended)
    conflict_count = len([r for r in all_rooms if r and r[0].get("_room_conflicts")])
    
    summary = {
        "total_students": total,
        "room_count": room_count,
        "suspended_count": suspended_count,
        "conflict_count": conflict_count,
        "room_size": room_size,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    
    if conflict_count > 0:
        warnings.append(f"⚠️ {conflict_count} 个寝室存在硬冲突，需人工复核")
    
    if suspended_count > 0:
        advice.append(f"💡 {suspended_count} 人在混寝挂起池，建议按同校/同城/同乡捆绑外调")
    
    return {
        "students": students,
        "rooms": all_rooms,
        "suspended": all_suspended,
        "warnings": warnings,
        "advice": advice,
        "summary": summary,
    }

# ==================== Excel 导出 ====================
def export_to_excel(plan_data):
    """导出寝室分配方案为 Excel"""
    wb = openpyxl.Workbook()
    
    rooms = plan_data["rooms"]
    suspended = plan_data["suspended"]
    students = plan_data["students"]
    room_size = plan_data.get("room_size", 4)
    
    # 工作表1：寝室分配结果
    ws1 = wb.active
    ws1.title = "寝室分配结果"
    headers1 = ["宿舍", "床位", "学号", "姓名", "性别", "生源地", "本科院校",
                "本科所在地", "抽烟", "作息", "游戏", "噪音", "空调", "午睡",
                "社交", "风险/纽带", "备注"]
    ws1.append(headers1)
    
    for room in rooms:
        room_id = room[0].get("_room_id", "?") if room else "?"
        conflicts = "; ".join(room[0].get("_room_conflicts", [])) if room else ""
        bonds = "; ".join(room[0].get("_room_bonds", [])) if room else ""
        risk_bond = ""
        if conflicts:
            risk_bond = f"⚠️{conflicts}"
        if bonds:
            risk_bond += f" | {bonds}" if risk_bond else bonds
        
        for bed, s in enumerate(room, 1):
            ws1.append([
                room_id, bed, s.get("id", ""), s.get("name", ""),
                s.get("gender", ""), s.get("origin", ""),
                s.get("undergrad_school", ""), s.get("undergrad_city", ""),
                s.get("smoke", ""), s.get("schedule", ""),
                s.get("game_freq", ""), "敏感" if s.get("noise_sensitive") else "一般",
                s.get("ac", ""), s.get("nap", ""), s.get("social", ""),
                risk_bond, s.get("remark", ""),
            ])
    
    # 工作表2：混寝挂起池
    ws2 = wb.create_sheet("混寝挂起池")
    headers2 = ["学号", "姓名", "性别", "生源地", "本科院校", "本科所在地",
                "生活标签", "风险/纽带", "备注"]
    ws2.append(headers2)
    
    for s in suspended:
        tags = []
        if s.get("missing_survey"):
            tags.append("漏报")
        if s.get("smoke") == "抽烟":
            tags.append("抽烟")
        tags_str = "/".join(tags) if tags else "无特殊"
        
        ws2.append([
            s.get("id", ""), s.get("name", ""), s.get("gender", ""),
            s.get("origin", ""), s.get("undergrad_school", ""),
            s.get("undergrad_city", ""), tags_str,
            "", "混寝/外调优先人工复核",
        ])
    
    # 工作表3：学生画像
    ws3 = wb.create_sheet("学生画像")
    headers3 = ["学号", "姓名", "性别", "手机", "生源地", "是否本校", "是否本地",
                "本科院校", "本科所在地", "备注意向室友", "备注回避室友",
                "宿舍要求", "完整偏好", "备注语义标签", "备注"]
    ws3.append(headers3)
    
    for s in students:
        intent_ids = extract_intent_ids(s)
        avoid_ids = extract_avoid_ids(s)
        tags = extract_remark_tags(s.get("remark", ""))
        
        prefs = (
            f"抽烟:{s.get('smoke','')} | "
            f"作息:{s.get('schedule','')} | "
            f"游戏:{s.get('game_freq','')} | "
            f"噪音:{'敏感' if s.get('noise_sensitive') else '一般'} | "
            f"卫生:{s.get('hygiene','')} | "
            f"空调:{s.get('ac','')} | "
            f"午睡:{s.get('nap','')} | "
            f"社交:{s.get('social','')} | "
            f"饮食:{s.get('diet','')}"
        )
        
        ws3.append([
            s.get("id", ""), s.get("name", ""), s.get("gender", ""),
            s.get("phone", ""), s.get("origin", ""),
            s.get("is_home_school", ""), s.get("is_local", ""),
            s.get("undergrad_school", ""), s.get("undergrad_city", ""),
            ", ".join(intent_ids) if intent_ids else "",
            ", ".join(avoid_ids) if avoid_ids else "",
            "", prefs,
            ", ".join(tags) if tags else "",
            s.get("remark", ""),
        ])
    
    # 保存到临时文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = DATA_DIR / f"dormitory_assignment_{timestamp}.xlsx"
    wb.save(str(output_path))
    
    return output_path

# ==================== 版本管理 ====================
def save_version(plan_id, version_name, plan_data):
    """保存方案版本"""
    version_id = str(uuid.uuid4())[:8]
    version_file = VERSIONS_DIR / f"{version_id}.json"
    
    version = {
        "version_id": version_id,
        "plan_id": plan_id,
        "version_name": version_name,
        "created_at": datetime.now().isoformat(),
        "data": plan_data,
    }
    
    with open(version_file, "w", encoding="utf-8") as f:
        json.dump(version, f, ensure_ascii=False, indent=2)
    
    return version_id

def list_versions(plan_id=None):
    """列出所有版本"""
    versions = []
    for vf in VERSIONS_DIR.glob("*.json"):
        with open(vf, "r", encoding="utf-8") as f:
            v = json.load(f)
        if plan_id is None or v.get("plan_id") == plan_id:
            versions.append({
                "version_id": v["version_id"],
                "version_name": v["version_name"],
                "plan_id": v["plan_id"],
                "created_at": v["created_at"],
            })
    return sorted(versions, key=lambda x: x["created_at"], reverse=True)

def restore_version(version_id):
    """恢复版本"""
    version_file = VERSIONS_DIR / f"{version_id}.json"
    if not version_file.exists():
        return None
    with open(version_file, "r", encoding="utf-8") as f:
        return json.load(f)["data"]

# ==================== Multipart 解析（替代 cgi） ====================
def parse_multipart(form_data_bytes, boundary):
    """解析 multipart/form-data 数据（替代 cgi.FieldStorage）"""
    parts = {}
    header = f"Content-Type: multipart/form-data; boundary={boundary}\r\n\r\n".encode()
    full_data = header + form_data_bytes
    
    parser = BytesParser(policy=default_policy)
    msg = parser.parsebytes(full_data)
    
    if msg.is_multipart():
        for part in msg.iter_parts():
            name = part.get_param("name", header="Content-Disposition")
            if name:
                filename = part.get_param("filename", header="Content-Disposition")
                if filename:
                    parts[name] = {
                        "filename": filename,
                        "data": part.get_payload(decode=True),
                    }
                else:
                    parts[name] = part.get_payload(decode=True).decode("utf-8")
    
    return parts

# ==================== HTTP 服务 ====================
# 内存存储当前方案
current_plans = {}

class DormHandler(BaseHTTPRequestHandler):
    """HTTP 请求处理器"""
    
    def log_message(self, format, *args):
        """简化日志"""
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {args[0]}")
    
    def send_json(self, data, status=200):
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization, X-API-Key")
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode("utf-8"))
    
    def do_OPTIONS(self):
        """CORS 预检请求"""
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization, X-API-Key")
        self.end_headers()
    
    def send_file(self, filepath, content_type="application/octet-stream"):
        with open(filepath, "rb") as f:
            content = f.read()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f"attachment; filename={filepath.name}")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)
    
    def read_body(self):
        length = int(self.headers.get("Content-Length", 0))
        return self.rfile.read(length)
    
    def do_GET(self):
        if self.path == "/api/health":
            self.send_json({"ok": True, "time": int(datetime.now().timestamp() * 1000)})
        
        elif self.path == "/api/demo":
            # 生成示例数据
            result = self._generate_demo()
            self.send_json(result)
        
        elif self.path == "/api/sample-official.xlsx":
            path = self._generate_sample_official()
            self.send_file(path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        elif self.path == "/api/sample-survey.xlsx":
            path = self._generate_sample_survey()
            self.send_file(path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        elif self.path == "/" or self.path == "/index.html":
            self._serve_static("index.html", "text/html")
        
        elif self.path == "/styles.css":
            self._serve_static("styles.css", "text/css")
        
        elif self.path == "/app.js":
            self._serve_static("app.js", "application/javascript")
        
        else:
            self.send_json({"error": "Not Found"}, 404)
    
    def do_POST(self):
        # 鉴权检查（排除 health 和 demo 等公开接口）
        public_paths = {"/api/health", "/api/demo", "/api/sample-official.xlsx", "/api/sample-survey.xlsx"}
        if self.path not in public_paths:
            ok, err = check_auth(self.headers)
            if not ok:
                self.send_json({"error": err}, 401)
                return
        
        if self.path == "/api/match":
            self._handle_match()
        elif self.path == "/api/export":
            self._handle_export()
        elif self.path == "/api/save_version":
            self._handle_save_version()
        elif self.path == "/api/list_versions":
            self._handle_list_versions()
        elif self.path == "/api/restore_version":
            self._handle_restore_version()
        elif self.path == "/api/move_student":
            self._handle_move_student()
        elif self.path == "/api/swap_students":
            self._handle_swap_students()
        elif self.path == "/api/move_to_suspended":
            self._handle_move_to_suspended()
        elif self.path == "/api/recompute_room":
            self._handle_recompute_room()
        else:
            self.send_json({"error": "Not Found"}, 404)
    
    def _serve_static(self, filename, content_type):
        filepath = STATIC_DIR / filename
        if filepath.exists():
            self.send_file(filepath, content_type)
        else:
            self.send_json({"error": f"Static file not found: {filename}"}, 404)
    
    def _handle_match(self):
        """处理匹配请求"""
        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            self.send_json({"error": "需要 multipart/form-data"}, 400)
            return
        
        # 提取 boundary
        boundary = content_type.split("boundary=")[-1].strip()
        body = self.read_body()
        
        try:
            parts = parse_multipart(body, boundary)
        except Exception as e:
            self.send_json({"error": f"解析上传数据失败: {str(e)}"}, 400)
            return
        
        # 获取参数
        official_data = None
        survey_data = None
        room_size = 4
        
        if "official" in parts:
            if isinstance(parts["official"], dict):
                official_data = parts["official"]["data"]
        if "survey" in parts:
            if isinstance(parts["survey"], dict):
                survey_data = parts["survey"]["data"]
        if "roomSize" in parts:
            try:
                room_size = int(parts["roomSize"])
            except (ValueError, TypeError):
                pass
        
        if not official_data or not survey_data:
            self.send_json({"error": "需要上传官方名单和问卷表两个文件"}, 400)
            return
        
        if room_size < 2 or room_size > 8:
            self.send_json({"error": "每寝人数必须在 2-8 之间"}, 400)
            return
        
        try:
            # 解析文件
            official_rows = try_read_file(official_data, "official")
            survey_rows = try_read_file(survey_data, "survey")
            
            official_students, _ = parse_students(official_rows, find_header_row(official_rows))
            survey_students, _ = parse_students(survey_rows, find_header_row(survey_rows))
            
            # 合并
            merged, merge_warnings = merge_students(official_students, survey_students)
            
            # 归一化
            normalized = [normalize_student(s) for s in merged]
            
            # 匹配
            result = match_dormitories(normalized, room_size)
            result["room_size"] = room_size
            result["warnings"].extend(merge_warnings)
            
            # 保存方案
            plan_id = str(uuid.uuid4())[:8]
            current_plans[plan_id] = result
            
            # 备份（可选，生产环境应关闭）
            # DATA_DIR.mkdir(parents=True, exist_ok=True)
            # (DATA_DIR / "last_import").mkdir(exist_ok=True)
            # with open(DATA_DIR / "last_import" / "official.xlsx", "wb") as f:
            #     f.write(official_data)
            # with open(DATA_DIR / "last_import" / "survey.xlsx", "wb") as f:
            #     f.write(survey_data)
            
            self.send_json({**result, "plan_id": plan_id})
            
        except Exception as e:
            self.send_json({"error": f"匹配失败: {str(e)}"}, 500)
    
    def _handle_export(self):
        """处理导出请求"""
        body = self.read_body()
        try:
            data = json.loads(body)
        except json.JSONDecodeError:
            self.send_json({"error": "JSON 解析失败"}, 400)
            return
        
        plan_id = data.get("plan_id")
        if plan_id and plan_id in current_plans:
            plan_data = current_plans[plan_id]
        else:
            plan_data = data  # 直接使用传入的数据
        
        plan_data["room_size"] = data.get("room_size", 4)
        
        try:
            output_path = export_to_excel(plan_data)
            self.send_file(output_path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            self.send_json({"error": f"导出失败: {str(e)}"}, 500)
    
    def _handle_save_version(self):
        body = self.read_body()
        data = json.loads(body)
        plan_id = data.get("plan_id", "default")
        version_name = data.get("version_name", datetime.now().strftime("%Y-%m-%d %H:%M"))
        
        plan_data = current_plans.get(plan_id, data)
        version_id = save_version(plan_id, version_name, plan_data)
        self.send_json({"version_id": version_id, "version_name": version_name})
    
    def _handle_list_versions(self):
        body = self.read_body()
        data = json.loads(body) if body else {}
        plan_id = data.get("plan_id")
        versions = list_versions(plan_id)
        self.send_json({"versions": versions})
    
    def _handle_restore_version(self):
        body = self.read_body()
        data = json.loads(body)
        version_id = data.get("version_id")
        plan_data = restore_version(version_id)
        if plan_data:
            plan_id = str(uuid.uuid4())[:8]
            current_plans[plan_id] = plan_data
            self.send_json({"ok": True, "plan_id": plan_id, "version_id": version_id})
        else:
            self.send_json({"error": "版本不存在"}, 404)
    
    def _handle_move_student(self):
        body = self.read_body()
        data = json.loads(body)
        plan_id = data.get("plan_id")
        student_key = data.get("student_key")  # 学号或姓名
        target_room_id = data.get("target_room_id")
        
        if not plan_id or plan_id not in current_plans:
            self.send_json({"error": "方案不存在"}, 404)
            return
        
        result = self._move_student_in_plan(current_plans[plan_id], student_key, target_room_id)
        self.send_json(result)
    
    def _handle_swap_students(self):
        body = self.read_body()
        data = json.loads(body)
        plan_id = data.get("plan_id")
        student_a = data.get("student_a")
        student_b = data.get("student_b")
        
        if not plan_id or plan_id not in current_plans:
            self.send_json({"error": "方案不存在"}, 404)
            return
        
        result = self._swap_students_in_plan(current_plans[plan_id], student_a, student_b)
        self.send_json(result)
    
    def _handle_move_to_suspended(self):
        body = self.read_body()
        data = json.loads(body)
        plan_id = data.get("plan_id")
        student_key = data.get("student_key")
        
        if not plan_id or plan_id not in current_plans:
            self.send_json({"error": "方案不存在"}, 404)
            return
        
        result = self._move_to_suspended_in_plan(current_plans[plan_id], student_key)
        self.send_json(result)
    
    def _handle_recompute_room(self):
        body = self.read_body()
        data = json.loads(body)
        plan_id = data.get("plan_id")
        
        if not plan_id or plan_id not in current_plans:
            self.send_json({"error": "方案不存在"}, 404)
            return
        
        plan = current_plans[plan_id]
        for room in plan["rooms"]:
            if not room:
                continue
            room_conflicts = []
            room_bonds = []
            score = 0
            for i in range(len(room)):
                for j in range(i + 1, len(room)):
                    conflicts = check_hard_conflict(room[i], room[j])
                    room_conflicts.extend(conflicts)
                    bond = calc_bond_score(room[i], room[j])
                    if bond < 0:
                        room_bonds.append(f"{room[i]['name']}-{room[j]['name']}: 纽带+{abs(bond)}")
                    if check_avoid_conflict(room[i], room[j]):
                        room_conflicts.append(f"{room[i]['name']}/{room[j]['name']}: 回避同住")
                    score += calc_preference_score(room[i], room[j]) + calc_bond_score(room[i], room[j])
            
            conflicts_set = list(set(room_conflicts))
            for s in room:
                s["_room_conflicts"] = conflicts_set
                s["_room_bonds"] = room_bonds
                s["_room_score"] = score
        
        self.send_json({"ok": True, "message": "冲突和纽带已重新计算"})
    
    def _find_student(self, plan, student_key):
        """在方案中查找学生"""
        for room in plan["rooms"]:
            for s in room:
                if s.get("id") == student_key or s.get("name") == student_key:
                    return s, room
        for s in plan["suspended"]:
            if s.get("id") == student_key or s.get("name") == student_key:
                return s, None
        return None, None
    
    def _move_student_in_plan(self, plan, student_key, target_room_id):
        student, source_room = self._find_student(plan, student_key)
        if not student:
            return {"error": f"未找到学生: {student_key}"}
        
        # 从原位置移除
        if source_room is not None:
            source_room.remove(student)
        else:
            plan["suspended"].remove(student)
        
        # 找到目标房间
        target_room = None
        for room in plan["rooms"]:
            if room[0].get("_room_id") == target_room_id:
                target_room = room
                break
        
        if target_room:
            target_room.append(student)
            student["_room_id"] = target_room_id
        else:
            # 加入挂起池
            plan["suspended"].append(student)
            student.pop("_room_id", None)
        
        return {"ok": True, "message": f"已将 {student.get('name')} 移动"}
    
    def _swap_students_in_plan(self, plan, student_a_key, student_b_key):
        sa, ra = self._find_student(plan, student_a_key)
        sb, rb = self._find_student(plan, student_b_key)
        
        if not sa or not sb:
            return {"error": "未找到一个或两个学生"}
        
        if ra is not None:
            ra.remove(sa)
        else:
            plan["suspended"].remove(sa)
        
        if rb is not None:
            rb.remove(sb)
        else:
            plan["suspended"].remove(sb)
        
        # 交换
        if rb is not None:
            rb.append(sa)
            sa["_room_id"] = rb[0].get("_room_id", "?") if rb else None
        else:
            plan["suspended"].append(sa)
            sa.pop("_room_id", None)
        
        if ra is not None:
            ra.append(sb)
            sb["_room_id"] = ra[0].get("_room_id", "?") if ra else None
        else:
            plan["suspended"].append(sb)
            sb.pop("_room_id", None)
        
        return {"ok": True, "message": f"已互换 {sa.get('name')} 和 {sb.get('name')}"}
    
    def _move_to_suspended_in_plan(self, plan, student_key):
        student, source_room = self._find_student(plan, student_key)
        if not student:
            return {"error": f"未找到学生: {student_key}"}
        
        if source_room is not None:
            source_room.remove(student)
        
        if student not in plan["suspended"]:
            plan["suspended"].append(student)
        student.pop("_room_id", None)
        
        return {"ok": True, "message": f"已将 {student.get('name')} 移入挂起池"}
    
    def _generate_demo(self):
        """生成示例数据"""
        names_male = ["张三", "李四", "王五", "赵六", "陈七", "周八", "吴九", "郑十",
                      "孙一", "钱二", "刘三", "杨四"]
        names_female = ["小红", "小明", "小华", "小丽", "小芳", "小燕", "小玲", "小敏",
                        "小婷", "小静"]
        
        cities = ["杭州", "宁波", "温州", "嘉兴", "绍兴", "金华", "台州", "上海", "北京"]
        schools = ["浙江工商大学", "浙江大学", "杭州电子科技大学", "宁波大学",
                   "浙江工业大学", "温州医科大学", "嘉兴学院", "绍兴文理学院"]
        
        students = []
        for i, name in enumerate(names_male + names_female):
            s = {
                "id": f"2026{i+1:04d}",
                "name": name,
                "gender": "男" if name in names_male else "女",
                "phone": f"138{i:08d}",
                "origin": f"{cities[i % len(cities)]}",
                "undergrad_school": schools[i % len(schools)],
                "undergrad_city": "",
                "is_local": "是" if cities[i % len(cities)] in ["杭州"] else "",
                "is_home_school": "是" if schools[i % len(schools)] == "浙江工商大学" else "",
                "smoke": "不抽" if i % 5 != 0 else "抽烟",
                "schedule": ["早睡", "23:00左右", "晚睡"][i % 3],
                "game_freq": ["不打", "偶尔", "高频"][i % 3],
                "noise_sensitive": i % 4 == 0,
                "hygiene": "一般",
                "ac": "无特殊",
                "nap": "偶尔",
                "social": "一般",
                "diet": "无特殊",
                "intent": "",
                "remark": "",
                "missing_survey": False,
            }
            # 设置几个强意向
            if i == 0:
                s["intent"] = "李四"
            elif i == 1:
                s["intent"] = "张三"
            
            students.append(s)
        
        result = match_dormitories(students, 4)
        plan_id = "demo"
        current_plans[plan_id] = result
        result["plan_id"] = plan_id
        
        return result
    
    def _generate_sample_official(self):
        """生成官方名单示例"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "官方名单"
        ws.append(["学号", "姓名", "性别", "生源地", "是否本地", "是否本校"])
        
        names = ["张三", "李四", "王五", "赵六", "陈七", "小红", "小明", "小华"]
        for i, name in enumerate(names):
            ws.append([
                f"2026{i+1:04d}", name,
                "男" if i < 5 else "女",
                ["杭州", "宁波", "温州", "嘉兴", "绍兴"][i % 5],
                "是" if i % 5 == 0 else "",
                "是" if i % 3 == 0 else "",
            ])
        
        path = DATA_DIR / "sample_official.xlsx"
        wb.save(str(path))
        return path
    
    def _generate_sample_survey(self):
        """生成问卷示例"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "问卷"
        ws.append(["学号", "姓名", "性别", "手机号", "本科院校",
                    "意向室友", "是否抽烟", "作息时间", "打游戏",
                    "噪音敏感度", "备注"])
        
        names = ["张三", "李四", "王五", "赵六", "陈七", "小红", "小明", "小华"]
        for i, name in enumerate(names):
            ws.append([
                f"2026{i+1:04d}", name,
                "男" if i < 5 else "女",
                f"138{i:08d}",
                ["浙江工商大学", "浙江大学", "杭州电子科技大学"][i % 3],
                "李四" if i == 0 else ("张三" if i == 1 else ""),
                "否" if i % 5 != 0 else "是",
                ["22:30", "23:00", "01:00"][i % 3],
                ["不打", "偶尔", "经常"][i % 3],
                "敏感" if i % 4 == 0 else "一般",
                "希望朝南" if i == 0 else "",
            ])
        
        path = DATA_DIR / "sample_survey.xlsx"
        wb.save(str(path))
        return path

# ==================== 启动 ====================
def main():
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", sys.argv[1] if len(sys.argv) > 1 else "8765"))
    
    server = HTTPServer((host, port), DormHandler)
    print(f"🦞 新生选寝系统启动: http://{host}:{port}")
    print(f"   健康检查: http://{host}:{port}/api/health")
    print(f"   示例数据: http://{host}:{port}/api/demo")
    
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n服务已停止")
        server.shutdown()

if __name__ == "__main__":
    main()
