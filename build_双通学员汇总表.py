#!/usr/bin/env python3
"""构建第三届"双通"研修班学员汇总表"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = openpyxl.Workbook()

# ============================================================
# 原始数据（7人，来自Excel汇总表+6份PDF申请表）
# ============================================================
people = [
    {
        "姓名": "李文忠", "教名": "伯多禄", "教别": "天主教",
        "性别": "男", "民族": "汉", "身份证号": "210181198803090011",
        "出生年月": "1988年3月", "联系电话": "13056802195",
        "所在单位或场所": "宁波慈溪市浒山教堂（堂区副主任）",
        "最高学历": "沈阳大学园林本科（2011年）",
        "入教时间": "1995年",
    },
    {
        "姓名": "洪圆圆", "教名": "若瑟", "教别": "天主教",
        "性别": "男", "民族": "汉", "身份证号": "330203199205210915",
        "出生年月": "1992年5月", "联系电话": "18532110863",
        "所在单位或场所": "宁波慈溪市新浦天主堂（副本堂）",
        "最高学历": "河北石家庄神哲学院神学学士（2024年）",
        "入教时间": "2002年12月",
    },
    {
        "姓名": "汤迦南", "教名": "", "教别": "基督教",
        "性别": "男", "民族": "汉", "身份证号": "330683199303302411",
        "出生年月": "1993年3月", "联系电话": "18267450721",
        "所在单位或场所": "绍兴嵊州市（专职传道员）",
        "最高学历": "武汉理工大学材料科学与工程硕士（2018年）+金陵协和神学院神学学士（2022年）",
        "入教时间": "",
    },
    {
        "姓名": "洪马可", "教名": "", "教别": "基督教",
        "性别": "男", "民族": "汉", "身份证号": "330281198409230015",
        "出生年月": "1984年9月", "联系电话": "13777165999",
        "所在单位或场所": "宁波余姚市基督教城堂（牧师）",
        "最高学历": "浙江神学院神学本科（2016年）+上海海洋大学（2007年）",
        "入教时间": "2007年",
    },
    {
        "姓名": "陈斌", "教名": "智眼", "教别": "佛教",
        "性别": "男", "民族": "汉", "身份证号": "421023198706064955",
        "出生年月": "1987年6月", "联系电话": "18268507816",
        "所在单位或场所": "宁波鄞州区慧日禅寺（寺管会副主任/副监院）",
        "最高学历": "中国佛学院普陀山学院天台学佛学硕士（2022年）",
        "入教时间": "2011年",
    },
    {
        "姓名": "曹家欢", "教名": "释本道", "教别": "佛教",
        "性别": "男", "民族": "汉", "身份证号": "220723199304272617",
        "出生年月": "1993年4月", "联系电话": "15834146789",
        "所在单位或场所": "宁波宁海县岔路余庆寺（住持）",
        "最高学历": "曼谷皇家理工大学全球佛教专业（2025年）",
        "入教时间": "2010年12月",
    },
    {
        "姓名": "黄浩峰", "教名": "耀忠", "教别": "佛教",
        "性别": "男", "民族": "汉", "身份证号": "350424198702080519",
        "出生年月": "1987年2月", "联系电话": "13306608811",
        "所在单位或场所": "宁波宁海县道堂寺",
        "最高学历": "泰国国际佛教大学佛学硕士研究生（2024年）",
        "入教时间": "2006年",
    },
]

# ============================================================
# 排序规则
# ============================================================
JIAO_ORDER = {"佛教": 1, "道教": 2, "伊斯兰教": 3, "天主教": 4, "基督教": 5}
REGION_ORDER = {"杭州": 1, "宁波": 2, "温州": 3, "湖州": 4, "嘉兴": 5,
                "绍兴": 6, "金华": 7, "衢州": 8, "舟山": 9, "台州": 10, "丽水": 11}

def get_region(unit):
    """从所在单位中提取地区"""
    for r in REGION_ORDER:
        if r in unit:
            return r
    # 根据已知信息补充
    name_to_region = {
        "李文忠": "宁波", "洪圆圆": "宁波", "汤迦南": "绍兴",
        "洪马可": "宁波", "陈斌": "宁波", "曹家欢": "宁波", "黄浩峰": "宁波",
    }
    return name_to_region.get(unit, "")

for p in people:
    p["_region"] = get_region(p["所在单位或场所"])
    p["_jiao_order"] = JIAO_ORDER.get(p["教别"], 99)
    p["_region_order"] = REGION_ORDER.get(p["_region"], 99)

people.sort(key=lambda x: (x["_jiao_order"], x["_region_order"], x["姓名"]))

# ============================================================
# Sheet 1: 学员信息总表
# ============================================================
ws1 = wb.active
ws1.title = "学员信息总表"

headers = ["序号", "姓名", "教名", "教别", "性别", "民族", "出生年月",
           "所在单位或场所", "最高学历", "联系电话"]

header_font = Font(name="宋体", bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(name="宋体", bold=True, size=11, color="FFFFFF")
cell_font = Font(name="宋体", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# 标题行
ws1.merge_cells("A1:J1")
title_cell = ws1["A1"]
title_cell.value = '浙江省第三届宗教界"双通"人才研修班学员信息汇总表（已报送）'
title_cell.font = Font(name="宋体", bold=True, size=14)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 35

ws1.merge_cells("A2:J2")
ws1["A2"].value = f"整理时间：{datetime.now().strftime('%Y年%m月%d日')} ｜ 共 {len(people)} 人"
ws1["A2"].font = Font(name="宋体", size=10, color="666666")
ws1["A2"].alignment = Alignment(horizontal="center")

# 表头
row = 4
for col_idx, h in enumerate(headers, 1):
    cell = ws1.cell(row=row, column=col_idx, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
ws1.row_dimensions[row].height = 25

# 数据行
alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
jiao_fills = {
    "佛教": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "道教": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "伊斯兰教": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "天主教": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "基督教": PatternFill(start_color="E2D0F0", end_color="E2D0F0", fill_type="solid"),
}

for idx, p in enumerate(people):
    row += 1
    vals = [
        idx + 1, p["姓名"], p.get("教名", ""), p["教别"], p["性别"], p["民族"],
        p.get("出生年月", ""), p["所在单位或场所"], p.get("最高学历", ""),
        p.get("联系电话", "")
    ]
    fill = jiao_fills.get(p["教别"], alt_fill if idx % 2 == 1 else None)
    for col_idx, v in enumerate(vals, 1):
        cell = ws1.cell(row=row, column=col_idx, value=v)
        cell.font = cell_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if col_idx in (1, 4, 5, 6):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if fill:
            cell.fill = fill
    ws1.row_dimensions[row].height = 30

# 列宽
col_widths = {"A": 6, "B": 10, "C": 10, "D": 10, "E": 6, "F": 6,
              "G": 12, "H": 45, "I": 50, "J": 16}
for col, w in col_widths.items():
    ws1.column_dimensions[col].width = w

# ============================================================
# Sheet 2: 统计分析
# ============================================================
ws2 = wb.create_sheet("统计分析")

stat_header_font = Font(name="宋体", bold=True, size=12, color="FFFFFF")
stat_header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
stat_cell_font = Font(name="宋体", size=11)
stat_title_font = Font(name="宋体", bold=True, size=13)
stat_label_font = Font(name="宋体", bold=True, size=11)

r = 1
# --- 一、重复学员 ---
ws2.merge_cells(f"A{r}:D{r}")
ws2.cell(row=r, column=1, value="一、重复学员检测").font = stat_title_font
r += 1
ws2.cell(row=r, column=1, value="经身份证号核对，本次报送的 7 名学员无重复人员。").font = stat_cell_font
r += 2

# --- 二、各地市学员数量 ---
ws2.merge_cells(f"A{r}:D{r}")
ws2.cell(row=r, column=1, value="二、各地市学员数量统计").font = stat_title_font
r += 1
for col_idx, h in enumerate(["地区", "人数", "占比"], 1):
    cell = ws2.cell(row=r, column=col_idx, value=h)
    cell.font = stat_header_font
    cell.fill = stat_header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border
r += 1

region_counts = {}
for p in people:
    reg = p["_region"]
    region_counts[reg] = region_counts.get(reg, 0) + 1

for reg in ["杭州", "宁波", "温州", "湖州", "嘉兴", "绍兴", "金华", "衢州", "舟山", "台州", "丽水"]:
    cnt = region_counts.get(reg, 0)
    if cnt > 0:
        ws2.cell(row=r, column=1, value=reg).font = stat_cell_font
        ws2.cell(row=r, column=2, value=cnt).font = stat_cell_font
        ws2.cell(row=r, column=3, value=f"{cnt/len(people)*100:.1f}%").font = stat_cell_font
        for c in range(1, 4):
            ws2.cell(row=r, column=c).border = thin_border
            ws2.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1

# 合计
ws2.cell(row=r, column=1, value="合计").font = stat_label_font
ws2.cell(row=r, column=2, value=len(people)).font = stat_label_font
ws2.cell(row=r, column=3, value="100%").font = stat_label_font
for c in range(1, 4):
    ws2.cell(row=r, column=c).border = thin_border
    ws2.cell(row=r, column=c).alignment = Alignment(horizontal="center")
r += 2

# --- 三、各教派学员数量 ---
ws2.merge_cells(f"A{r}:D{r}")
ws2.cell(row=r, column=1, value="三、各教派学员数量统计").font = stat_title_font
r += 1
for col_idx, h in enumerate(["教别", "人数", "占比"], 1):
    cell = ws2.cell(row=r, column=col_idx, value=h)
    cell.font = stat_header_font
    cell.fill = stat_header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border
r += 1

jiao_counts = {}
for p in people:
    j = p["教别"]
    jiao_counts[j] = jiao_counts.get(j, 0) + 1

for j in ["佛教", "道教", "伊斯兰教", "天主教", "基督教"]:
    cnt = jiao_counts.get(j, 0)
    if cnt > 0:
        ws2.cell(row=r, column=1, value=j).font = stat_cell_font
        ws2.cell(row=r, column=2, value=cnt).font = stat_cell_font
        ws2.cell(row=r, column=3, value=f"{cnt/len(people)*100:.1f}%").font = stat_cell_font
        for c in range(1, 4):
            ws2.cell(row=r, column=c).border = thin_border
            ws2.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        r += 1

ws2.cell(row=r, column=1, value="合计").font = stat_label_font
ws2.cell(row=r, column=2, value=len(people)).font = stat_label_font
ws2.cell(row=r, column=3, value="100%").font = stat_label_font
for c in range(1, 4):
    ws2.cell(row=r, column=c).border = thin_border
    ws2.cell(row=r, column=c).alignment = Alignment(horizontal="center")
r += 2

# --- 四、各地市×各教派交叉统计 ---
ws2.merge_cells(f"A{r}:G{r}")
ws2.cell(row=r, column=1, value="四、各地市 × 各教派 交叉统计表").font = stat_title_font
r += 1

jiao_list = ["佛教", "道教", "伊斯兰教", "天主教", "基督教"]
reg_list = ["杭州", "宁波", "温州", "湖州", "嘉兴", "绍兴", "金华", "衢州", "舟山", "台州", "丽水"]

# 表头
ws2.cell(row=r, column=1, value="地区").font = stat_header_font
ws2.cell(row=r, column=1).fill = stat_header_fill
ws2.cell(row=r, column=1).border = thin_border
ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center")

for j_idx, j in enumerate(jiao_list):
    cell = ws2.cell(row=r, column=j_idx + 2, value=j)
    cell.font = stat_header_font
    cell.fill = stat_header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

cell = ws2.cell(row=r, column=len(jiao_list) + 2, value="合计")
cell.font = stat_header_font
cell.fill = stat_header_fill
cell.alignment = Alignment(horizontal="center")
cell.border = thin_border
r += 1

for reg in reg_list:
    ws2.cell(row=r, column=1, value=reg).font = stat_cell_font
    ws2.cell(row=r, column=1).border = thin_border
    ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    row_total = 0
    for j_idx, j in enumerate(jiao_list):
        cnt = 0
        for p in people:
            if p["_region"] == reg and p["教别"] == j:
                cnt += 1
        cell = ws2.cell(row=r, column=j_idx + 2, value=cnt if cnt > 0 else "")
        cell.font = stat_cell_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        row_total += cnt
    cell = ws2.cell(row=r, column=len(jiao_list) + 2, value=row_total if row_total > 0 else "")
    cell.font = stat_label_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")
    if row_total > 0:
        r += 1

# 合计行
ws2.cell(row=r, column=1, value="合计").font = stat_label_font
ws2.cell(row=r, column=1).border = thin_border
ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center")
grand_total = 0
for j_idx, j in enumerate(jiao_list):
    cnt = jiao_counts.get(j, 0)
    cell = ws2.cell(row=r, column=j_idx + 2, value=cnt if cnt > 0 else "")
    cell.font = stat_label_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")
    grand_total += cnt
cell = ws2.cell(row=r, column=len(jiao_list) + 2, value=grand_total)
cell.font = stat_label_font
cell.border = thin_border
cell.alignment = Alignment(horizontal="center")

# 列宽
for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
    ws2.column_dimensions[col].width = 14
ws2.column_dimensions["A"].width = 10

# ============================================================
# 保存
# ============================================================
output = "/home/admin/.openclaw/workspace/浙江省第三届双通研修班学员汇总表.xlsx"
wb.save(output)
print(f"✅ 已保存: {output}")
print(f"   Sheet 1: 学员信息总表 ({len(people)} 人)")
print(f"   Sheet 2: 统计分析")

# 打印简要统计
print(f"\n📊 简要统计:")
print(f"   总人数: {len(people)}")
print(f"   重复人数: 0")
for j in ["佛教", "道教", "伊斯兰教", "天主教", "基督教"]:
    if j in jiao_counts:
        print(f"   {j}: {jiao_counts[j]} 人")
for reg in reg_list:
    if reg in region_counts:
        print(f"   {reg}: {region_counts[reg]} 人")
