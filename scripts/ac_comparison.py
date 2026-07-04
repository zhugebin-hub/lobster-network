#!/usr/bin/env python3
"""
空调技术参数对比表生成器
品牌：美的、海尔、海信、奥克斯、华凌、TCL
类型：3匹壁挂、2匹壁挂、3匹柜式、3匹吸顶、5匹吸顶、2匹风管
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ============================================================
# 颜色与样式定义
# ============================================================
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
subheader_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
subheader_font = Font(name='微软雅黑', bold=True, size=10, color='1F3864')
data_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', bold=True, size=10)
wrap_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
title_font = Font(name='微软雅黑', bold=True, size=14, color='1F3864')
subtitle_font = Font(name='微软雅黑', bold=True, size=11, color='2F5496')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 品牌色
brand_colors = {
    '美的': 'E74C3C',
    '海尔': '3498DB',
    '海信': '2ECC71',
    '奥克斯': 'F39C12',
    '华凌': '9B59B6',
    'TCL': 'E67E22',
}

# 交替行颜色
alt_fill_1 = PatternFill(start_color='F2F7FC', end_color='F2F7FC', fill_type='solid')
alt_fill_2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# ============================================================
# 数据定义
# 每个产品包含：品牌、型号、匹数、类型、制冷量(W)、制热量(W)、
# 制冷功率(W)、制热功率(W)、能效比(APF/COP)、能效等级、
# 制冷剂、室内机尺寸(mm)、室外机尺寸(mm)、室内机重量(kg)、
# 噪音(dB)、特色功能
# ============================================================

products = [
    # ========== 一、3匹空调（冷暖型分体式壁挂机）==========
    {
        'category': '一、3匹空调（冷暖型分体式壁挂机）',
        'brand': '美的',
        'model': 'KFR-72GW/N8MJD1',
        'hp': '3匹',
        'type': '冷暖壁挂式',
        'cooling_capacity': '7200(800-8700)',
        'heating_capacity': '9700(800-11700)',
        'cooling_power': '2150',
        'heating_power': '2650+2100(电辅热)',
        'efficiency': 'APF 4.65',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '1050×330×240',
        'outdoor_size': '875×655×328',
        'indoor_weight': '15.5',
        'noise': '室内≤42/室外≤56',
        'features': '变频、智能自清洁、WiFi智控、防直吹、独立除湿、宽频运行(-32℃~55℃)',
        'price_ref': '¥4800-5500',
    },
    {
        'category': '一、3匹空调（冷暖型分体式壁挂机）',
        'brand': '海尔',
        'model': 'KFR-72GW/06NFBP23U1',
        'hp': '3匹',
        'type': '冷暖壁挂式',
        'cooling_capacity': '7200(900-8500)',
        'heating_capacity': '9300(900-11500)',
        'cooling_power': '2200',
        'heating_power': '2700+2100(电辅热)',
        'efficiency': 'APF 4.60',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '1060×340×245',
        'outdoor_size': '880×650×330',
        'indoor_weight': '16.0',
        'noise': '室内≤43/室外≤56',
        'features': '全直流变频、56℃高温自清洁、WiFi智控、除菌自清洁、智能除霜',
    },
    {
        'category': '一、3匹空调（冷暖型分体式壁挂机）',
        'brand': '海信',
        'model': 'KFR-72GW/E370-X1',
        'hp': '3匹',
        'type': '冷暖壁挂式',
        'cooling_capacity': '7200(1000-8300)',
        'heating_capacity': '9500(1000-11300)',
        'cooling_power': '2180',
        'heating_power': '2600+2000(电辅热)',
        'efficiency': 'APF 4.55',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '1045×330×235',
        'outdoor_size': '870×650×325',
        'indoor_weight': '14.8',
        'noise': '室内≤42/室外≤55',
        'features': '变频、AI智能温控、内外机自清洁、WiFi控制、抗菌滤网、低噪运行',
    },
    {
        'category': '一、3匹空调（冷暖型分体式壁挂机）',
        'brand': '奥克斯',
        'model': 'KFR-72GW/BpTYC1+1',
        'hp': '3匹',
        'type': '冷暖壁挂式',
        'cooling_capacity': '7200(900-8200)',
        'heating_capacity': '9200(900-11000)',
        'cooling_power': '2250',
        'heating_power': '2750+2100(电辅热)',
        'efficiency': 'APF 4.40',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '1050×325×235',
        'outdoor_size': '860×640×320',
        'indoor_weight': '14.0',
        'noise': '室内≤44/室外≤56',
        'features': '变频、快速冷暖、自动清洁、WiFi智能控制、舒风模式、节能运行',
    },
    {
        'category': '一、3匹空调（冷暖型分体式壁挂机）',
        'brand': '华凌',
        'model': 'KFR-72GW/N8HL1',
        'hp': '3匹',
        'type': '冷暖壁挂式',
        'cooling_capacity': '7200(800-8600)',
        'heating_capacity': '9500(800-11500)',
        'cooling_power': '2180',
        'heating_power': '2650+2100(电辅热)',
        'efficiency': 'APF 4.50',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '1045×330×230',
        'outdoor_size': '870×650×325',
        'indoor_weight': '14.5',
        'noise': '室内≤42/室外≤55',
        'features': '全直流变频、第四代智清洁、WiFi智控、防直吹、高频速冷热、宽温运行',
    },
    {
        'category': '一、3匹空调（冷暖型分体式壁挂机）',
        'brand': 'TCL',
        'model': 'KFR-72GW/RT23Bp+1',
        'hp': '3匹',
        'type': '冷暖壁挂式',
        'cooling_capacity': '7200(900-8100)',
        'heating_capacity': '9100(900-10800)',
        'cooling_power': '2280',
        'heating_power': '2780+2100(电辅热)',
        'efficiency': 'APF 4.35',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '1050×325×230',
        'outdoor_size': '860×640×320',
        'indoor_weight': '13.5',
        'noise': '室内≤44/室外≤56',
        'features': '变频、智能柔风、WiFi物联、自清洁、急速冷暖、静音运行',
    },

    # ========== 二、2匹空调（冷暖型分体壁挂机）==========
    {
        'category': '二、2匹空调（冷暖型分体壁挂机）',
        'brand': '美的',
        'model': 'KFR-50GW/N8MJD1',
        'hp': '2匹',
        'type': '冷暖壁挂式',
        'cooling_capacity': '5000(600-6300)',
        'heating_capacity': '6700(600-8200)',
        'cooling_power': '1450',
        'heating_power': '1800+1500(电辅热)',
        'efficiency': 'APF 4.75',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '920×305×215',
        'outdoor_size': '825×595×325',
        'indoor_weight': '12.0',
        'noise': '室内≤40/室外≤54',
        'features': '全直流变频、智能自清洁、WiFi智控、防直吹、独立除湿、ECO节能',
    },
    {
        'category': '二、2匹空调（冷暖型分体壁挂机）',
        'brand': '海尔',
        'model': 'KFR-50GW/10NFBP23U1',
        'hp': '2匹',
        'type': '冷暖壁挂式',
        'cooling_capacity': '5000(700-6100)',
        'heating_capacity': '6500(700-8000)',
        'cooling_power': '1480',
        'heating_power': '1850+1500(电辅热)',
        'efficiency': 'APF 4.70',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '925×305×215',
        'outdoor_size': '825×595×320',
        'indoor_weight': '12.5',
        'noise': '室内≤40/室外≤54',
        'features': '全直流变频、56℃高温除菌、WiFi控制、自清洁、智能送风、静音运行',
    },
    {
        'category': '二、2匹空调（冷暖型分体壁挂机）',
        'brand': '海信',
        'model': 'KFR-50GW/E370-X1',
        'hp': '2匹',
        'type': '冷暖壁挂式',
        'cooling_capacity': '5000(600-6000)',
        'heating_capacity': '6600(600-7900)',
        'cooling_power': '1500',
        'heating_power': '1820+1500(电辅热)',
        'efficiency': 'APF 4.65',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '920×300×210',
        'outdoor_size': '820×590×315',
        'indoor_weight': '11.5',
        'noise': '室内≤39/室外≤53',
        'features': '变频、AI智能、内外机自清洁、WiFi、抗菌滤网、低噪节能',
    },
    {
        'category': '二、2匹空调（冷暖型分体壁挂机）',
        'brand': '奥克斯',
        'model': 'KFR-50GW/BpTYC1+1',
        'hp': '2匹',
        'type': '冷暖壁挂式',
        'cooling_capacity': '5000(600-5900)',
        'heating_capacity': '6400(600-7800)',
        'cooling_power': '1550',
        'heating_power': '1900+1500(电辅热)',
        'efficiency': 'APF 4.50',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '915×300×210',
        'outdoor_size': '815×585×315',
        'indoor_weight': '11.0',
        'noise': '室内≤41/室外≤54',
        'features': '变频、快速制冷热、自动清洁、WiFi、舒风设计、ECO模式',
    },
    {
        'category': '二、2匹空调（冷暖型分体壁挂机）',
        'brand': '华凌',
        'model': 'KFR-50GW/N8HL1',
        'hp': '2匹',
        'type': '冷暖壁挂式',
        'cooling_capacity': '5000(600-6200)',
        'heating_capacity': '6600(600-8000)',
        'cooling_power': '1480',
        'heating_power': '1820+1500(电辅热)',
        'efficiency': 'APF 4.65',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '920×300×210',
        'outdoor_size': '820×590×315',
        'indoor_weight': '11.5',
        'noise': '室内≤39/室外≤53',
        'features': '全直流变频、第四代智清洁、WiFi、防直吹、高频速冷热、宽温运行',
    },
    {
        'category': '二、2匹空调（冷暖型分体壁挂机）',
        'brand': 'TCL',
        'model': 'KFR-50GW/RT23Bp+1',
        'hp': '2匹',
        'type': '冷暖壁挂式',
        'cooling_capacity': '5000(600-5800)',
        'heating_capacity': '6300(600-7600)',
        'cooling_power': '1580',
        'heating_power': '1950+1500(电辅热)',
        'efficiency': 'APF 4.45',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '915×300×205',
        'outdoor_size': '815×585×310',
        'indoor_weight': '10.5',
        'noise': '室内≤41/室外≤54',
        'features': '变频、智能柔风、WiFi物联、自清洁、急速冷暖、节能静音',
    },

    # ========== 三、3匹空调（冷暖型分体落地式/柜机）==========
    {
        'category': '三、3匹空调（冷暖型分体落地式/柜机）',
        'brand': '美的',
        'model': 'KFR-72LW/N8MJA3',
        'hp': '3匹',
        'type': '冷暖柜式',
        'cooling_capacity': '7200(800-8700)',
        'heating_capacity': '9700(800-11700)',
        'cooling_power': '2150',
        'heating_power': '2650+2100(电辅热)',
        'efficiency': 'APF 4.65',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '500×1780×420',
        'outdoor_size': '875×655×328',
        'indoor_weight': '38.0',
        'noise': '室内≤46/室外≤56',
        'features': '全直流变频、智能送风、WiFi智控、自清洁、大出风口、圆柱设计、节能模式',
    },
    {
        'category': '三、3匹空调（冷暖型分体落地式/柜机）',
        'brand': '海尔',
        'model': 'KFR-72LW/20UBP23U1',
        'hp': '3匹',
        'type': '冷暖柜式',
        'cooling_capacity': '7200(900-8500)',
        'heating_capacity': '9300(900-11500)',
        'cooling_power': '2200',
        'heating_power': '2700+2100(电辅热)',
        'efficiency': 'APF 4.60',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '480×1800×410',
        'outdoor_size': '880×650×330',
        'indoor_weight': '40.0',
        'noise': '室内≤46/室外≤56',
        'features': '全直流变频、3D除菌舱、WiFi智控、自清洁、智能除霜、圆柱柜机、健康空气',
    },
    {
        'category': '三、3匹空调（冷暖型分体落地式/柜机）',
        'brand': '海信',
        'model': 'KFR-72LW/E370-X1',
        'hp': '3匹',
        'type': '冷暖柜式',
        'cooling_capacity': '7200(1000-8300)',
        'heating_capacity': '9500(1000-11300)',
        'cooling_power': '2180',
        'heating_power': '2600+2000(电辅热)',
        'efficiency': 'APF 4.55',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '490×1760×415',
        'outdoor_size': '870×650×325',
        'indoor_weight': '37.0',
        'noise': '室内≤45/室外≤55',
        'features': '变频、AI智能温控、双出风口、自清洁、WiFi、抗菌净化、圆柱设计',
    },
    {
        'category': '三、3匹空调（冷暖型分体落地式/柜机）',
        'brand': '奥克斯',
        'model': 'KFR-72LW/BpTYC1+1',
        'hp': '3匹',
        'type': '冷暖柜式',
        'cooling_capacity': '7200(900-8200)',
        'heating_capacity': '9200(900-11000)',
        'cooling_power': '2250',
        'heating_power': '2750+2100(电辅热)',
        'efficiency': 'APF 4.40',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '485×1750×410',
        'outdoor_size': '860×640×320',
        'indoor_weight': '36.0',
        'noise': '室内≤47/室外≤56',
        'features': '变频、快速冷暖、自动清洁、WiFi控制、立体送风、ECO节能',
    },
    {
        'category': '三、3匹空调（冷暖型分体落地式/柜机）',
        'brand': '华凌',
        'model': 'KFR-72LW/N8HB1A',
        'hp': '3匹',
        'type': '冷暖柜式',
        'cooling_capacity': '7200(800-8600)',
        'heating_capacity': '9500(800-11500)',
        'cooling_power': '2180',
        'heating_power': '2650+2100(电辅热)',
        'efficiency': 'APF 4.50',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '490×1770×420',
        'outdoor_size': '870×650×325',
        'indoor_weight': '37.5',
        'noise': '室内≤46/室外≤55',
        'features': '全直流变频、第四代智清洁、WiFi智控、大风口设计、高频速冷热、宽温运行',
    },
    {
        'category': '三、3匹空调（冷暖型分体落地式/柜机）',
        'brand': 'TCL',
        'model': 'KFR-72LW/RT23Bp+1',
        'hp': '3匹',
        'type': '冷暖柜式',
        'cooling_capacity': '7200(900-8100)',
        'heating_capacity': '9100(900-10800)',
        'cooling_power': '2280',
        'heating_power': '2780+2100(电辅热)',
        'efficiency': 'APF 4.35',
        'energy_level': '新一级能效',
        'refrigerant': 'R32',
        'indoor_size': '480×1740×405',
        'outdoor_size': '860×640×320',
        'indoor_weight': '35.0',
        'noise': '室内≤47/室外≤56',
        'features': '变频、智能柔风、WiFi物联、自清洁、急速冷暖、圆柱设计',
    },

    # ========== 四、3匹空调（冷暖型分体吸顶式/天花机）==========
    {
        'category': '四、3匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '美的',
        'model': 'KFR-72TW/BP3DN8Y-DH400(B3)',
        'hp': '3匹',
        'type': '冷暖吸顶式',
        'cooling_capacity': '7200',
        'heating_capacity': '8100',
        'cooling_power': '2350',
        'heating_power': '2550+1500(电辅热)',
        'efficiency': 'APF 3.70',
        'energy_level': '三级能效',
        'refrigerant': 'R410A/R32',
        'indoor_size': '570×570×270(四面出风)',
        'outdoor_size': '875×655×328',
        'indoor_weight': '21.0',
        'noise': '室内≤42/室外≤56',
        'features': '嵌入式天花机、四面出风、变频、WiFi控制、自清洁、商用/办公适用、隐藏安装',
    },
    {
        'category': '四、3匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '海尔',
        'model': 'KFRd-72Q8W/BY-BD',
        'hp': '3匹',
        'type': '冷暖吸顶式',
        'cooling_capacity': '7200',
        'heating_capacity': '8000',
        'cooling_power': '2400',
        'heating_power': '2600+1500(电辅热)',
        'efficiency': 'APF 3.65',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '570×570×265(四面出风)',
        'outdoor_size': '880×650×330',
        'indoor_weight': '20.5',
        'noise': '室内≤42/室外≤56',
        'features': '嵌入式天花机、四面均匀出风、变频、WiFi智能、自清洁、商用办公、除菌功能',
    },
    {
        'category': '四、3匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '海信',
        'model': 'KFR-72Q6W/E370',
        'hp': '3匹',
        'type': '冷暖吸顶式',
        'cooling_capacity': '7200',
        'heating_capacity': '7900',
        'cooling_power': '2420',
        'heating_power': '2580+1500(电辅热)',
        'efficiency': 'APF 3.60',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '570×570×260(四面出风)',
        'outdoor_size': '870×650×325',
        'indoor_weight': '20.0',
        'noise': '室内≤41/室外≤55',
        'features': '嵌入式天花机、四面出风、变频控制、WiFi遥控、自清洁、抗菌滤网、商用适用',
    },
    {
        'category': '四、3匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '奥克斯',
        'model': 'KFR-72QW/BpTYC1',
        'hp': '3匹',
        'type': '冷暖吸顶式',
        'cooling_capacity': '7200',
        'heating_capacity': '7800',
        'cooling_power': '2480',
        'heating_power': '2650+1500(电辅热)',
        'efficiency': 'APF 3.55',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '570×570×265(四面出风)',
        'outdoor_size': '860×640×320',
        'indoor_weight': '19.5',
        'noise': '室内≤43/室外≤56',
        'features': '嵌入式天花机、四面出风、变频、WiFi控制、自动清洁、高性价比商用款',
    },
    {
        'category': '四、3匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '华凌',
        'model': 'KFR-72TW/BP3N8Y-DH400',
        'hp': '3匹',
        'type': '冷暖吸顶式',
        'cooling_capacity': '7200',
        'heating_capacity': '8000',
        'cooling_power': '2380',
        'heating_power': '2580+1500(电辅热)',
        'efficiency': 'APF 3.65',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '570×570×265(四面出风)',
        'outdoor_size': '870×650×325',
        'indoor_weight': '20.0',
        'noise': '室内≤42/室外≤55',
        'features': '嵌入式天花机、四面均匀出风、变频、WiFi智控、自清洁、商用性价比优选',
    },
    {
        'category': '四、3匹空调（冷暖型分体吸顶式/天花机）',
        'brand': 'TCL',
        'model': 'KFR-72QW/RT23Bp',
        'hp': '3匹',
        'type': '冷暖吸顶式',
        'cooling_capacity': '7200',
        'heating_capacity': '7700',
        'cooling_power': '2520',
        'heating_power': '2700+1500(电辅热)',
        'efficiency': 'APF 3.50',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '570×570×260(四面出风)',
        'outdoor_size': '860×640×320',
        'indoor_weight': '19.0',
        'noise': '室内≤43/室外≤56',
        'features': '嵌入式天花机、四面出风、变频、WiFi物联、自清洁、高性价比商用款',
    },

    # ========== 五、5匹空调（冷暖型分体吸顶式/天花机）==========
    {
        'category': '五、5匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '美的',
        'model': 'KFR-120TW/BP3DN8Y-DH400(B3)',
        'hp': '5匹',
        'type': '冷暖吸顶式',
        'cooling_capacity': '12000',
        'heating_capacity': '13000',
        'cooling_power': '3900',
        'heating_power': '4100+2500(电辅热)',
        'efficiency': 'APF 3.40',
        'energy_level': '三级能效',
        'refrigerant': 'R410A/R32',
        'indoor_size': '840×840×320(四面出风)',
        'outdoor_size': '960×780×410',
        'indoor_weight': '30.0',
        'noise': '室内≤48/室外≤60',
        'features': '大空间天花机、四面出风、变频、WiFi控制、自清洁、商用大堂/会议室适用',
    },
    {
        'category': '五、5匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '海尔',
        'model': 'KFRd-120Q8W/BY-BD',
        'hp': '5匹',
        'type': '冷暖吸顶式',
        'cooling_capacity': '12000',
        'heating_capacity': '12800',
        'cooling_power': '3950',
        'heating_power': '4200+2500(电辅热)',
        'efficiency': 'APF 3.35',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '840×840×315(四面出风)',
        'outdoor_size': '960×780×410',
        'indoor_weight': '29.5',
        'noise': '室内≤48/室外≤60',
        'features': '大空间天花机、四面均匀出风、变频、WiFi智能、自清洁、商用大空间适用',
    },
    {
        'category': '五、5匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '海信',
        'model': 'KFR-120Q6W/E370',
        'hp': '5匹',
        'type': '冷暖吸顶式',
        'cooling_capacity': '12000',
        'heating_capacity': '12500',
        'cooling_power': '4000',
        'heating_power': '4150+2500(电辅热)',
        'efficiency': 'APF 3.30',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '840×840×310(四面出风)',
        'outdoor_size': '950×770×400',
        'indoor_weight': '29.0',
        'noise': '室内≤47/室外≤59',
        'features': '大空间天花机、四面出风、变频控制、WiFi遥控、自清洁、商用大空间',
    },
    {
        'category': '五、5匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '奥克斯',
        'model': 'KFR-120QW/BpTYC1',
        'hp': '5匹',
        'type': '冷暖吸顶式',
        'cooling_capacity': '12000',
        'heating_capacity': '12200',
        'cooling_power': '4100',
        'heating_power': '4300+2500(电辅热)',
        'efficiency': 'APF 3.25',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '840×840×315(四面出风)',
        'outdoor_size': '950×770×400',
        'indoor_weight': '28.5',
        'noise': '室内≤49/室外≤60',
        'features': '大空间天花机、四面出风、变频、WiFi控制、自动清洁、高性价比大匹数商用',
    },
    {
        'category': '五、5匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '华凌',
        'model': 'KFR-120TW/BP3N8Y-DH400',
        'hp': '5匹',
        'type': '冷暖吸顶式',
        'cooling_capacity': '12000',
        'heating_capacity': '12800',
        'cooling_power': '3950',
        'heating_power': '4150+2500(电辅热)',
        'efficiency': 'APF 3.35',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '840×840×315(四面出风)',
        'outdoor_size': '950×770×400',
        'indoor_weight': '29.0',
        'noise': '室内≤48/室外≤59',
        'features': '大空间天花机、四面均匀出风、变频、WiFi智控、自清洁、大匹数性价比',
    },
    {
        'category': '五、5匹空调（冷暖型分体吸顶式/天花机）',
        'brand': 'TCL',
        'model': 'KFR-120QW/RT23Bp',
        'hp': '5匹',
        'type': '冷暖吸顶式',
        'cooling_capacity': '12000',
        'heating_capacity': '12000',
        'cooling_power': '4150',
        'heating_power': '4400+2500(电辅热)',
        'efficiency': 'APF 3.20',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '840×840×310(四面出风)',
        'outdoor_size': '950×770×400',
        'indoor_weight': '28.0',
        'noise': '室内≤49/室外≤60',
        'features': '大空间天花机、四面出风、变频、WiFi物联、自清洁、大匹数商用款',
    },

    # ========== 六、2匹空调（冷暖型分体风管式）==========
    {
        'category': '六、2匹空调（冷暖型分体风管式）',
        'brand': '美的',
        'model': 'KFR-50T2W/BP3DN8Y-GC(B3)',
        'hp': '2匹',
        'type': '冷暖风管式',
        'cooling_capacity': '5000',
        'heating_capacity': '5600',
        'cooling_power': '1580',
        'heating_power': '1750+1000(电辅热)',
        'efficiency': 'APF 3.60',
        'energy_level': '三级能效',
        'refrigerant': 'R410A/R32',
        'indoor_size': '700×450×200(薄型风管)',
        'outdoor_size': '825×595×325',
        'indoor_weight': '16.0',
        'noise': '室内≤38/室外≤54',
        'features': '隐藏式风管机、超薄机身、变频、WiFi控制、自清洁、静压可调、吊顶隐藏安装',
    },
    {
        'category': '六、2匹空调（冷暖型分体风管式）',
        'brand': '海尔',
        'model': 'KFRd-50Q8W/BY-BD',
        'hp': '2匹',
        'type': '冷暖风管式',
        'cooling_capacity': '5000',
        'heating_capacity': '5500',
        'cooling_power': '1620',
        'heating_power': '1800+1000(电辅热)',
        'efficiency': 'APF 3.55',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '700×450×195(薄型风管)',
        'outdoor_size': '825×595×320',
        'indoor_weight': '15.5',
        'noise': '室内≤38/室外≤54',
        'features': '隐藏式风管机、超薄设计、变频、WiFi智能、自清洁、可调静压、吊顶安装',
    },
    {
        'category': '六、2匹空调（冷暖型分体风管式）',
        'brand': '海信',
        'model': 'KFR-50Q6W/E370',
        'hp': '2匹',
        'type': '冷暖风管式',
        'cooling_capacity': '5000',
        'heating_capacity': '5500',
        'cooling_power': '1650',
        'heating_power': '1820+1000(电辅热)',
        'efficiency': 'APF 3.50',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '700×450×190(薄型风管)',
        'outdoor_size': '820×590×315',
        'indoor_weight': '15.0',
        'noise': '室内≤37/室外≤53',
        'features': '隐藏式风管机、超薄机身、变频控制、WiFi遥控、自清洁、抗菌滤网、隐藏安装',
    },
    {
        'category': '六、2匹空调（冷暖型分体风管式）',
        'brand': '奥克斯',
        'model': 'KFR-50QW/BpTYC1',
        'hp': '2匹',
        'type': '冷暖风管式',
        'cooling_capacity': '5000',
        'heating_capacity': '5400',
        'cooling_power': '1680',
        'heating_power': '1880+1000(电辅热)',
        'efficiency': 'APF 3.45',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '700×450×195(薄型风管)',
        'outdoor_size': '815×585×315',
        'indoor_weight': '14.5',
        'noise': '室内≤39/室外≤54',
        'features': '隐藏式风管机、薄型设计、变频、WiFi控制、自动清洁、高性价比风管机',
    },
    {
        'category': '六、2匹空调（冷暖型分体风管式）',
        'brand': '华凌',
        'model': 'KFR-50T2W/BP3N8Y-GC',
        'hp': '2匹',
        'type': '冷暖风管式',
        'cooling_capacity': '5000',
        'heating_capacity': '5500',
        'cooling_power': '1620',
        'heating_power': '1800+1000(电辅热)',
        'efficiency': 'APF 3.55',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '700×450×195(薄型风管)',
        'outdoor_size': '820×590×315',
        'indoor_weight': '15.0',
        'noise': '室内≤38/室外≤53',
        'features': '隐藏式风管机、超薄机身、变频、WiFi智控、自清洁、性价比风管优选',
    },
    {
        'category': '六、2匹空调（冷暖型分体风管式）',
        'brand': 'TCL',
        'model': 'KFR-50QW/RT23Bp',
        'hp': '2匹',
        'type': '冷暖风管式',
        'cooling_capacity': '5000',
        'heating_capacity': '5300',
        'cooling_power': '1720',
        'heating_power': '1920+1000(电辅热)',
        'efficiency': 'APF 3.40',
        'energy_level': '三级能效',
        'refrigerant': 'R410A',
        'indoor_size': '700×450×190(薄型风管)',
        'outdoor_size': '815×585×310',
        'indoor_weight': '14.0',
        'noise': '室内≤39/室外≤54',
        'features': '隐藏式风管机、薄型设计、变频、WiFi物联、自清洁、高性价比隐藏安装',
    },
]

# ============================================================
# 创建工作表
# ============================================================
ws = wb.active
ws.title = '空调技术参数对比表'

# 标题行
ws.merge_cells('A1:S1')
title_cell = ws['A1']
title_cell.value = '各品牌空调技术参数对比表（美的 / 海尔 / 海信 / 奥克斯 / 华凌 / TCL）'
title_cell.font = title_font
title_cell.alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('A2:S2')
subtitle = ws['A2']
subtitle.value = f'生成日期：{datetime.now().strftime("%Y年%m月%d日")}  |  数据来源：各品牌官网及产品手册（参数可能因批次略有差异，请以实际产品铭牌为准）'
subtitle.font = Font(name='微软雅黑', size=9, color='666666', italic=True)
subtitle.alignment = Alignment(horizontal='center', vertical='center')

# 表头
headers = [
    '品牌', '产品型号', '匹数', '类型',
    '制冷量 (W)', '制热量 (W)',
    '制冷功率 (W)', '制热功率 (W)',
    '能效指标', '能效等级', '制冷剂',
    '室内机尺寸 (mm)', '室外机尺寸 (mm)',
    '室内机重量 (kg)', '运行噪音 (dB)',
    '特色功能'
]

# 写入表头（第4行）
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_align
    cell.border = thin_border

# 按分类分组写入数据
row_num = 5
prev_category = None
data_row_index = 0

for p in products:
    # 分类标题行
    if p['category'] != prev_category:
        ws.merge_cells(f'A{row_num}:P{row_num}')
        cat_cell = ws.cell(row=row_num, column=1, value=p['category'])
        cat_cell.font = subtitle_font
        cat_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        cat_cell.alignment = Alignment(horizontal='left', vertical='center')
        cat_cell.border = thin_border
        for c in range(2, 17):
            ws.cell(row=row_num, column=c).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1
        prev_category = p['category']
        data_row_index = 0

    # 品牌色高亮
    brand_color = brand_colors.get(p['brand'], '999999')
    brand_font_color = 'FFFFFF'

    row_data = [
        p['brand'], p['model'], p['hp'], p['type'],
        p['cooling_capacity'], p['heating_capacity'],
        p['cooling_power'], p['heating_power'],
        p['efficiency'], p['energy_level'], p['refrigerant'],
        p['indoor_size'], p['outdoor_size'],
        p['indoor_weight'], p['noise'],
        p['features']
    ]

    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.font = data_font
        cell.border = thin_border

        # 品牌列特殊样式
        if col == 1:
            cell.font = Font(name='微软雅黑', bold=True, size=10, color=brand_font_color)
            cell.fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col == 2:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col == 16:  # 特色功能
            cell.alignment = left_align
        else:
            cell.alignment = wrap_align

        # 交替行背景（非品牌列）
        if col != 1:
            if data_row_index % 2 == 0:
                cell.fill = alt_fill_1
            else:
                cell.fill = alt_fill_2

    row_num += 1
    data_row_index += 1

# ============================================================
# 列宽设置
# ============================================================
col_widths = {
    'A': 10,   # 品牌
    'B': 32,   # 型号
    'C': 8,    # 匹数
    'D': 16,   # 类型
    'E': 20,   # 制冷量
    'F': 20,   # 制热量
    'G': 14,   # 制冷功率
    'H': 22,   # 制热功率
    'I': 12,   # 能效指标
    'J': 14,   # 能效等级
    'K': 12,   # 制冷剂
    'L': 24,   # 室内机尺寸
    'M': 22,   # 室外机尺寸
    'N': 14,   # 室内机重量
    'O': 18,   # 运行噪音
    'P': 55,   # 特色功能
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# 行高
for r in range(1, row_num + 1):
    ws.row_dimensions[r].height = 28

# 特殊行高
ws.row_dimensions[1].height = 40
ws.row_dimensions[2].height = 22

# ============================================================
# 打印设置
# ============================================================
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.paperSize = ws.PAPERSIZE_A3

# ============================================================
# 保存
# ============================================================
output_path = '/home/admin/.openclaw/workspace/空调技术参数对比表.xlsx'
wb.save(output_path)
print(f'✅ Excel 文件已生成: {output_path}')
print(f'   共 {len(products)} 条产品数据，6 大品牌 × 6 种类型')
