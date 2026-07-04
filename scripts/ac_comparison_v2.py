#!/usr/bin/env python3
"""
空调技术参数对比表生成器（含参考价格）
品牌：美的、海尔、海信、奥克斯、华凌、TCL
类型：3匹壁挂、2匹壁挂、3匹柜式、3匹吸顶、5匹吸顶、2匹风管
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ============================================================
# 颜色与样式定义
# ============================================================
header_font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
data_font = Font(name='微软雅黑', size=9)
bold_font = Font(name='微软雅黑', bold=True, size=9)
wrap_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
title_font = Font(name='微软雅黑', bold=True, size=14, color='1F3864')
subtitle_font = Font(name='微软雅黑', bold=True, size=10, color='2F5496')

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

brand_colors = {
    '美的': 'E74C3C', '海尔': '3498DB', '海信': '2ECC71',
    '奥克斯': 'F39C12', '华凌': '9B59B6', 'TCL': 'E67E22',
}
brand_font_color = 'FFFFFF'
alt_fill_1 = PatternFill(start_color='F2F7FC', end_color='F2F7FC', fill_type='solid')
alt_fill_2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
cat_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
price_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')

# ============================================================
# 数据定义 (含参考价格)
# ============================================================

products = [
    # ========== 一、3匹空调（冷暖型分体式壁挂机）==========
    {
        'category': '一、3匹空调（冷暖型分体式壁挂机）',
        'brand': '美的', 'model': 'KFR-72GW/N8MJD1', 'hp': '3匹', 'type': '冷暖壁挂式',
        'cooling': '7200(800-8700)', 'heating': '9700(800-11700)',
        'cooling_power': '2150', 'heating_power': '2650+2100(电辅)',
        'efficiency': 'APF 4.65', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '1050×330×240', 'outdoor_size': '875×655×328',
        'indoor_weight': '15.5', 'noise': '室内≤42/室外≤56',
        'features': '变频、智能自清洁、WiFi智控、防直吹、独立除湿、宽频运行(-32℃~55℃)',
        'price': '¥4800-5500',
    },
    {
        'category': '一、3匹空调（冷暖型分体式壁挂机）',
        'brand': '海尔', 'model': 'KFR-72GW/06NFBP23U1', 'hp': '3匹', 'type': '冷暖壁挂式',
        'cooling': '7200(900-8500)', 'heating': '9300(900-11500)',
        'cooling_power': '2200', 'heating_power': '2700+2100(电辅)',
        'efficiency': 'APF 4.60', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '1060×340×245', 'outdoor_size': '880×650×330',
        'indoor_weight': '16.0', 'noise': '室内≤43/室外≤56',
        'features': '全直流变频、56℃高温自清洁、WiFi智控、除菌自清洁、智能除霜',
        'price': '¥4600-5300',
    },
    {
        'category': '一、3匹空调（冷暖型分体式壁挂机）',
        'brand': '海信', 'model': 'KFR-72GW/E370-X1', 'hp': '3匹', 'type': '冷暖壁挂式',
        'cooling': '7200(1000-8300)', 'heating': '9500(1000-11300)',
        'cooling_power': '2180', 'heating_power': '2600+2000(电辅)',
        'efficiency': 'APF 4.55', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '1045×330×235', 'outdoor_size': '870×650×325',
        'indoor_weight': '14.8', 'noise': '室内≤42/室外≤55',
        'features': '变频、AI智能温控、内外机自清洁、WiFi控制、抗菌滤网、低噪运行',
        'price': '¥4200-4900',
    },
    {
        'category': '一、3匹空调（冷暖型分体式壁挂机）',
        'brand': '奥克斯', 'model': 'KFR-72GW/BpTYC1+1', 'hp': '3匹', 'type': '冷暖壁挂式',
        'cooling': '7200(900-8200)', 'heating': '9200(900-11000)',
        'cooling_power': '2250', 'heating_power': '2750+2100(电辅)',
        'efficiency': 'APF 4.40', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '1050×325×235', 'outdoor_size': '860×640×320',
        'indoor_weight': '14.0', 'noise': '室内≤44/室外≤56',
        'features': '变频、快速冷暖、自动清洁、WiFi智能控制、舒风模式、节能运行',
        'price': '¥3800-4500',
    },
    {
        'category': '一、3匹空调（冷暖型分体式壁挂机）',
        'brand': '华凌', 'model': 'KFR-72GW/N8HL1', 'hp': '3匹', 'type': '冷暖壁挂式',
        'cooling': '7200(800-8600)', 'heating': '9500(800-11500)',
        'cooling_power': '2180', 'heating_power': '2650+2100(电辅)',
        'efficiency': 'APF 4.50', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '1045×330×230', 'outdoor_size': '870×650×325',
        'indoor_weight': '14.5', 'noise': '室内≤42/室外≤55',
        'features': '全直流变频、第四代智清洁、WiFi智控、防直吹、高频速冷热、宽温运行',
        'price': '¥3500-4200',
    },
    {
        'category': '一、3匹空调（冷暖型分体式壁挂机）',
        'brand': 'TCL', 'model': 'KFR-72GW/RT23Bp+1', 'hp': '3匹', 'type': '冷暖壁挂式',
        'cooling': '7200(900-8100)', 'heating': '9100(900-10800)',
        'cooling_power': '2280', 'heating_power': '2780+2100(电辅)',
        'efficiency': 'APF 4.35', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '1050×325×230', 'outdoor_size': '860×640×320',
        'indoor_weight': '13.5', 'noise': '室内≤44/室外≤56',
        'features': '变频、智能柔风、WiFi物联、自清洁、急速冷暖、静音运行',
        'price': '¥3600-4300',
    },

    # ========== 二、2匹空调（冷暖型分体壁挂机）==========
    {
        'category': '二、2匹空调（冷暖型分体壁挂机）',
        'brand': '美的', 'model': 'KFR-50GW/N8MJD1', 'hp': '2匹', 'type': '冷暖壁挂式',
        'cooling': '5000(600-6300)', 'heating': '6700(600-8200)',
        'cooling_power': '1450', 'heating_power': '1800+1500(电辅)',
        'efficiency': 'APF 4.75', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '920×305×215', 'outdoor_size': '825×595×325',
        'indoor_weight': '12.0', 'noise': '室内≤40/室外≤54',
        'features': '全直流变频、智能自清洁、WiFi智控、防直吹、独立除湿、ECO节能',
        'price': '¥3200-3800',
    },
    {
        'category': '二、2匹空调（冷暖型分体壁挂机）',
        'brand': '海尔', 'model': 'KFR-50GW/10NFBP23U1', 'hp': '2匹', 'type': '冷暖壁挂式',
        'cooling': '5000(700-6100)', 'heating': '6500(700-8000)',
        'cooling_power': '1480', 'heating_power': '1850+1500(电辅)',
        'efficiency': 'APF 4.70', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '925×305×215', 'outdoor_size': '825×595×320',
        'indoor_weight': '12.5', 'noise': '室内≤40/室外≤54',
        'features': '全直流变频、56℃高温除菌、WiFi控制、自清洁、智能送风、静音运行',
        'price': '¥3100-3700',
    },
    {
        'category': '二、2匹空调（冷暖型分体壁挂机）',
        'brand': '海信', 'model': 'KFR-50GW/E370-X1', 'hp': '2匹', 'type': '冷暖壁挂式',
        'cooling': '5000(600-6000)', 'heating': '6600(600-7900)',
        'cooling_power': '1500', 'heating_power': '1820+1500(电辅)',
        'efficiency': 'APF 4.65', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '920×300×210', 'outdoor_size': '820×590×315',
        'indoor_weight': '11.5', 'noise': '室内≤39/室外≤53',
        'features': '变频、AI智能、内外机自清洁、WiFi、抗菌滤网、低噪节能',
        'price': '¥2800-3500',
    },
    {
        'category': '二、2匹空调（冷暖型分体壁挂机）',
        'brand': '奥克斯', 'model': 'KFR-50GW/BpTYC1+1', 'hp': '2匹', 'type': '冷暖壁挂式',
        'cooling': '5000(600-5900)', 'heating': '6400(600-7800)',
        'cooling_power': '1550', 'heating_power': '1900+1500(电辅)',
        'efficiency': 'APF 4.50', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '915×300×210', 'outdoor_size': '815×585×315',
        'indoor_weight': '11.0', 'noise': '室内≤41/室外≤54',
        'features': '变频、快速制冷热、自动清洁、WiFi、舒风设计、ECO模式',
        'price': '¥2500-3200',
    },
    {
        'category': '二、2匹空调（冷暖型分体壁挂机）',
        'brand': '华凌', 'model': 'KFR-50GW/N8HL1', 'hp': '2匹', 'type': '冷暖壁挂式',
        'cooling': '5000(600-6200)', 'heating': '6600(600-8000)',
        'cooling_power': '1480', 'heating_power': '1820+1500(电辅)',
        'efficiency': 'APF 4.65', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '920×300×210', 'outdoor_size': '820×590×315',
        'indoor_weight': '11.5', 'noise': '室内≤39/室外≤53',
        'features': '全直流变频、第四代智清洁、WiFi、防直吹、高频速冷热、宽温运行',
        'price': '¥2300-3000',
    },
    {
        'category': '二、2匹空调（冷暖型分体壁挂机）',
        'brand': 'TCL', 'model': 'KFR-50GW/RT23Bp+1', 'hp': '2匹', 'type': '冷暖壁挂式',
        'cooling': '5000(600-5800)', 'heating': '6300(600-7600)',
        'cooling_power': '1580', 'heating_power': '1950+1500(电辅)',
        'efficiency': 'APF 4.45', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '915×300×205', 'outdoor_size': '815×585×310',
        'indoor_weight': '10.5', 'noise': '室内≤41/室外≤54',
        'features': '变频、智能柔风、WiFi物联、自清洁、急速冷暖、节能静音',
        'price': '¥2400-3100',
    },

    # ========== 三、3匹空调（冷暖型分体落地式/柜机）==========
    {
        'category': '三、3匹空调（冷暖型分体落地式/柜机）',
        'brand': '美的', 'model': 'KFR-72LW/N8MJA3', 'hp': '3匹', 'type': '冷暖柜式',
        'cooling': '7200(800-8700)', 'heating': '9700(800-11700)',
        'cooling_power': '2150', 'heating_power': '2650+2100(电辅)',
        'efficiency': 'APF 4.65', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '500×1780×420', 'outdoor_size': '875×655×328',
        'indoor_weight': '38.0', 'noise': '室内≤46/室外≤56',
        'features': '全直流变频、智能送风、WiFi智控、自清洁、大出风口、圆柱设计、节能模式',
        'price': '¥5500-6500',
    },
    {
        'category': '三、3匹空调（冷暖型分体落地式/柜机）',
        'brand': '海尔', 'model': 'KFR-72LW/20UBP23U1', 'hp': '3匹', 'type': '冷暖柜式',
        'cooling': '7200(900-8500)', 'heating': '9300(900-11500)',
        'cooling_power': '2200', 'heating_power': '2700+2100(电辅)',
        'efficiency': 'APF 4.60', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '480×1800×410', 'outdoor_size': '880×650×330',
        'indoor_weight': '40.0', 'noise': '室内≤46/室外≤56',
        'features': '全直流变频、3D除菌舱、WiFi智控、自清洁、智能除霜、圆柱柜机、健康空气',
        'price': '¥5300-6300',
    },
    {
        'category': '三、3匹空调（冷暖型分体落地式/柜机）',
        'brand': '海信', 'model': 'KFR-72LW/E370-X1', 'hp': '3匹', 'type': '冷暖柜式',
        'cooling': '7200(1000-8300)', 'heating': '9500(1000-11300)',
        'cooling_power': '2180', 'heating_power': '2600+2000(电辅)',
        'efficiency': 'APF 4.55', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '490×1760×415', 'outdoor_size': '870×650×325',
        'indoor_weight': '37.0', 'noise': '室内≤45/室外≤55',
        'features': '变频、AI智能温控、双出风口、自清洁、WiFi、抗菌净化、圆柱设计',
        'price': '¥4800-5800',
    },
    {
        'category': '三、3匹空调（冷暖型分体落地式/柜机）',
        'brand': '奥克斯', 'model': 'KFR-72LW/BpTYC1+1', 'hp': '3匹', 'type': '冷暖柜式',
        'cooling': '7200(900-8200)', 'heating': '9200(900-11000)',
        'cooling_power': '2250', 'heating_power': '2750+2100(电辅)',
        'efficiency': 'APF 4.40', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '485×1750×410', 'outdoor_size': '860×640×320',
        'indoor_weight': '36.0', 'noise': '室内≤47/室外≤56',
        'features': '变频、快速冷暖、自动清洁、WiFi控制、立体送风、ECO节能',
        'price': '¥4200-5200',
    },
    {
        'category': '三、3匹空调（冷暖型分体落地式/柜机）',
        'brand': '华凌', 'model': 'KFR-72LW/N8HB1A', 'hp': '3匹', 'type': '冷暖柜式',
        'cooling': '7200(800-8600)', 'heating': '9500(800-11500)',
        'cooling_power': '2180', 'heating_power': '2650+2100(电辅)',
        'efficiency': 'APF 4.50', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '490×1770×420', 'outdoor_size': '870×650×325',
        'indoor_weight': '37.5', 'noise': '室内≤46/室外≤55',
        'features': '全直流变频、第四代智清洁、WiFi智控、大风口设计、高频速冷热、宽温运行',
        'price': '¥3800-4800',
    },
    {
        'category': '三、3匹空调（冷暖型分体落地式/柜机）',
        'brand': 'TCL', 'model': 'KFR-72LW/RT23Bp+1', 'hp': '3匹', 'type': '冷暖柜式',
        'cooling': '7200(900-8100)', 'heating': '9100(900-10800)',
        'cooling_power': '2280', 'heating_power': '2780+2100(电辅)',
        'efficiency': 'APF 4.35', 'energy': '新一级', 'refrigerant': 'R32',
        'indoor_size': '480×1740×405', 'outdoor_size': '860×640×320',
        'indoor_weight': '35.0', 'noise': '室内≤47/室外≤56',
        'features': '变频、智能柔风、WiFi物联、自清洁、急速冷暖、圆柱设计',
        'price': '¥3900-4900',
    },

    # ========== 四、3匹空调（冷暖型分体吸顶式/天花机）==========
    {
        'category': '四、3匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '美的', 'model': 'KFR-72TW/BP3DN8Y-DH400(B3)', 'hp': '3匹', 'type': '冷暖吸顶式',
        'cooling': '7200', 'heating': '8100',
        'cooling_power': '2350', 'heating_power': '2550+1500(电辅)',
        'efficiency': 'APF 3.70', 'energy': '三级', 'refrigerant': 'R32',
        'indoor_size': '570×570×270(四面)', 'outdoor_size': '875×655×328',
        'indoor_weight': '21.0', 'noise': '室内≤42/室外≤56',
        'features': '嵌入式天花机、四面出风、变频、WiFi控制、自清洁、商用/办公适用、隐藏安装',
        'price': '¥5000-6000',
    },
    {
        'category': '四、3匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '海尔', 'model': 'KFRd-72Q8W/BY-BD', 'hp': '3匹', 'type': '冷暖吸顶式',
        'cooling': '7200', 'heating': '8000',
        'cooling_power': '2400', 'heating_power': '2600+1500(电辅)',
        'efficiency': 'APF 3.65', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '570×570×265(四面)', 'outdoor_size': '880×650×330',
        'indoor_weight': '20.5', 'noise': '室内≤42/室外≤56',
        'features': '嵌入式天花机、四面均匀出风、变频、WiFi智能、自清洁、商用办公、除菌功能',
        'price': '¥4800-5800',
    },
    {
        'category': '四、3匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '海信', 'model': 'KFR-72Q6W/E370', 'hp': '3匹', 'type': '冷暖吸顶式',
        'cooling': '7200', 'heating': '7900',
        'cooling_power': '2420', 'heating_power': '2580+1500(电辅)',
        'efficiency': 'APF 3.60', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '570×570×260(四面)', 'outdoor_size': '870×650×325',
        'indoor_weight': '20.0', 'noise': '室内≤41/室外≤55',
        'features': '嵌入式天花机、四面出风、变频控制、WiFi遥控、自清洁、抗菌滤网、商用适用',
        'price': '¥4500-5500',
    },
    {
        'category': '四、3匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '奥克斯', 'model': 'KFR-72QW/BpTYC1', 'hp': '3匹', 'type': '冷暖吸顶式',
        'cooling': '7200', 'heating': '7800',
        'cooling_power': '2480', 'heating_power': '2650+1500(电辅)',
        'efficiency': 'APF 3.55', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '570×570×265(四面)', 'outdoor_size': '860×640×320',
        'indoor_weight': '19.5', 'noise': '室内≤43/室外≤56',
        'features': '嵌入式天花机、四面出风、变频、WiFi控制、自动清洁、高性价比商用款',
        'price': '¥3800-4800',
    },
    {
        'category': '四、3匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '华凌', 'model': 'KFR-72TW/BP3N8Y-DH400', 'hp': '3匹', 'type': '冷暖吸顶式',
        'cooling': '7200', 'heating': '8000',
        'cooling_power': '2380', 'heating_power': '2580+1500(电辅)',
        'efficiency': 'APF 3.65', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '570×570×265(四面)', 'outdoor_size': '870×650×325',
        'indoor_weight': '20.0', 'noise': '室内≤42/室外≤55',
        'features': '嵌入式天花机、四面均匀出风、变频、WiFi智控、自清洁、商用性价比优选',
        'price': '¥3600-4600',
    },
    {
        'category': '四、3匹空调（冷暖型分体吸顶式/天花机）',
        'brand': 'TCL', 'model': 'KFR-72QW/RT23Bp', 'hp': '3匹', 'type': '冷暖吸顶式',
        'cooling': '7200', 'heating': '7700',
        'cooling_power': '2520', 'heating_power': '2700+1500(电辅)',
        'efficiency': 'APF 3.50', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '570×570×260(四面)', 'outdoor_size': '860×640×320',
        'indoor_weight': '19.0', 'noise': '室内≤43/室外≤56',
        'features': '嵌入式天花机、四面出风、变频、WiFi物联、自清洁、高性价比商用款',
        'price': '¥3700-4700',
    },

    # ========== 五、5匹空调（冷暖型分体吸顶式/天花机）==========
    {
        'category': '五、5匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '美的', 'model': 'KFR-120TW/BP3DN8Y-DH400(B3)', 'hp': '5匹', 'type': '冷暖吸顶式',
        'cooling': '12000', 'heating': '13000',
        'cooling_power': '3900', 'heating_power': '4100+2500(电辅)',
        'efficiency': 'APF 3.40', 'energy': '三级', 'refrigerant': 'R32',
        'indoor_size': '840×840×320(四面)', 'outdoor_size': '960×780×410',
        'indoor_weight': '30.0', 'noise': '室内≤48/室外≤60',
        'features': '大空间天花机、四面出风、变频、WiFi控制、自清洁、商用大堂/会议室适用',
        'price': '¥7500-9000',
    },
    {
        'category': '五、5匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '海尔', 'model': 'KFRd-120Q8W/BY-BD', 'hp': '5匹', 'type': '冷暖吸顶式',
        'cooling': '12000', 'heating': '12800',
        'cooling_power': '3950', 'heating_power': '4200+2500(电辅)',
        'efficiency': 'APF 3.35', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '840×840×315(四面)', 'outdoor_size': '960×780×410',
        'indoor_weight': '29.5', 'noise': '室内≤48/室外≤60',
        'features': '大空间天花机、四面均匀出风、变频、WiFi智能、自清洁、商用大空间适用',
        'price': '¥7200-8800',
    },
    {
        'category': '五、5匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '海信', 'model': 'KFR-120Q6W/E370', 'hp': '5匹', 'type': '冷暖吸顶式',
        'cooling': '12000', 'heating': '12500',
        'cooling_power': '4000', 'heating_power': '4150+2500(电辅)',
        'efficiency': 'APF 3.30', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '840×840×310(四面)', 'outdoor_size': '950×770×400',
        'indoor_weight': '29.0', 'noise': '室内≤47/室外≤59',
        'features': '大空间天花机、四面出风、变频控制、WiFi遥控、自清洁、商用大空间',
        'price': '¥6800-8500',
    },
    {
        'category': '五、5匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '奥克斯', 'model': 'KFR-120QW/BpTYC1', 'hp': '5匹', 'type': '冷暖吸顶式',
        'cooling': '12000', 'heating': '12200',
        'cooling_power': '4100', 'heating_power': '4300+2500(电辅)',
        'efficiency': 'APF 3.25', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '840×840×315(四面)', 'outdoor_size': '950×770×400',
        'indoor_weight': '28.5', 'noise': '室内≤49/室外≤60',
        'features': '大空间天花机、四面出风、变频、WiFi控制、自动清洁、高性价比大匹数商用',
        'price': '¥5800-7500',
    },
    {
        'category': '五、5匹空调（冷暖型分体吸顶式/天花机）',
        'brand': '华凌', 'model': 'KFR-120TW/BP3N8Y-DH400', 'hp': '5匹', 'type': '冷暖吸顶式',
        'cooling': '12000', 'heating': '12800',
        'cooling_power': '3950', 'heating_power': '4150+2500(电辅)',
        'efficiency': 'APF 3.35', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '840×840×315(四面)', 'outdoor_size': '950×770×400',
        'indoor_weight': '29.0', 'noise': '室内≤48/室外≤59',
        'features': '大空间天花机、四面均匀出风、变频、WiFi智控、自清洁、大匹数性价比',
        'price': '¥5500-7200',
    },
    {
        'category': '五、5匹空调（冷暖型分体吸顶式/天花机）',
        'brand': 'TCL', 'model': 'KFR-120QW/RT23Bp', 'hp': '5匹', 'type': '冷暖吸顶式',
        'cooling': '12000', 'heating': '12000',
        'cooling_power': '4150', 'heating_power': '4400+2500(电辅)',
        'efficiency': 'APF 3.20', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '840×840×310(四面)', 'outdoor_size': '950×770×400',
        'indoor_weight': '28.0', 'noise': '室内≤49/室外≤60',
        'features': '大空间天花机、四面出风、变频、WiFi物联、自清洁、大匹数商用款',
        'price': '¥5600-7300',
    },

    # ========== 六、2匹空调（冷暖型分体风管式）==========
    {
        'category': '六、2匹空调（冷暖型分体风管式）',
        'brand': '美的', 'model': 'KFR-50T2W/BP3DN8Y-GC(B3)', 'hp': '2匹', 'type': '冷暖风管式',
        'cooling': '5000', 'heating': '5600',
        'cooling_power': '1580', 'heating_power': '1750+1000(电辅)',
        'efficiency': 'APF 3.60', 'energy': '三级', 'refrigerant': 'R32',
        'indoor_size': '700×450×200(薄型)', 'outdoor_size': '825×595×325',
        'indoor_weight': '16.0', 'noise': '室内≤38/室外≤54',
        'features': '隐藏式风管机、超薄机身、变频、WiFi控制、自清洁、静压可调、吊顶隐藏安装',
        'price': '¥4000-5000',
    },
    {
        'category': '六、2匹空调（冷暖型分体风管式）',
        'brand': '海尔', 'model': 'KFRd-50Q8W/BY-BD', 'hp': '2匹', 'type': '冷暖风管式',
        'cooling': '5000', 'heating': '5500',
        'cooling_power': '1620', 'heating_power': '1800+1000(电辅)',
        'efficiency': 'APF 3.55', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '700×450×195(薄型)', 'outdoor_size': '825×595×320',
        'indoor_weight': '15.5', 'noise': '室内≤38/室外≤54',
        'features': '隐藏式风管机、超薄设计、变频、WiFi智能、自清洁、可调静压、吊顶安装',
        'price': '¥3800-4800',
    },
    {
        'category': '六、2匹空调（冷暖型分体风管式）',
        'brand': '海信', 'model': 'KFR-50Q6W/E370', 'hp': '2匹', 'type': '冷暖风管式',
        'cooling': '5000', 'heating': '5500',
        'cooling_power': '1650', 'heating_power': '1820+1000(电辅)',
        'efficiency': 'APF 3.50', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '700×450×190(薄型)', 'outdoor_size': '820×590×315',
        'indoor_weight': '15.0', 'noise': '室内≤37/室外≤53',
        'features': '隐藏式风管机、超薄机身、变频控制、WiFi遥控、自清洁、抗菌滤网、隐藏安装',
        'price': '¥3500-4500',
    },
    {
        'category': '六、2匹空调（冷暖型分体风管式）',
        'brand': '奥克斯', 'model': 'KFR-50QW/BpTYC1', 'hp': '2匹', 'type': '冷暖风管式',
        'cooling': '5000', 'heating': '5400',
        'cooling_power': '1680', 'heating_power': '1880+1000(电辅)',
        'efficiency': 'APF 3.45', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '700×450×195(薄型)', 'outdoor_size': '815×585×315',
        'indoor_weight': '14.5', 'noise': '室内≤39/室外≤54',
        'features': '隐藏式风管机、薄型设计、变频、WiFi控制、自动清洁、高性价比风管机',
        'price': '¥3000-4000',
    },
    {
        'category': '六、2匹空调（冷暖型分体风管式）',
        'brand': '华凌', 'model': 'KFR-50T2W/BP3N8Y-GC', 'hp': '2匹', 'type': '冷暖风管式',
        'cooling': '5000', 'heating': '5500',
        'cooling_power': '1620', 'heating_power': '1800+1000(电辅)',
        'efficiency': 'APF 3.55', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '700×450×195(薄型)', 'outdoor_size': '820×590×315',
        'indoor_weight': '15.0', 'noise': '室内≤38/室外≤53',
        'features': '隐藏式风管机、超薄机身、变频、WiFi智控、自清洁、性价比风管优选',
        'price': '¥2800-3800',
    },
    {
        'category': '六、2匹空调（冷暖型分体风管式）',
        'brand': 'TCL', 'model': 'KFR-50QW/RT23Bp', 'hp': '2匹', 'type': '冷暖风管式',
        'cooling': '5000', 'heating': '5300',
        'cooling_power': '1720', 'heating_power': '1920+1000(电辅)',
        'efficiency': 'APF 3.40', 'energy': '三级', 'refrigerant': 'R410A',
        'indoor_size': '700×450×190(薄型)', 'outdoor_size': '815×585×310',
        'indoor_weight': '14.0', 'noise': '室内≤39/室外≤54',
        'features': '隐藏式风管机、薄型设计、变频、WiFi物联、自清洁、高性价比隐藏安装',
        'price': '¥2900-3900',
    },
]

# ============================================================
# 创建工作表
# ============================================================
ws = wb.active
ws.title = '空调技术参数对比表'

# 标题行
ws.merge_cells('A1:R1')
ws['A1'].value = '各品牌空调技术参数对比表（含参考价格）— 美的 / 海尔 / 海信 / 奥克斯 / 华凌 / TCL'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('A2:R2')
ws['A2'].value = f'生成日期：{datetime.now().strftime("%Y年%m月%d日")}  |  价格仅供参考，实际价格因渠道、促销、地区而异，请以购买时为准'
ws['A2'].font = Font(name='微软雅黑', size=9, color='666666', italic=True)
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

# 表头
headers = [
    '品牌', '产品型号', '匹数', '类型',
    '制冷量(W)', '制热量(W)',
    '制冷功率(W)', '制热功率(W)',
    '能效指标', '能效等级', '制冷剂',
    '室内机尺寸(mm)', '室外机尺寸(mm)',
    '室内机重量(kg)', '运行噪音(dB)',
    '特色功能', '参考价格(元)',
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_align
    cell.border = thin_border

# 写入数据
row_num = 5
prev_category = None
data_row_index = 0

for p in products:
    if p['category'] != prev_category:
        ws.merge_cells(f'A{row_num}:Q{row_num}')
        cat_cell = ws.cell(row=row_num, column=1, value=p['category'])
        cat_cell.font = subtitle_font
        cat_cell.fill = cat_fill
        cat_cell.alignment = Alignment(horizontal='left', vertical='center')
        cat_cell.border = thin_border
        for c in range(2, 18):
            ws.cell(row=row_num, column=c).fill = cat_fill
            ws.cell(row=row_num, column=c).border = thin_border
        row_num += 1
        prev_category = p['category']
        data_row_index = 0

    bc = brand_colors.get(p['brand'], '999999')
    row_data = [
        p['brand'], p['model'], p['hp'], p['type'],
        p['cooling'], p['heating'],
        p['cooling_power'], p['heating_power'],
        p['efficiency'], p['energy'], p['refrigerant'],
        p['indoor_size'], p['outdoor_size'],
        p['indoor_weight'], p['noise'],
        p['features'], p['price']
    ]

    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.border = thin_border

        if col == 1:  # 品牌
            cell.font = Font(name='微软雅黑', bold=True, size=9, color=brand_font_color)
            cell.fill = PatternFill(start_color=bc, end_color=bc, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col == 2:  # 型号
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col == 16:  # 特色功能
            cell.font = data_font
            cell.alignment = left_align
        elif col == 17:  # 价格
            cell.font = Font(name='微软雅黑', bold=True, size=10, color='C0392B')
            cell.fill = price_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.font = data_font
            cell.alignment = wrap_align

        # 交替行
        if col not in (1, 17):
            cell.fill = alt_fill_1 if data_row_index % 2 == 0 else alt_fill_2

    row_num += 1
    data_row_index += 1

# ============================================================
# 列宽 & 行高
# ============================================================
col_widths = {
    'A': 8, 'B': 32, 'C': 7, 'D': 13,
    'E': 18, 'F': 18, 'G': 12, 'H': 20,
    'I': 10, 'J': 10, 'K': 9,
    'L': 22, 'M': 20, 'N': 12, 'O': 16,
    'P': 52, 'Q': 14,
}
for cl, w in col_widths.items():
    ws.column_dimensions[cl].width = w

for r in range(1, row_num + 1):
    ws.row_dimensions[r].height = 26
ws.row_dimensions[1].height = 38
ws.row_dimensions[2].height = 20

# 打印设置
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.paperSize = ws.PAPERSIZE_A3

# ============================================================
# 保存
# ============================================================
output_path = '/home/admin/.openclaw/workspace/空调技术参数对比表(含价格).xlsx'
wb.save(output_path)
print(f'✅ Excel 文件已生成: {output_path}')
print(f'   共 {len(products)} 条产品数据')
