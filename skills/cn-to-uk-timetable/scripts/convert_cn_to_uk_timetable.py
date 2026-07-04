#!/usr/bin/env python3
"""
Redraw a UK-style Teaching Calendar from a single Chinese course timetable.

This version does NOT copy or reuse an English template sheet. It creates a new
workbook from scratch and draws a timetable layout that mimics the UK format:
title row, headers, Morning/Afternoon/Evening blocks, break rows, borders,
merged cells, row heights, column widths, and color-coded scheduled blocks.

For mobile WeChat preview, prefer outputting .xls. WeChat tends to preserve
legacy Excel 97-2003 indexed fills better than modern .xlsx fills.

Typical usage:
    python scripts/convert_cn_to_uk_timetable.py \
      --source-file "AII040-数字信号处理.xls" \
      --output "AII040_数字信号处理_英方课表.xls" \
      --log "AII040_数字信号处理_转换日志.json"
"""

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, asdict, replace
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Missing dependency: openpyxl. Install with `pip install openpyxl`.") from exc

SKILL_VERSION = "v37_reference_merge_contiguous_weeks_same_slot"

DAY_BY_COL = {
    3: "Monday",
    4: "Tuesday",
    5: "Wednesday",
    6: "Thursday",
    7: "Friday",
    8: "Saturday",
    9: "Sunday",
}

DEFAULT_PERIOD_BY_ROW = {
    3: (1, 2),
    4: (3, 4),
    5: (6, 7),
    6: (8, 9),
    7: (10, 11),
}

CLASS_TO_ROW = {
    1: 3,
    2: 4,
    3: 5,
    4: 6,
    5: 7,
    6: 9,
    7: 10,
    8: 11,
    9: 12,
    10: 14,
    11: 15,
    12: 16,
}

CLASS_TIMES = {
    1: "8: 05-8: 50",
    2: "8: 50-9: 35",
    3: "9: 50-10: 35",
    4: "10: 40-11: 25",
    5: "11: 30-12: 15",
    6: "13: 40-14: 25",
    7: "14: 35-15: 20",
    8: "15: 30-16: 15",
    9: "16: 25-17: 10",
    10: "18: 30-19: 15",
    11: "19: 25-20: 10",
    12: "20: 20-21: 05",
}

# Days are laid out dynamically. Weekdays keep at least two subcolumns,
# because English timetable samples often use split day columns. If several
# events occur at the same time in different rooms, extra subcolumns are added
# so they appear as separate cells instead of being appended into one cell.
DAY_SEQUENCE = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
DEFAULT_DAY_MIN_LANES = {
    "Monday": 2,
    "Tuesday": 2,
    "Wednesday": 2,
    "Thursday": 2,
    "Friday": 2,
    "Saturday": 1,
    "Sunday": 1,
}

DEFAULT_MAPPING = {
    "course_sheet_map": {},
    "course_title_map": {},
    "type_overrides": {},
}

FILL_TITLE = PatternFill(fill_type="solid", start_color="FFA6A6A6", end_color="FFA6A6A6")
FILL_BREAK = PatternFill(fill_type="solid", start_color="FFA6A6A6", end_color="FFA6A6A6")
# Colors are matched to the provided English/WPS sample instead of using the
# over-bright default palette. These are opaque ARGB values so desktop WPS,
# Excel and WeChat preview have the best chance to preserve fills.
FILL_LECTURE = PatternFill(fill_type="solid", start_color="FFFFC000", end_color="FFFFC000")  # warm yellow/orange
FILL_LAB = PatternFill(fill_type="solid", start_color="FF92D050", end_color="FF92D050")      # sample light green
FILL_SPECIAL = PatternFill(fill_type="solid", start_color="FF9DC3E6", end_color="FF9DC3E6")  # sample soft blue
FILL_EMPTY = PatternFill(fill_type="solid", start_color="FFFFFFFF", end_color="FFFFFFFF")
# In mobile WeChat .xls preview, truly transparent/no-fill cells may render as black.
# Use solid white for blank timetable cells: visually blank, but not black.
FILL_BLANK = FILL_EMPTY

BORDER_THIN = Side(style="thin", color="FF000000")
BORDER_TABLE = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THIN)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
FONT_BODY = Font(name="Times New Roman", size=14)
# v24: keep v17 horizontal blank-cell merging layout, Times New Roman fonts,
# fixed requested row heights, and center alignment in every timetable cell.
FONT_BLOCK = Font(name="Times New Roman", size=14)
FONT_HEADER = Font(name="Times New Roman", size=14, bold=True)
FONT_TITLE = Font(name="Times New Roman", size=20, bold=True)


def period_span_to_display_hours(start_period: int, end_period: int) -> int:
    """Convert Chinese timetable period span to displayed UK hours.

    User-corrected v31 rule:
    - A one-period record is shown as 1h.
    - A normal two-period record is shown as 2h.
    - Three or more consecutive periods must be judged by actual clock time.
      For example, periods 6-8 run from 13:40 to 16:15, which is only
      2h35m, so it must display 2h, not 3h. Periods 6-9 run from
      13:40 to 17:10, which is 3h30m, so it displays 3h.
    In short, never round a 2h35m block up to 3h. Use completed natural
    hours for 3+ period spans, while preserving the standard 1-period/
    2-period teaching-hour labels.
    """
    periods = max(1, end_period - start_period + 1)
    if periods <= 2:
        return periods
    elapsed_minutes = _clock_minutes_between_periods(start_period, end_period)
    return max(2, int(elapsed_minutes // 60))


@dataclass
class CourseEvent:
    course_code: str
    course_name: str
    source_file: str
    raw: str
    teacher: str
    day: str
    start_period: int
    end_period: int
    weeks: str
    location: str
    room_text: str
    class_text: str
    section_code: str
    class_type: str
    explicit_period: bool
    # display_hours is the original Chinese teaching-hour count parsed from the
    # source timetable. It must stay unchanged even when the visual block is
    # extended to cover enough clock time in the UK-style timetable.
    display_hours: Optional[int] = None

    @property
    def duration(self) -> int:
        return max(1, self.display_hours if self.display_hours is not None else self.end_period - self.start_period + 1)

    def to_uk_cell_text(self) -> str:
        lines = [f"Week {self.weeks}", f"{self.duration}h {self.class_type}"]
        if self.class_text:
            if self.class_type.lower() == "lab":
                lines.append(self.class_text)
            else:
                lines.append(f"Class: {self.class_text}")
        if self.room_text:
            lines.append(self.room_text)
        return "\n".join(lines)


class ConversionError(RuntimeError):
    pass


def estimate_text_tokens(text: str) -> int:
    """Roughly estimate LLM tokens for mixed Chinese/English timetable text.

    This is intentionally dependency-free. It is not exact API accounting; it
    gives the user a practical per-course estimate when one Chinese timetable
    file is sent to an agent and summarized/rewritten into the UK format.
    """
    if not text:
        return 0
    cjk = len(re.findall(r"[\u4e00-\u9fff]", text))
    non_ws = len(re.sub(r"\s+", "", text))
    non_cjk = max(0, non_ws - cjk)
    # Chinese is often close to 1 char/token; English/numbers are roughly 4 chars/token.
    return int(math.ceil(cjk * 1.05 + non_cjk / 4.0))


def estimate_conversion_tokens(
    source_file: Path,
    course_code: str,
    course_name: str,
    events: List[CourseEvent],
    warnings: List[dict],
) -> dict:
    """Return a visible, conservative token estimate for one-course conversion.

    The script itself does not call an LLM, so there is no exact usage bill here.
    The estimate reflects the tokens an LLM-style agent would likely consume for
    the extracted course content, conversion instructions, generated course
    blocks, and a small amount of bookkeeping/log text.
    """
    source_text_parts = [course_code or "", course_name or ""]
    source_text_parts.extend(e.raw for e in events if e.raw)
    source_text_parts.extend(str(w.get("raw", "")) for w in warnings if isinstance(w, dict))
    source_text = "\n".join(source_text_parts)

    output_text_parts = []
    for event in events:
        output_text_parts.append(event.to_uk_cell_text())
    output_text = "\n".join(output_text_parts)

    instruction_overhead_tokens = 900  # skill rules, layout rules, color rules, logging request
    spreadsheet_overhead_tokens = 350  # title/headers/row labels/summary metadata
    input_tokens = estimate_text_tokens(source_text) + instruction_overhead_tokens
    output_tokens = estimate_text_tokens(output_text) + spreadsheet_overhead_tokens
    total_tokens = input_tokens + output_tokens

    try:
        size_bytes = source_file.stat().st_size
    except OSError:
        size_bytes = None

    return {
        "note": "估算值，不是精确 API 计费；脚本本身不调用大模型。用于判断上传一门中方课表给 Agent 转换时的大致 token 消耗。",
        "method": "mixed_zh_en_chars_estimate_plus_skill_overhead",
        "source_file_size_bytes": size_bytes,
        "source_text_chars": len(source_text),
        "generated_cell_text_chars": len(output_text),
        "event_count": len(events),
        "warning_count": len(warnings),
        "estimated_input_tokens": int(input_tokens),
        "estimated_output_tokens": int(output_tokens),
        "estimated_total_tokens": int(total_tokens),
        "human_readable": f"约 {total_tokens:,} tokens / 门课",
    }


def find_soffice() -> Optional[str]:
    for name in ("libreoffice", "soffice"):
        path = shutil.which(name)
        if path:
            return path
    return None


def convert_to_xlsx(path: Path, workdir: Path) -> Path:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return path
    if suffix != ".xls":
        raise ConversionError(f"Unsupported spreadsheet type: {path}")
    soffice = find_soffice()
    if not soffice:
        raise ConversionError(
            f"Cannot convert {path.name}: LibreOffice/soffice not found. "
            "Please save .xls as .xlsx first or install LibreOffice."
        )
    outdir = workdir / "xlsx"
    outdir.mkdir(parents=True, exist_ok=True)
    result = subprocess.run(
        [soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(outdir), str(path)],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )
    if result.returncode != 0:
        raise ConversionError(f"LibreOffice conversion failed for {path}: {result.stderr or result.stdout}")
    converted = outdir / (path.stem + ".xlsx")
    if not converted.exists():
        matches = list(outdir.glob(path.stem + "*.xlsx"))
        if matches:
            converted = matches[0]
    if not converted.exists():
        raise ConversionError(f"LibreOffice did not create .xlsx for {path}")
    return converted


def load_mapping(mapping_path: Optional[Path]) -> dict:
    mapping = json.loads(json.dumps(DEFAULT_MAPPING))
    if mapping_path and mapping_path.exists():
        with mapping_path.open("r", encoding="utf-8") as f:
            user_mapping = json.load(f)
        for key, value in user_mapping.items():
            if isinstance(value, dict) and isinstance(mapping.get(key), dict):
                mapping[key].update(value)
            else:
                mapping[key] = value
    return mapping


def clean_course_name(text: Optional[str]) -> str:
    if not text:
        return ""
    s = str(text).strip()
    s = re.sub(r"^\d{4}-\d{4}学年第\d学期", "", s)
    s = s.replace("课程课表", "").replace("课表", "")
    return s.strip()


def normalize_weeks(raw: str) -> str:
    s = raw.strip()
    s = re.sub(r"^\(?\d+\s*-\s*\d+节\)?", "", s)
    s = s.replace("（", "(").replace("）", ")")
    s = s.replace("，", ",").replace(" ", "")
    s = s.replace("周", "")

    def expand_odd_even(match: re.Match[str]) -> str:
        a, b, kind = int(match.group(1)), int(match.group(2)), match.group(3)
        want_odd = kind == "单"
        nums = [n for n in range(a, b + 1) if (n % 2 == 1) == want_odd]
        return ",".join(map(str, nums))

    s = re.sub(r"(\d+)\s*-\s*(\d+)\((单|双)\)", expand_odd_even, s)
    s = s.replace("(单)", " odd").replace("(双)", " even")
    s = re.sub(r",+", ",", s).strip(",")
    return s or raw.strip()


def convert_class_name(token: str) -> str:
    token = token.strip()
    m = re.match(r"AI电子(\d{2})(\d{2})", token)
    if m:
        return f"REE {m.group(1)}{m.group(2)}"
    m = re.match(r"AI通信(\d{2})(\d{2})", token)
    if m:
        return f"CE {m.group(1)}{m.group(2)}"
    return token


def normalize_classes(raw: str) -> str:
    if not raw:
        return ""
    parts = [convert_class_name(x) for x in re.split(r"[;；,，]", raw) if x.strip()]
    grouped: Dict[str, List[str]] = {}
    leftovers: List[str] = []
    for p in parts:
        m = re.match(r"^(REE|CE)\s+(\d{4})$", p)
        if m:
            grouped.setdefault(m.group(1), []).append(m.group(2))
        else:
            leftovers.append(p)
    lines: List[str] = []
    for prefix in ("REE", "CE"):
        if prefix in grouped:
            nums = grouped[prefix]
            lines.append(prefix + " " + ", ".join(nums))
    lines.extend(leftovers)
    return "\n".join(lines)


def normalize_room(raw: str) -> str:
    if not raw:
        return ""
    s = str(raw).strip()
    s = s.replace("下沙", "").strip()
    s = re.sub(r"\s+", " ", s)
    m = re.search(r"萨塞克斯\s*(\d+)", s)
    if m:
        return f"Building Lab, Room {m.group(1)}"
    m = re.search(r"信电楼电路\s*(\d+)", s)
    if m:
        return f"Building SIEE, Room {m.group(1)}"
    m = re.search(r"\b([A-Za-z])\s*-?\s*(\d{2,4})\b", s)
    if m:
        return f"Building {m.group(1).upper()}, Room {m.group(2)}"
    m = re.search(r"([A-Za-z])(\d{2,4})", s)
    if m:
        return f"Building {m.group(1).upper()}, Room {m.group(2)}"
    return s


def parse_period_and_week(fields: List[str], source_row: int) -> Tuple[int, int, str, bool]:
    default_start, default_end = DEFAULT_PERIOD_BY_ROW.get(source_row, (1, 2))
    week_field = fields[1].strip() if len(fields) > 1 else ""
    m = re.match(r"^\(?\s*(\d+)\s*-\s*(\d+)节\s*\)?\s*(.*)$", week_field)
    if m:
        return int(m.group(1)), int(m.group(2)), normalize_weeks(m.group(3)), True
    return default_start, default_end, normalize_weeks(week_field), False


def extract_course_code(value: Optional[str], filename: str = "") -> str:
    text = f"{value or ''} {filename}"
    m = re.search(r"AII\d{3}", text, re.I)
    return m.group(0).upper() if m else ""



def normalize_teacher(raw: str) -> str:
    """Normalize the teacher name from the first field of a Chinese timetable record."""
    s = str(raw or "").strip()
    s = re.sub(r"\s+", "", s)
    for _ in range(3):
        new_s = re.sub(r"(教授|副教授|讲师|助教|实验师|高级实验师|无)$", "", s)
        if new_s == s or not new_s:
            break
        s = new_s
    return s or str(raw or "").strip()


def _ordered_unique(items: List[str]) -> List[str]:
    seen = set()
    out: List[str] = []
    for item in items:
        item = str(item or "").strip()
        if not item:
            continue
        key = _norm_key_text(item)
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def _weeks_to_numbers(value: str) -> Optional[List[int]]:
    """Parse normalized week text into sorted week numbers when possible.

    Supports values like ``2-3``, ``4``, ``2-3,4`` or ``13,14,15``.
    If the value contains non-numeric qualifiers that cannot be represented as
    exact week numbers, return None so callers can keep the original text.
    """
    text = str(value or "").strip().replace("周", "").replace("，", ",").replace(" ", "")
    if not text:
        return []
    nums = set()
    for part in [p for p in text.split(",") if p]:
        m_range = re.fullmatch(r"(\d+)\s*-\s*(\d+)", part)
        if m_range:
            a, b = int(m_range.group(1)), int(m_range.group(2))
            if a > b:
                a, b = b, a
            nums.update(range(a, b + 1))
            continue
        m_num = re.fullmatch(r"\d+", part)
        if m_num:
            nums.add(int(part))
            continue
        return None
    return sorted(nums)


def _format_week_numbers(nums: List[int]) -> str:
    nums = sorted(set(nums))
    if not nums:
        return ""
    ranges = []
    start = prev = nums[0]
    for n in nums[1:]:
        if n == prev + 1:
            prev = n
            continue
        ranges.append(f"{start}-{prev}" if start != prev else str(start))
        start = prev = n
    ranges.append(f"{start}-{prev}" if start != prev else str(start))
    return ",".join(ranges)


def _weeks_are_contiguous_or_overlapping(a: str, b: str) -> bool:
    nums_a = _weeks_to_numbers(a)
    nums_b = _weeks_to_numbers(b)
    if nums_a is None or nums_b is None or not nums_a or not nums_b:
        return False
    combined = sorted(set(nums_a) | set(nums_b))
    return combined[-1] - combined[0] + 1 == len(combined)


def combine_weeks(values: List[str]) -> str:
    parsed_all: List[int] = []
    all_parseable = True
    for v in values:
        nums = _weeks_to_numbers(str(v or ""))
        if nums is None:
            all_parseable = False
            break
        parsed_all.extend(nums)
    if all_parseable and parsed_all:
        return _format_week_numbers(parsed_all)

    tokens: List[str] = []
    for v in values:
        for part in re.split(r"[,，\n]+", str(v or "")):
            part = part.strip().replace("周", "")
            if part:
                tokens.append(part)
    return ",".join(_ordered_unique(tokens))


def combine_class_texts(values: List[str]) -> str:
    grouped: Dict[str, List[str]] = {}
    leftovers: List[str] = []
    for v in values:
        for line in str(v or "").splitlines():
            line = line.strip()
            if not line:
                continue
            m = re.match(r"^(REE|CE)\s+(.+)$", line)
            if m:
                prefix = m.group(1)
                nums = [x.strip() for x in re.split(r"[,，]", m.group(2)) if x.strip()]
                grouped.setdefault(prefix, []).extend(nums)
            else:
                leftovers.append(line)
    lines: List[str] = []
    for prefix in ("REE", "CE"):
        if prefix in grouped:
            lines.append(prefix + " " + ", ".join(_ordered_unique(grouped[prefix])))
    lines.extend(_ordered_unique(leftovers))
    return "\n".join(lines)


def combine_room_texts(values: List[str]) -> str:
    rooms_by_building: Dict[str, List[str]] = {}
    leftovers: List[str] = []
    for v in values:
        for line in str(v or "").splitlines():
            line = line.strip()
            if not line:
                continue
            m = re.match(r"^Building\s+(.+?),\s*Room\s+(.+)$", line)
            if m:
                building = m.group(1).strip()
                nums = [x.strip() for x in re.split(r"[,，]", m.group(2)) if x.strip()]
                rooms_by_building.setdefault(building, []).extend(nums)
            else:
                leftovers.append(line)
    lines: List[str] = []
    for building, nums in rooms_by_building.items():
        lines.append(f"Building {building}, Room " + ", ".join(_ordered_unique(nums)))
    lines.extend(_ordered_unique(leftovers))
    return "\n".join(lines)

def determine_class_type(course_code: str, section_code: str, explicit_period: bool, mapping: dict) -> str:
    section_tail = section_code.split("-")[-1] if section_code else ""
    overrides = mapping.get("type_overrides", {}).get(course_code, {})
    if not explicit_period and overrides.get("implicit_base"):
        return overrides["implicit_base"]
    if explicit_period and overrides.get("explicit_base"):
        return overrides["explicit_base"]
    if re.search(r"[A-Za-z]$", section_tail):
        return "Lab"
    if overrides.get("base"):
        return overrides["base"]
    return "Lecture"


def parse_event_line(
    line: str,
    source_file: Path,
    course_code: str,
    course_name: str,
    day: str,
    source_row: int,
    mapping: dict,
) -> Optional[CourseEvent]:
    raw = line.strip()
    if not raw:
        return None
    fields = [f.strip() for f in raw.split("/")]
    if len(fields) < 5:
        return None
    start, end, weeks, explicit = parse_period_and_week(fields, source_row)
    location = fields[2].strip() if len(fields) > 2 else ""
    section_code = fields[3].strip() if len(fields) > 3 else ""
    classes = fields[4].strip() if len(fields) > 4 else ""
    class_type = determine_class_type(course_code, section_code, explicit, mapping)
    teacher = normalize_teacher(fields[0] if fields else "")
    return CourseEvent(
        course_code=course_code,
        course_name=course_name,
        source_file=str(source_file),
        raw=raw,
        teacher=teacher,
        day=day,
        start_period=start,
        end_period=end,
        weeks=weeks,
        location=location,
        room_text=normalize_room(location),
        class_text=normalize_classes(classes),
        section_code=section_code,
        class_type=class_type,
        explicit_period=explicit,
        display_hours=period_span_to_display_hours(start, end),
    )


def parse_chinese_workbook(path: Path, mapping: dict, workdir: Path) -> Tuple[str, str, List[CourseEvent], List[dict]]:
    xlsx = convert_to_xlsx(path, workdir)
    wb = load_workbook(xlsx, data_only=True)
    ws = wb.active
    course_name = clean_course_name(ws.cell(1, 4).value) or clean_course_name(path.stem)
    course_code = extract_course_code(ws.cell(1, 8).value, path.name)
    events: List[CourseEvent] = []
    warnings: List[dict] = []
    for row in range(3, 8):
        for col in range(3, 10):
            value = ws.cell(row, col).value
            if not value:
                continue
            day = DAY_BY_COL.get(col)
            if not day:
                continue
            for line in str(value).splitlines():
                if not line.strip():
                    continue
                event = parse_event_line(line, path, course_code, course_name, day, row, mapping)
                if event:
                    events.append(event)
                else:
                    warnings.append({
                        "file": str(path),
                        "course": course_name,
                        "cell": f"{ws.cell(row, col).coordinate}",
                        "raw": line,
                        "warning": "Could not parse event line",
                    })
    return course_code, course_name, events, warnings


def safe_sheet_title(course_code: str, course_name: str, mapping: dict) -> str:
    mapped = mapping.get("course_sheet_map", {}).get(course_name)
    base = mapped or f"{course_code} {course_name}".strip() or "Converted Course"
    base = re.sub(r"[\\/*?:\[\]]", " ", base)
    base = re.sub(r"\s+", " ", base).strip()
    return (base[:31] or "Converted Course")


def display_title(course_code: str, course_name: str, mapping: dict) -> str:
    return mapping.get("course_title_map", {}).get(course_code) or f"{course_code} {course_name}".strip() or "Converted Course"




def _norm_key_text(value: str) -> str:
    """Normalize text used for visual de-duplication.

    Different punctuation/spacing from parsing should not create duplicate cells
    when the displayed meaning is the same.
    """
    value = str(value or "").strip()
    value = value.replace("，", ",").replace("；", ";").replace("：", ":")
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"\s*,\s*", ",", value)
    return value.lower()


def event_layout_key(event: CourseEvent) -> Tuple[str, int, int, str, str, str, str]:
    """Key for merging truly identical visible course blocks.

    If two source records would display the same text in the same day/time slot,
    they must become one merged visual block rather than two separate cells.
    Different rooms, different classes, different weeks, or different class
    types intentionally remain separate lanes.
    """
    return (
        event.day,
        _norm_key_text(event.teacher),
        event.start_period,
        event.end_period,
        _norm_key_text(event.weeks),
        _norm_key_text(event.class_type),
        _norm_key_text(event.class_text),
        _norm_key_text(event.room_text),
    )


def collapse_identical_events(events: List[CourseEvent]) -> List[CourseEvent]:
    """Remove duplicate visible course blocks while preserving order.

    The source Chinese timetable sometimes contains repeated entries for the
    same week/type/class/room/time. In the UK-style output, these should be one
    block. This is different from same-time different-room events, which must be
    split into separate cells.
    """
    seen = set()
    collapsed: List[CourseEvent] = []
    for event in events:
        key = event_layout_key(event)
        if key in seen:
            continue
        seen.add(key)
        collapsed.append(event)
    return collapsed


def _mergeable_contiguous_key(event: CourseEvent) -> Tuple[str, str, str, str, str]:
    """Key for merging adjacent periods of the same visible course.

    If the same day has adjacent blocks with the same weeks/type/classes/room,
    they should be one longer UK timetable block rather than separate cells.
    This is the behavior shown in the user's reference sheet. Different rooms,
    classes, weeks or teaching types remain separate cells.
    """
    return (
        event.day,
        _norm_key_text(event.teacher),
        _norm_key_text(event.weeks),
        _norm_key_text(event.class_type),
        _norm_key_text(event.class_text),
        _norm_key_text(event.room_text),
    )


def _same_teacher_base_key(event: CourseEvent) -> Tuple[str, int, int, str, str, str]:
    """Base key for same-teacher same-slot merging.

    v36 correction: match the official English table behavior.  First merge
    contiguous records for the same week/class/room; only then merge same-time
    records by teacher.  Same-room/different-class records may be merged only
    when they also share the same week range.  If week ranges differ, keep
    them as separate blocks, such as Week 13 / Week 14 / Week 15 lab groups.
    """
    return (
        event.day,
        event.start_period,
        event.end_period,
        _norm_key_text(event.teacher),
        _norm_key_text(event.class_type),
        _norm_key_text(event.weeks),
    )


def _same_teacher_room_or_class_mergeable(a: CourseEvent, b: CourseEvent) -> bool:
    """Return True when two same-teacher same-slot records should be one block.

    v36 correction:
    - Same teacher + same exact time + same week + same room may be merged even
      when the class/student group differs.
    - Same teacher + same exact time + same week + same class may still merge
      across rooms.
    - Different time ranges or different week ranges are excluded by the base key.
    - Different room AND different class remain separate cells.
    """
    same_room = bool(_norm_key_text(a.room_text)) and _norm_key_text(a.room_text) == _norm_key_text(b.room_text)
    same_class = bool(_norm_key_text(a.class_text)) and _norm_key_text(a.class_text) == _norm_key_text(b.class_text)
    return same_room or same_class


def _same_teacher_components(group: List[CourseEvent]) -> List[List[CourseEvent]]:
    """Split a same-teacher/time/type group into mergeable components.

    Components are connected by either same room or same class.  This allows
    same-room/different-class records to merge, while still preventing unrelated
    different-room/different-class records from collapsing into one cell.
    """
    components: List[List[CourseEvent]] = []
    for event in group:
        target_indexes = []
        for idx, comp in enumerate(components):
            if any(_same_teacher_room_or_class_mergeable(event, other) for other in comp):
                target_indexes.append(idx)
        if not target_indexes:
            components.append([event])
            continue
        first = target_indexes[0]
        components[first].append(event)
        # If this event bridges multiple components, merge those components too.
        for idx in reversed(target_indexes[1:]):
            components[first].extend(components[idx])
            del components[idx]
    return components


def _merge_teacher_slot_group(group: List[CourseEvent]) -> CourseEvent:
    base = group[0]
    return replace(
        base,
        weeks=combine_weeks([e.weeks for e in group]),
        class_text=combine_class_texts([e.class_text for e in group]),
        room_text=combine_room_texts([e.room_text for e in group]),
        raw="\n".join(e.raw for e in group if e.raw),
        explicit_period=any(e.explicit_period for e in group),
        display_hours=period_span_to_display_hours(base.start_period, base.end_period),
    )




def _merge_adjacent_weeks_key(event: CourseEvent) -> Tuple[str, int, int, str, str, str, str]:
    """Key for merging same-slot records whose only difference is week range.

    v37 rule: if same day/time/teacher/type/class/room records differ only by
    week, and those week ranges are contiguous or overlapping, display them as
    one block with a compressed Week range, e.g. Week 2-3 + Week 4 -> Week 2-4.
    Different class, different room, different time or different teacher remain
    separate blocks.
    """
    return (
        event.day,
        event.start_period,
        event.end_period,
        _norm_key_text(event.teacher),
        _norm_key_text(event.class_type),
        _norm_key_text(event.class_text),
        _norm_key_text(event.room_text),
    )


def merge_contiguous_weeks_same_slot(events: List[CourseEvent]) -> List[CourseEvent]:
    """Merge adjacent/overlapping week ranges for otherwise identical slots.

    This fixes cases such as Week 2-3 and Week 4 for the same teacher, room,
    class and time being rendered as two side-by-side cells. They should become
    one cell labelled Week 2-4. The merge is deliberately conservative: if the
    class or room differs, or if the weeks are not contiguous, keep them split.
    """
    grouped: Dict[Tuple[str, int, int, str, str, str, str], List[CourseEvent]] = {}
    order: List[Tuple[str, int, int, str, str, str, str]] = []
    for event in events:
        key = _merge_adjacent_weeks_key(event)
        if key not in grouped:
            grouped[key] = []
            order.append(key)
        grouped[key].append(event)

    merged: List[CourseEvent] = []
    for key in order:
        group = grouped[key]
        parseable = []
        non_parseable = []
        for event in group:
            nums = _weeks_to_numbers(event.weeks)
            if nums is None or not nums:
                non_parseable.append(event)
            else:
                parseable.append((min(nums), max(nums), event, set(nums)))
        parseable.sort(key=lambda x: (x[0], x[1], x[2].weeks))

        current_event: Optional[CourseEvent] = None
        current_nums: set[int] = set()
        current_raws: List[str] = []
        for _min_w, _max_w, event, nums in parseable:
            if current_event is None:
                current_event = event
                current_nums = set(nums)
                current_raws = [event.raw]
                continue
            combined = sorted(current_nums | nums)
            contiguous = bool(combined) and (combined[-1] - combined[0] + 1 == len(combined))
            if contiguous:
                current_nums |= nums
                current_raws.append(event.raw)
                current_event = replace(
                    current_event,
                    weeks=_format_week_numbers(list(current_nums)),
                    raw="\n".join(current_raws),
                    explicit_period=current_event.explicit_period or event.explicit_period,
                    display_hours=period_span_to_display_hours(current_event.start_period, current_event.end_period),
                )
            else:
                merged.append(current_event)
                current_event = event
                current_nums = set(nums)
                current_raws = [event.raw]
        if current_event is not None:
            merged.append(current_event)
        merged.extend(non_parseable)

    day_order = {d: i for i, d in enumerate(DAY_SEQUENCE)}
    return sorted(merged, key=lambda e: (day_order.get(e.day, 99), e.start_period, e.end_period, e.teacher, e.weeks, e.class_type, e.class_text, e.room_text))

def merge_same_teacher_same_time(events: List[CourseEvent]) -> List[CourseEvent]:
    """Merge safe same-teacher records before layout.

    v36 rule: same teacher records may merge only within the same day, exact
    original period range, same week range, and same teaching type.  Inside
    that slot, merge records that share the same room even if the class/student
    group differs; also allow same class/student group to merge across rooms.
    Records with different time ranges or different week ranges remain separate.
    """
    grouped: Dict[Tuple[str, int, int, str, str, str], List[CourseEvent]] = {}
    order: List[Tuple[str, int, int, str, str, str]] = []
    for event in events:
        key = _same_teacher_base_key(event)
        if key not in grouped:
            grouped[key] = []
            order.append(key)
        grouped[key].append(event)

    merged: List[CourseEvent] = []
    for key in order:
        components = _same_teacher_components(grouped[key])
        for comp in components:
            merged.append(_merge_teacher_slot_group(comp) if len(comp) > 1 else comp[0])

    day_order = {d: i for i, d in enumerate(DAY_SEQUENCE)}
    return sorted(merged, key=lambda e: (day_order.get(e.day, 99), e.start_period, e.end_period, e.teacher, e.weeks, e.class_type, e.class_text, e.room_text))


def _periods_cross_break(start_period: int, end_period: int) -> bool:
    """Do not merge course blocks across Lunch/Dinner boundaries."""
    # Morning ends at period 5, afternoon starts at 6; afternoon ends at 9, evening starts at 10.
    return (start_period <= 5 < end_period) or (start_period <= 9 < end_period)


def _parse_class_time_minutes(time_text: str) -> Tuple[int, int]:
    """Parse a class time such as '9: 50-10: 35' into minute offsets."""
    nums = [int(x) for x in re.findall(r"\d+", time_text)]
    if len(nums) >= 4:
        return nums[0] * 60 + nums[1], nums[2] * 60 + nums[3]
    raise ValueError(f"Cannot parse class time: {time_text}")


PERIOD_MINUTES = {cls: _parse_class_time_minutes(t) for cls, t in CLASS_TIMES.items()}


def _session_end_period(start_period: int) -> int:
    if start_period <= 5:
        return 5
    if start_period <= 9:
        return 9
    return 12


def _clock_minutes_between_periods(start_period: int, end_period: int) -> int:
    """Elapsed clock minutes from the start of the first period to the end of the last."""
    start_min = PERIOD_MINUTES[start_period][0]
    end_min = PERIOD_MINUTES[end_period][1]
    return max(0, end_min - start_min)


def extend_visual_span_by_clock_time(event: CourseEvent) -> CourseEvent:
    """Extend the *visual* block span when the scheduled periods do not cover the displayed hours.

    Important user rule: do NOT change the Chinese/parsed teaching-hour label.
    Example: a source event on periods 3-4 is still shown as '2h Lecture/Lab',
    but periods 3-4 only cover 95 clock minutes, which is less than two hours.
    Therefore the visible block should extend to period 5 (3-5) so the UK-style
    timetable block visually covers at least two hours of clock time.

    This rule starts from v24 and replaces the abandoned v25 behavior that
    changed the displayed hour number.
    """
    required_hours = event.duration
    # Keep 1h source blocks as-is. The user's correction was specifically about
    # 2h/3h+ blocks such as 3-4 needing to occupy 3-5 visually.
    if required_hours < 2:
        return event
    target_minutes = required_hours * 60
    cap = _session_end_period(event.start_period)
    visual_end = max(event.end_period, event.start_period)
    while visual_end < cap and _clock_minutes_between_periods(event.start_period, visual_end) < target_minutes:
        visual_end += 1
    if visual_end == event.end_period:
        return event
    return replace(event, end_period=visual_end, display_hours=event.duration)


def apply_visual_span_rules(events: List[CourseEvent]) -> List[CourseEvent]:
    return [extend_visual_span_by_clock_time(e) for e in events]


def merge_contiguous_same_events(events: List[CourseEvent]) -> List[CourseEvent]:
    """Merge adjacent same course arrangements into one longer visual block.

    Example: (6-7节) Week 8 Lecture + (8-8节) Week 8 Lecture in the same
    room/classes becomes one visual block, but the displayed hour label is kept
    from the main/source teaching-hour value instead of being summed.
    Example: a 2h arrangement represented across 6-7 plus an extra visual
    continuation row 8 should still display 2h, not 3h.
    """
    grouped: Dict[Tuple[str, str, str, str, str], List[CourseEvent]] = {}
    for event in events:
        grouped.setdefault(_mergeable_contiguous_key(event), []).append(event)

    merged: List[CourseEvent] = []
    for key, group in grouped.items():
        group = sorted(group, key=lambda e: (e.start_period, e.end_period))
        current: Optional[CourseEvent] = None
        raw_parts: List[str] = []
        for event in group:
            if current is None:
                current = event
                raw_parts = [event.raw]
                continue
            can_merge = (
                event.start_period <= current.end_period + 1
                and not _periods_cross_break(current.start_period, max(current.end_period, event.end_period))
            )
            if can_merge:
                current = replace(
                    current,
                    end_period=max(current.end_period, event.end_period),
                    raw="\n".join(raw_parts + [event.raw]),
                    explicit_period=current.explicit_period or event.explicit_period,
                    display_hours=period_span_to_display_hours(current.start_period, max(current.end_period, event.end_period)),
                )
                raw_parts.append(event.raw)
            else:
                merged.append(current)
                current = event
                raw_parts = [event.raw]
        if current is not None:
            merged.append(current)

    day_order = {d: i for i, d in enumerate(DAY_SEQUENCE)}
    return sorted(merged, key=lambda e: (day_order.get(e.day, 99), e.start_period, e.end_period, e.weeks, e.class_type, e.class_text, e.room_text))

def compute_day_columns(events: List[CourseEvent]) -> Dict[str, List[int]]:
    """Return dynamic timetable columns for each day.

    Same-time events in different rooms/classes must be displayed in separate
    cells. We therefore calculate the maximum number of simultaneous events per
    day and allocate that many lanes. Weekdays still keep at least two lanes;
    weekends keep one lane unless conflicts require more.
    """
    lanes_by_day: Dict[str, int] = dict(DEFAULT_DAY_MIN_LANES)
    for day in DAY_SEQUENCE:
        day_events = [e for e in events if e.day == day]
        if not day_events:
            # v27: if the whole day has no scheduled classes, keep only one
            # narrow placeholder column. This avoids a wide blank Tuesday-like
            # area and matches the user's requested 12.66-character width.
            lanes_by_day[day] = 1
            continue
        max_concurrent = 1
        for period in range(1, 13):
            active_keys = {
                event_layout_key(e) for e in day_events
                if e.start_period <= period <= e.end_period
            }
            max_concurrent = max(max_concurrent, len(active_keys))
        lanes_by_day[day] = max(lanes_by_day.get(day, 1), max_concurrent)

    day_cols: Dict[str, List[int]] = {}
    col = 4
    for day in DAY_SEQUENCE:
        lane_count = max(1, lanes_by_day.get(day, 1))
        day_cols[day] = list(range(col, col + lane_count))
        col += lane_count
    return day_cols


def apply_requested_row_heights(ws) -> None:
    """Apply the user's fixed row-height requirements in points.

    - Course title row: 24.6 pt.
    - Header row, Lunch Break, Dinner Break, and all timetable body rows: 40.05 pt.
    - Rows outside the timetable are left unchanged.
    """
    ws.row_dimensions[1].height = 24.6
    for row in range(2, 17):
        ws.row_dimensions[row].height = 40.05


def setup_sheet_layout(ws, title: str, day_cols: Dict[str, List[int]], active_days: set[str]) -> int:
    ws.title = ws.title[:31]
    ws.sheet_view.showGridLines = False
    last_col = max(max(cols) for cols in day_cols.values())

    # v24 visual-size tuning: keep the user's requested 20pt/14pt fonts,
    # use fixed reference-like row heights, and make columns wide enough so
    # content remains readable without very tall rows.
    fixed_widths = {"A": 10, "B": 9, "C": 19}
    for col, width in fixed_widths.items():
        ws.column_dimensions[col].width = width
    # v27: days with no scheduled classes should not occupy a large blank
    # timetable width. Give those days a single narrow column, 12.66 chars.
    # Days with any scheduled class keep the wide readable columns.
    for day in DAY_SEQUENCE:
        width = 12.66 if day not in active_days else 42
        for col_idx in day_cols.get(day, []):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    apply_requested_row_heights(ws)

    # Title and headers.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(1, 1).value = title
    ws.cell(1, 1).fill = FILL_TITLE
    ws.cell(1, 1).font = FONT_TITLE
    ws.cell(1, 1).alignment = ALIGN_CENTER

    headers = {1: "Day", 2: "Class", 3: "ZJSU Timetable\n(FYI)"}
    for col, value in headers.items():
        ws.cell(2, col).value = value
        ws.cell(2, col).font = FONT_HEADER
        ws.cell(2, col).alignment = ALIGN_CENTER

    for day in DAY_SEQUENCE:
        cols = day_cols[day]
        first, last = cols[0], cols[-1]
        if first != last:
            ws.merge_cells(start_row=2, start_column=first, end_row=2, end_column=last)
        cell = ws.cell(2, first)
        cell.value = day
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER

    # Day/session blocks.
    ws.merge_cells("A3:A7")
    ws.cell(3, 1).value = "Morning"
    ws.merge_cells("A9:A12")
    ws.cell(9, 1).value = "Afternoon"
    ws.merge_cells("A14:A16")
    ws.cell(14, 1).value = "Evening"

    # Class/time labels.
    for cls, row in CLASS_TO_ROW.items():
        ws.cell(row, 2).value = cls
        ws.cell(row, 3).value = CLASS_TIMES[cls]

    # Break rows.
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=last_col)
    ws.cell(8, 1).value = "Lunch Break"
    ws.cell(8, 1).fill = FILL_BREAK
    ws.cell(8, 1).font = FONT_BODY
    ws.cell(8, 1).alignment = ALIGN_CENTER

    ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=last_col)
    ws.cell(13, 1).value = "Dinner Break"
    ws.cell(13, 1).fill = FILL_BREAK
    ws.cell(13, 1).font = FONT_BODY
    ws.cell(13, 1).alignment = ALIGN_CENTER

    # Base style and borders over visible timetable area.
    for row in range(1, 17):
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.border = BORDER_TABLE
            cell.alignment = ALIGN_CENTER
            # v23 font rule: course title row uses 20pt; all other visible
            # English/numeric text uses Times New Roman 14pt. Apply fonts
            # unconditionally so class numbers and time labels do not remain
            # in Excel/LibreOffice default Calibri.
            if row == 1:
                cell.font = FONT_TITLE
            elif row == 2:
                cell.font = FONT_HEADER
            else:
                cell.font = FONT_BODY
            if row not in (1, 8, 13):
                # Keep all empty timetable cells visually blank white. Do not use no-fill,
                # because WeChat mobile preview may render transparent .xls cells as black.
                if cell.value is None:
                    cell.fill = FILL_BLANK

    # Ensure merged section labels have the correct font/fill.
    for coord in ("A3", "A9", "A14"):
        ws[coord].font = FONT_BODY
        ws[coord].alignment = ALIGN_CENTER

    ws.freeze_panes = "D3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:{get_column_letter(last_col)}16"
    ws.sheet_view.zoomScale = 55
    return last_col


def event_fill(event: CourseEvent):
    """Return the visual fill for a course block.

    User v17 color rule:
    - Every pure Lecture block is yellow, regardless of duration.
    - Workshop and Lecture/Workshop blocks are blue.
    - Lab/experiment/practical blocks are green, regardless of duration.
    - Blue is also used for Project/Special/Intensive or explicitly special arrangements.
    """
    class_type = (event.class_type or "").lower()
    if "workshop" in class_type or "project" in class_type or "special" in class_type or "intensive" in class_type:
        return FILL_SPECIAL
    if "lab" in class_type:
        return FILL_LAB
    return FILL_LECTURE


def apply_course_block_style(ws, min_row: int, max_row: int, min_col: int, max_col: int, fill) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            cell.fill = fill
            cell.border = BORDER_TABLE
            cell.alignment = ALIGN_CENTER
            cell.font = FONT_BLOCK


def write_course_block(ws, min_row: int, max_row: int, min_col: int, max_col: int, text: str, fill, merge: bool = True) -> None:
    """Write one visible course block and preserve its color in .xls/WeChat.

    Important: style every cell in the target rectangle BEFORE merging.
    Some mobile WeChat/WPS .xls previews do not reliably infer a merged
    region's fill from only the top-left cell. If we merge first and then
    later clear blank cells, the child cells may be treated as blank/white and
    the whole course block can appear uncolored.
    """
    apply_course_block_style(ws, min_row, max_row, min_col, max_col, fill)
    if merge and (max_row > min_row or max_col > min_col):
        ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
    anchor = ws.cell(min_row, min_col)
    anchor.value = text
    anchor.fill = fill
    anchor.border = BORDER_TABLE
    anchor.alignment = ALIGN_CENTER
    anchor.font = FONT_BLOCK


def _scheduled_merged_ranges(ws):
    """Merged course ranges that must not be cleared as empty cells."""
    ranges = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.max_col < 4 or merged_range.min_row in (1, 2, 8, 13):
            continue
        text = str(ws.cell(merged_range.min_row, merged_range.min_col).value or "").strip()
        if text and ("Week" in text or "Lecture" in text or "Lab" in text or "Project" in text):
            ranges.append(merged_range)
    return ranges


def _in_ranges(row: int, col: int, ranges) -> bool:
    return any(r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col for r in ranges)


def _event_rows(event: CourseEvent) -> Tuple[Optional[int], Optional[int]]:
    return CLASS_TO_ROW.get(event.start_period), CLASS_TO_ROW.get(event.end_period)


def _events_overlap(a: CourseEvent, b: CourseEvent) -> bool:
    if a.day != b.day:
        return False
    a_start, a_end = _event_rows(a)
    b_start, b_end = _event_rows(b)
    if not a_start or not a_end or not b_start or not b_end:
        return False
    return max(a_start, b_start) <= min(a_end, b_end)


def _apply_dynamic_row_heights(ws, last_col: int) -> None:
    """Increase row heights after writing blocks so wrapped text is visible.

    Excel does not reliably auto-fit merged cells, so calculate a safe height
    from the number of lines in scheduled course blocks. Height is distributed
    across vertically merged rows when a course spans several class periods.
    """
    # v23: compact rows but keep 14pt text readable. A 4-line single-class
    # block normally needs about 82-88pt; multi-row blocks can share height.
    min_body_height = 72
    points_per_line = 17.5
    extra_padding = 16
    max_row_height = 100
    for row in list(CLASS_TO_ROW.values()):
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, min_body_height)

    for merged in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        if max_row < 3 or min_row > 16 or max_col < 4:
            continue
        text = str(ws.cell(min_row, min_col).value or "").strip()
        if not text:
            continue
        line_count = max(1, text.count("\n") + 1)
        row_count = max(1, max_row - min_row + 1)
        needed_each = min(max_row_height, max(min_body_height, (line_count * points_per_line + extra_padding) / row_count))
        for row in range(min_row, max_row + 1):
            if row not in (8, 13):
                ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, needed_each)

    # Non-merged single cells also need enough height.
    for row in range(3, 17):
        if row in (8, 13):
            continue
        max_lines = 1
        for col in range(4, last_col + 1):
            text = str(ws.cell(row, col).value or "").strip()
            if text:
                max_lines = max(max_lines, text.count("\n") + 1)
        if max_lines > 1:
            needed = min(max_row_height, max(min_body_height, max_lines * points_per_line + extra_padding))
            ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, needed)


def write_events(ws, events: List[CourseEvent], day_cols: Dict[str, List[int]], last_col: int) -> List[dict]:
    warnings: List[dict] = []
    occupied: Dict[Tuple[int, int], bool] = {}
    day_order = {d: i for i, d in enumerate(DAY_SEQUENCE)}
    # Events are already normalized in create_redrawn_workbook. Do not re-merge
    # here, otherwise v26 visual-span extension can accidentally combine blocks
    # and change the source teaching-hour label.
    events = sorted(events, key=lambda e: (day_order.get(e.day, 99), e.start_period, e.end_period, e.weeks, e.class_type, e.class_text, e.room_text))

    for event in events:
        if event.start_period < 1 or event.end_period > 12:
            warnings.append({"event": asdict(event), "warning": "Period outside 1-12; skipped"})
            continue
        start_row = CLASS_TO_ROW.get(event.start_period)
        end_row = CLASS_TO_ROW.get(event.end_period)
        if not start_row or not end_row:
            warnings.append({"event": asdict(event), "warning": "No target row for event period"})
            continue
        cols = day_cols.get(event.day, [])
        if not cols:
            warnings.append({"event": asdict(event), "warning": "No target columns for event day"})
            continue

        crosses_break = (start_row <= 8 <= end_row) or (start_row <= 13 <= end_row)
        has_overlap = any(other is not event and _events_overlap(event, other) for other in events)

        # If a slot has no overlapping course, use the whole day width
        # (for example D:E) instead of only one narrow subcolumn. If the same
        # time contains different rooms/classes, has_overlap=True and each
        # event is forced into a separate lane/cell.
        use_full_day_width = len(cols) > 1 and not has_overlap

        if use_full_day_width and all(
            not occupied.get((row, col))
            for row in range(start_row, end_row + 1)
            for col in cols
        ):
            min_col, max_col = cols[0], cols[-1]
            target_end_row = end_row if not crosses_break else start_row
            write_course_block(ws, start_row, target_end_row, min_col, max_col, event.to_uk_cell_text(), event_fill(event), merge=not crosses_break)
            for row in range(start_row, end_row + 1):
                for col in cols:
                    occupied[(row, col)] = True
            continue

        chosen_col: Optional[int] = None
        for col in cols:
            if all(not occupied.get((row, col)) for row in range(start_row, end_row + 1)):
                chosen_col = col
                break

        if chosen_col is None:
            # No empty subcolumn. Append to the first column's top cell without merging again.
            chosen_col = cols[0]
            anchor = ws.cell(start_row, chosen_col)
            existing = str(anchor.value or "").strip()
            anchor.value = (existing + "\n\n" if existing else "") + event.to_uk_cell_text()
            apply_course_block_style(ws, start_row, min(end_row, start_row), chosen_col, chosen_col, event_fill(event))
            warnings.append({"event": asdict(event), "warning": "No empty subcolumn available; appended into occupied cell"})
            continue

        target_end_row = end_row if not crosses_break else start_row
        write_course_block(ws, start_row, target_end_row, chosen_col, chosen_col, event.to_uk_cell_text(), event_fill(event), merge=(end_row > start_row and not crosses_break))
        for row in range(start_row, end_row + 1):
            occupied[(row, chosen_col)] = True

    # v24: user requested fixed row heights; do not dynamically enlarge rows.
    apply_requested_row_heights(ws)
    return warnings


def clear_empty_fills(ws, last_col: int) -> None:
    """Safety rule: blanks are white, but scheduled merged blocks keep color.

    Do not treat child cells inside a merged course block as empty. They have no
    value by design, but they are still part of a scheduled colored block.
    """
    scheduled_ranges = _scheduled_merged_ranges(ws)
    for row in range(3, 17):
        if row in (8, 13):
            continue
        for col in range(4, last_col + 1):
            if _in_ranges(row, col, scheduled_ranges):
                continue
            cell = ws.cell(row, col)
            if cell.value is None or str(cell.value).strip() == "":
                cell.fill = FILL_BLANK
                cell.border = BORDER_TABLE
                cell.alignment = ALIGN_CENTER
                cell.font = FONT_BODY



def _is_cell_in_any_merged_range(ws, row: int, col: int) -> bool:
    """Return True if a cell belongs to any existing merged range.

    Used before merging blank cells so we never overlap course blocks,
    day headers, title, or break rows. Overlapping merges are a common cause of
    corrupted-looking Excel files in WPS/WeChat preview.
    """
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return True
    return False


def merge_adjacent_blank_cells_by_day(ws, day_cols: Dict[str, List[int]]) -> int:
    """Merge adjacent blank cells inside each day row for a cleaner layout.

    User-facing rule: within the same day, neighbouring empty timetable cells
    should be merged so the sheet does not look unnecessarily fragmented.
    Important boundaries:
    - Merge only horizontally within the same day; never across different days.
    - Do not merge across class-period rows, Lunch Break, or Dinner Break.
    - Do not merge cells that are already part of a scheduled course block.
    - Only truly blank cells are merged; coloured course cells are untouched.
    """
    merged_count = 0
    body_rows = [3, 4, 5, 6, 7, 9, 10, 11, 12, 14, 15, 16]
    for row in body_rows:
        for day in DAY_SEQUENCE:
            cols = day_cols.get(day, [])
            if len(cols) <= 1:
                continue
            run: List[int] = []

            def flush_run() -> None:
                nonlocal merged_count, run
                if len(run) > 1:
                    min_col, max_col = run[0], run[-1]
                    # Style every cell before merging for stable WPS/WeChat display.
                    for c in run:
                        cell = ws.cell(row, c)
                        cell.fill = FILL_BLANK
                        cell.border = BORDER_TABLE
                        cell.font = FONT_BODY
                        cell.alignment = ALIGN_CENTER
                    ws.merge_cells(start_row=row, start_column=min_col, end_row=row, end_column=max_col)
                    anchor = ws.cell(row, min_col)
                    anchor.value = None
                    anchor.fill = FILL_BLANK
                    anchor.border = BORDER_TABLE
                    anchor.font = FONT_BODY
                    anchor.alignment = ALIGN_CENTER
                    merged_count += 1
                run = []

            for col in cols:
                cell = ws.cell(row, col)
                is_blank = (cell.value is None or str(cell.value).strip() == "")
                already_merged = _is_cell_in_any_merged_range(ws, row, col)
                if is_blank and not already_merged:
                    run.append(col)
                else:
                    flush_run()
            flush_run()
    return merged_count


def center_all_timetable_cells(ws, last_col: int) -> None:
    """Force center alignment for all visible timetable cells."""
    for row in range(1, 17):
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.alignment = ALIGN_CENTER


def enforce_break_rows(ws, last_col: int) -> None:
    """Force Lunch Break and Dinner Break to be single merged grey rows.

    In redraw mode, later course-block merges and empty-cell cleanup can leave
    WPS/Excel showing the break row as many small cells if the break merge is
    not enforced at the very end. This function is intentionally called after
    all timetable writing has finished.
    """
    for row, label in ((8, "Lunch Break"), (13, "Dinner Break")):
        # Remove any old/partial merge that touches the break row, then rebuild
        # the exact full-width merge. This makes the rule robust for any course.
        touching = []
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row <= row <= merged_range.max_row:
                touching.append(str(merged_range))
        for coord in touching:
            ws.unmerge_cells(coord)

        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.fill = FILL_BREAK
            cell.border = BORDER_TABLE
            cell.font = FONT_BODY
            cell.alignment = ALIGN_CENTER

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
        anchor = ws.cell(row, 1)
        anchor.value = label
        anchor.fill = FILL_BREAK
        anchor.border = BORDER_TABLE
        anchor.font = FONT_BODY
        anchor.alignment = ALIGN_CENTER
        ws.row_dimensions[row].height = 40.05

def validate_output_layout(ws, events: List[CourseEvent], last_col: int) -> List[dict]:
    """Lightweight safety checks for the generated workbook.

    These checks catch the exact failures reported by the user: break rows not
    merged, missing weekend headers, pure-blue lecture blocks, and duplicate
    visible blocks that should have been collapsed before layout.
    """
    warnings: List[dict] = []
    if str(ws.cell(8, 1).value or "") != "Lunch Break":
        warnings.append({"warning": "Lunch Break row is not correctly anchored at A8"})
    if str(ws.cell(13, 1).value or "") != "Dinner Break":
        warnings.append({"warning": "Dinner Break row is not correctly anchored at A13"})
    header_texts = {str(ws.cell(2, c).value or "").strip() for c in range(1, last_col + 1)}
    for day in DAY_SEQUENCE:
        if day not in header_texts:
            warnings.append({"warning": f"Missing day header: {day}"})

    seen = set()
    dup_count = 0
    for event in events:
        key = event_layout_key(event)
        if key in seen:
            dup_count += 1
        seen.add(key)
    if dup_count:
        warnings.append({"warning": f"Found {dup_count} duplicate visual events before collapse; they were collapsed in output"})

    # Check that ordinary lecture cells are not assigned the old ugly pure blue.
    # ARGB values here are those used by the converter, not palette indices after xls conversion.
    for row in range(3, 17):
        if row in (8, 13):
            continue
        for col in range(4, last_col + 1):
            value = str(ws.cell(row, col).value or "")
            if "Lecture" in value and "Project" not in value:
                fill = ws.cell(row, col).fill
                rgb = (fill.start_color.rgb or "").upper() if fill and fill.start_color else ""
                if rgb in {"FF0000FF", "000000FF"}:
                    warnings.append({"cell": f"{get_column_letter(col)}{row}", "warning": "Ordinary Lecture has pure blue fill; should not be pure blue; all Lecture blocks should be yellow, regardless of duration"})
    return warnings


def create_redrawn_workbook(course_code: str, course_name: str, events: List[CourseEvent], mapping: dict) -> Tuple[Workbook, List[dict]]:
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(course_code, course_name, mapping)
    title = display_title(course_code, course_name, mapping)
    # v36 follows the official English table behavior: merge contiguous source
    # rows for the same teacher/week/class/room/type first, so split source rows
    # such as 6-7 plus 8-9 become one 3h block.  Then merge same-teacher same-
    # time records only within the same week range when they share a room or
    # class.  This prevents Week 13, Week 14, Week 15 lab groups from being
    # collapsed into one block while still allowing real same-slot groups to
    # combine.
    layout_events = apply_visual_span_rules(merge_same_teacher_same_time(merge_contiguous_weeks_same_slot(merge_contiguous_same_events(collapse_identical_events(events)))))
    day_cols = compute_day_columns(layout_events)
    active_days = {e.day for e in layout_events}
    last_col = setup_sheet_layout(ws, title, day_cols, active_days)
    warnings = write_events(ws, layout_events, day_cols, last_col)
    clear_empty_fills(ws, last_col)
    # Final safety pass: break rows must be full-width merged grey bands,
    # never split into individual cells by WPS/Excel.
    enforce_break_rows(ws, last_col)
    # After enforcing merged rows, re-apply white to all truly empty timetable cells.
    # This prevents black backgrounds in WeChat's legacy .xls preview.
    clear_empty_fills(ws, last_col)
    blank_merge_count = merge_adjacent_blank_cells_by_day(ws, day_cols)
    # Final pass after all merges: every visible timetable cell remains centered
    # and row heights match the exact requested point sizes.
    center_all_timetable_cells(ws, last_col)
    apply_requested_row_heights(ws)
    warnings.extend(validate_output_layout(ws, layout_events, last_col))
    inactive_days = [d for d in DAY_SEQUENCE if d not in active_days]
    if inactive_days:
        warnings.append({"info": f"No-class day columns narrowed to 12.66 characters: {', '.join(inactive_days)}"})
    if blank_merge_count:
        warnings.append({"info": f"Merged {blank_merge_count} adjacent blank cell runs within same-day lanes for cleaner layout"})
    return wb, warnings



def save_workbook_for_output(wb: Workbook, output_path: Path, workdir: Path) -> None:
    """Save workbook as .xlsx or .xls without losing colors/layout.

    For .xls, prefer LibreOffice conversion from the correctly styled .xlsx.
    This preserves custom RGB colors, merged cells, row heights and column widths
    much better than a manual writer. If LibreOffice is unavailable, fall back to
    xlwt with an explicitly customized BIFF8 palette. This fixes the old v13 bug
    where #FFC000/#92D050/#9DC3E6 became pure blue/dark green/black in .xls.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    suffix = output_path.suffix.lower()
    if suffix in ("", ".xlsx"):
        wb.save(output_path if suffix else output_path.with_suffix(".xlsx"))
        return
    if suffix != ".xls":
        raise ConversionError(f"output must end with .xls or .xlsx: {output_path}")

    # Preferred path: save .xlsx then let LibreOffice write a real BIFF8 .xls.
    # Use an isolated temporary LibreOffice profile and timeout so agent runs do
    # not hang or reuse a stale office process.
    soffice = find_soffice()
    if soffice:
        tmp_xlsx = workdir / (output_path.stem + "__tmp.xlsx")
        wb.save(tmp_xlsx)
        lo_profile = workdir / "lo_profile"
        lo_profile.mkdir(parents=True, exist_ok=True)
        result = subprocess.run(
            [
                soffice,
                f"-env:UserInstallation=file://{lo_profile}",
                "--headless",
                "--nologo",
                "--nofirststartwizard",
                "--nolockcheck",
                "--nodefault",
                "--convert-to",
                "xls",
                "--outdir",
                str(output_path.parent),
                str(tmp_xlsx),
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=45,
        )
        converted = output_path.parent / (tmp_xlsx.stem + ".xls")
        if result.returncode == 0 and converted.exists():
            if output_path.exists():
                output_path.unlink()
            converted.rename(output_path)
            return
        # Fall through to xlwt only if LibreOffice failed.

    try:
        import xlwt
    except ImportError:
        raise ConversionError(
            "Cannot create .xls output because neither LibreOffice/soffice nor xlwt is available. "
            "Use an .xlsx output path, install LibreOffice, or install xlwt."
        )

    # Fallback path: convert openpyxl workbook to xlwt and explicitly set the palette.
    xls_wb = xlwt.Workbook()
    # Custom palette indices must be in 8..63.
    IDX_WHITE = 0x21
    IDX_YELLOW = 0x22
    IDX_GREEN = 0x23
    IDX_BLUE = 0x24
    IDX_GREY = 0x25
    xls_wb.set_colour_RGB(IDX_WHITE, 255, 255, 255)
    xls_wb.set_colour_RGB(IDX_YELLOW, 255, 192, 0)
    xls_wb.set_colour_RGB(IDX_GREEN, 146, 208, 80)
    xls_wb.set_colour_RGB(IDX_BLUE, 157, 195, 230)
    xls_wb.set_colour_RGB(IDX_GREY, 166, 166, 166)

    xls_ws = xls_wb.add_sheet(wb.active.title[:31])
    ws = wb.active

    # Preserve row heights and column widths.
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = ws.column_dimensions[letter].width or 10
        xls_ws.col(col_idx - 1).width = int(width * 256)
    for row_idx in range(1, ws.max_row + 1):
        height = ws.row_dimensions[row_idx].height or 15
        xls_ws.row(row_idx - 1).height_mismatch = True
        xls_ws.row(row_idx - 1).height = int(height * 20)

    def cell_rgb(cell) -> str:
        fill = cell.fill
        if not fill or fill.fill_type != "solid":
            return "FFFFFFFF"
        color = fill.fgColor
        if color.type == "rgb" and color.rgb:
            rgb = color.rgb.upper()
            if len(rgb) == 6:
                rgb = "FF" + rgb
            return rgb
        return "FFFFFFFF"

    def palette_index_for(cell) -> int:
        rgb = cell_rgb(cell)
        if rgb.endswith("FFC000"):
            return IDX_YELLOW
        if rgb.endswith("92D050"):
            return IDX_GREEN
        if rgb.endswith("9DC3E6"):
            return IDX_BLUE
        if rgb.endswith("A6A6A6"):
            return IDX_GREY
        # Header/no-fill/blank cells should be white, not black.
        return IDX_WHITE

    style_cache = {}
    def make_style(cell):
        fill_idx = palette_index_for(cell)
        font_name = cell.font.name or "Times New Roman"
        font_size = float(cell.font.size or 10)
        font_bold = bool(cell.font.bold)
        key = (fill_idx, font_name, font_size, font_bold)
        if key in style_cache:
            return style_cache[key]
        style = xlwt.XFStyle()
        font = xlwt.Font()
        font.name = font_name
        font.height = int(font_size * 20)
        font.bold = font_bold
        style.font = font
        alignment = xlwt.Alignment()
        alignment.horz = xlwt.Alignment.HORZ_CENTER
        alignment.vert = xlwt.Alignment.VERT_CENTER
        alignment.wrap = xlwt.Alignment.WRAP_AT_RIGHT
        alignment.shrink_to_fit = 1
        style.alignment = alignment
        pattern = xlwt.Pattern()
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        pattern.pattern_fore_colour = fill_idx
        style.pattern = pattern
        borders = xlwt.Borders()
        borders.left = xlwt.Borders.THIN
        borders.right = xlwt.Borders.THIN
        borders.top = xlwt.Borders.THIN
        borders.bottom = xlwt.Borders.THIN
        style.borders = borders
        style_cache[key] = style
        return style

    cell_to_merge = {}
    for merged_range in ws.merged_cells.ranges:
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                cell_to_merge[(r, c)] = (merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col)

    written_cells = set()
    for row in ws.iter_rows():
        for cell in row:
            if (cell.row, cell.column) in written_cells:
                continue
            style = make_style(cell)
            value = cell.value if cell.value is not None else ""
            if (cell.row, cell.column) in cell_to_merge:
                min_row, max_row, min_col, max_col = cell_to_merge[(cell.row, cell.column)]
                if (cell.row, cell.column) == (min_row, min_col):
                    xls_ws.write_merge(min_row - 1, max_row - 1, min_col - 1, max_col - 1, str(value), style)
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            written_cells.add((r, c))
                else:
                    written_cells.add((cell.row, cell.column))
            else:
                xls_ws.write(cell.row - 1, cell.column - 1, str(value), style)
                written_cells.add((cell.row, cell.column))

    xls_wb.save(output_path)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Redraw a UK-style Teaching Calendar from a single Chinese course timetable file.")
    parser.add_argument("--source-file", required=True, type=Path, help="Single uploaded Chinese course timetable .xls/.xlsx file")
    parser.add_argument("--output", required=True, type=Path, help="Output workbook path. Use .xls for mobile WeChat fill-color compatibility; .xlsx also supported.")
    parser.add_argument("--log", type=Path, default=None, help="Optional JSON conversion log path")
    parser.add_argument("--mapping", type=Path, default=None, help="Optional course mapping JSON")
    args = parser.parse_args(argv)

    if not args.source_file.exists():
        raise SystemExit(f"source-file not found: {args.source_file}")
    if args.source_file.suffix.lower() not in {".xls", ".xlsx"}:
        raise SystemExit(f"source-file must be .xls or .xlsx: {args.source_file}")

    script_dir = Path(__file__).resolve().parent
    default_mapping_path = script_dir.parent / "assets" / "course_mapping.json"
    mapping = load_mapping(args.mapping if args.mapping else default_mapping_path)

    log = {
        "skill_version": SKILL_VERSION,
        "mode": "redraw_single_file_no_template",
        "source_file": str(args.source_file),
        "output": str(args.output),
        "courses": [],
        "warnings": [],
        "token_estimate": None,
    }

    with tempfile.TemporaryDirectory(prefix="cn_to_uk_redraw_") as tmp:
        workdir = Path(tmp)
        try:
            code, course_name, events, parse_warnings = parse_chinese_workbook(args.source_file, mapping, workdir)
            token_estimate = estimate_conversion_tokens(args.source_file, code, course_name, events, parse_warnings)
            log["token_estimate"] = token_estimate
            wb, write_warnings = create_redrawn_workbook(code, course_name, events, mapping)
            log["warnings"].extend(parse_warnings)
            log["warnings"].extend(write_warnings)
            log["courses"].append({
                "file": str(args.source_file),
                "course_code": code,
                "course_name": course_name,
                "event_count": len(events),
                "sheet": wb.active.title,
                "token_estimate": token_estimate,
                "status": "redrawn_converted",
            })
            save_workbook_for_output(wb, args.output, workdir)
        except Exception as exc:
            log["warnings"].append({"file": str(args.source_file), "warning": f"Conversion failed: {exc}"})
            log["courses"].append({"file": str(args.source_file), "status": "failed", "error": str(exc)})
            if args.log:
                args.log.parent.mkdir(parents=True, exist_ok=True)
                with args.log.open("w", encoding="utf-8") as f:
                    json.dump(log, f, ensure_ascii=False, indent=2)
            raise SystemExit(str(exc)) from exc

    if args.log:
        args.log.parent.mkdir(parents=True, exist_ok=True)
        with args.log.open("w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False, indent=2)

    print(f"{SKILL_VERSION}: redrawn converted 1/1 course file -> {args.output}")
    if log.get("token_estimate"):
        te = log["token_estimate"]
        print(
            "Token estimate: "
            f"input≈{te['estimated_input_tokens']:,}, "
            f"output≈{te['estimated_output_tokens']:,}, "
            f"total≈{te['estimated_total_tokens']:,} tokens / course"
        )
    if log["warnings"]:
        print(f"Warnings: {len(log['warnings'])}. See log: {args.log or '(not written)'}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
