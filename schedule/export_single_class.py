#!/usr/bin/env python3
"""
导出单个班级课表为 Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 701 班课表数据
class_name = "七年级 701 班"
teacher_info = {
    '班主任': '魏金芬',
    '语文': '武隆峰',
    '数学': '魏金芬',
    '英语': '李梦梦',
    '科学': '潘诗婕',
    '历史与社会': '徐利花',
    '体育': '陆琪',
    '音乐': '沈典轶',
    '美术': '陈辰',
    '劳动': '金根森',
    '综合实践': '李怡卓',
    '班队心理': '魏金芬',
    '地方课程': '鲁博帆'
}

# 课表数据 {day: {period: subject}}
schedule = {
    '周一': {1: ('数学', '魏金芬'), 2: None, 3: ('语文', '武隆峰'), 4: ('科学', '潘诗婕'), 
            5: ('英语', '李梦梦'), 6: ('历史与社会', '徐利花'), 7: ('体育', '陆琪'), 
            8: ('地方课程', '鲁博帆')},
    '周二': {1: ('英语', '李梦梦'), 2: ('科学', '潘诗婕'), 3: ('数学', '魏金芬'), 
            4: ('历史与社会', '徐利花'), 5: None, 6: ('语文', '武隆峰'), 
            7: ('班队心理', '魏金芬'), 8: None},
    '周三': {1: None, 2: ('英语', '李梦梦'), 3: ('数学', '魏金芬'), 
            4: ('历史与社会', '徐利花'), 5: ('科学', '潘诗婕'), 6: ('语文', '武隆峰'), 
            7: ('音乐', '沈典轶'), 8: ('美术', '陈辰')},
    '周四': {1: ('历史与社会', '徐利花'), 2: None, 3: ('科学', '潘诗婕'), 
            4: ('语文', '武隆峰'), 5: ('数学', '魏金芬'), 6: ('英语', '李梦梦'), 
            7: ('劳动', '金根森'), 8: None},
    '周五': {1: ('历史与社会', '徐利花'), 2: ('科学', '潘诗婕'), 3: ('数学', '魏金芬'), 
            4: ('语文', '武隆峰'), 5: ('体育', '陆琪'), 6: ('英语', '李梦梦'), 
            7: ('体育', '陆琪'), 8: ('综合实践', '李怡卓')}
}

DAYS = ['周一', '周二', '周三', '周四', '周五']
PERIODS = list(range(1, 9))

# 创建 Excel
wb = Workbook()
ws = wb.active
ws.title = "701 班课表"

# 样式
TITLE_FONT = Font(bold=True, size=18, color='000000')
HEADER_FONT = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
MAIN_FILL = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
SKILL_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 标题
ws.merge_cells('A1:F1')
ws['A1'] = f'2025 学年第二学期 {class_name} 课程表'
ws['A1'].font = TITLE_FONT
ws['A1'].alignment = CENTER_ALIGN

# 表头
headers = ['节次'] + DAYS
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col)].width = 16

# 课表数据
main_subjects = ['语文', '数学', '英语', '科学', '历史与社会']
skill_subjects = ['体育', '音乐', '美术', '劳动', '综合实践', '班队心理', '地方课程']

for row_idx, period in enumerate(PERIODS, 3):
    ws.cell(row=row_idx, column=1, value=f'第{period}节').alignment = CENTER_ALIGN
    ws.cell(row=row_idx, column=1).border = THIN_BORDER
    
    for col_idx, day in enumerate(DAYS, 2):
        cell = ws.cell(row=row_idx, column=col_idx)
        
        if period in schedule[day] and schedule[day][period]:
            subject, teacher = schedule[day][period]
            cell.value = f'{subject}\n({teacher})'
            
            if subject in main_subjects:
                cell.fill = MAIN_FILL
            elif subject in skill_subjects:
                cell.fill = SKILL_FILL
        else:
            cell.value = ''
        
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

# 教师信息
row_offset = 12
ws.merge_cells(f'A{row_offset}:F{row_offset}')
ws.cell(row=row_offset, column=1, value='任课教师').font = Font(bold=True, size=14)
ws.cell(row=row_offset, column=1).alignment = CENTER_ALIGN

row_offset += 1
for subject, teacher in teacher_info.items():
    ws.cell(row=row_offset, column=1, value=subject)
    ws.cell(row=row_offset, column=2, value=teacher)
    ws.cell(row=row_offset, column=1).border = THIN_BORDER
    ws.cell(row=row_offset, column=2).border = THIN_BORDER
    row_offset += 1

# 保存
output_path = '/home/admin/.openclaw/workspace/schedule/701 班课表.xlsx'
wb.save(output_path)
print(f"已保存：{output_path}")
