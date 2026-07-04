#!/usr/bin/env python3
"""
课表导出为 Excel 格式
"""

import zipfile
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re

# 配置
DAYS = ['周一', '周二', '周三', '周四', '周五']
PERIODS = list(range(1, 9))
main_subjects = ['语文', '数学', '英语', '科学', '历史与社会']
skill_subjects = ['体育', '音乐', '美术', '劳动', '综合实践', '班队心理', '地方课程']

# 样式
HEADER_FONT = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
TITLE_FONT = Font(bold=True, size=14, color='000000')
NORMAL_FONT = Font(size=11)
MAIN_SUBJECT_FILL = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')  # 主科浅灰色
SKILL_SUBJECT_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # 技能课浅黄色

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

def parse_xlsx(filepath):
    """解析原始 xlsx 文件"""
    with zipfile.ZipFile(filepath, 'r') as z:
        strings = []
        if 'xl/sharedStrings.xml' in z.namelist():
            with z.open('xl/sharedStrings.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for si in root.findall('.//ss:si', ns):
                    t = si.find('ss:t', ns)
                    if t is not None:
                        strings.append(t.text if t.text else '')
        
        sheets_data = {}
        for idx, sheet_name in enumerate(['sheet1', 'sheet2', 'sheet3']):
            sheet_path = f'xl/worksheets/{sheet_name}.xml'
            if sheet_path in z.namelist():
                with z.open(sheet_path) as f:
                    tree = ET.parse(f)
                    root = tree.getroot()
                    ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    
                    rows = []
                    for row in root.findall('.//ss:row', ns):
                        row_data = {}
                        for cell in row.findall('ss:c', ns):
                            cell_ref = cell.get('r', '')
                            cell_type = cell.get('t', 'n')
                            v = cell.find('ss:v', ns)
                            value = v.text if v is not None else ''
                            
                            if cell_type == 's' and value.isdigit():
                                value = strings[int(value)] if int(value) < len(strings) else value
                            
                            row_data[cell_ref] = value
                        if row_data:
                            rows.append(row_data)
                    
                    sheets_data[f'sheet{idx+1}'] = rows
        
        return sheets_data, strings

def extract_assignments(sheets_data):
    """提取任课安排"""
    grade_names = {'sheet1': '九年级', 'sheet2': '八年级', 'sheet3': '七年级'}
    COL_MAP = {
        'D': '语文', 'E': '数学', 'F': '英语', 'G': '科学',
        'H': '历史与社会', 'I': '体育', 'J': '音乐', 'K': '美术',
        'L': '劳动', 'M': '综合实践', 'N': '班队心理', 'O': '地方课程'
    }
    
    assignments = {}
    
    for sheet_key, grade_name in grade_names.items():
        if sheet_key not in sheets_data:
            continue
        
        rows = sheets_data[sheet_key]
        for i, row in enumerate(rows):
            if i < 3:
                continue
            
            class_name = row.get(f'A{i+1}', '')
            if not class_name or not class_name.isdigit():
                continue
            
            assignments[class_name] = {'grade': grade_name}
            
            for col, subject in COL_MAP.items():
                teacher = row.get(f'{col}{i+1}', '').strip()
                if teacher:
                    assignments[class_name][subject] = teacher
    
    return assignments

def read_generated_schedules():
    """读取已生成的课表文件"""
    class_schedules = {}
    
    with open('/home/admin/.openclaw/workspace/schedule/class_schedules.txt', 'r', encoding='utf-8') as f:
        content = f.read()
    
    blocks = content.split('【')
    
    for block in blocks[1:]:  # 跳过第一个空块
        if '班】' not in block:
            continue
        
        # 提取班级名
        class_match = re.match(r'(七年级 | 八年级 | 九年级) (\d+) 班', block)
        if not class_match:
            continue
        
        grade = class_match.group(1)
        class_name = class_match.group(2)
        full_class_name = f"{grade}{class_name}班"
        
        # 解析课表
        schedule = {day: {} for day in DAYS}
        lines = block.split('\n')
        
        for line in lines:
            for period in PERIODS:
                if f'第{period}节' in line:
                    # 提取每个老师的课
                    for subject in main_subjects + skill_subjects:
                        pattern = rf'{subject}\(([^)]+)\)'
                        matches = re.findall(pattern, line)
                        for teacher in matches:
                            # 确定是哪一天（根据列位置）
                            parts = line.split()
                            for i, part in enumerate(parts):
                                if subject in part and teacher in part:
                                    day_idx = i - 1  # 第 1 个是节次名
                                    if 0 <= day_idx < len(DAYS):
                                        day = DAYS[day_idx]
                                        schedule[day][period] = (subject, teacher)
                                    break
        
        class_schedules[full_class_name] = {
            'grade': grade,
            'schedule': schedule
        }
    
    return class_schedules

def create_class_schedule_excel(class_schedules, output_path):
    """创建班级课表 Excel"""
    wb = Workbook()
    
    # 删除默认 sheet
    wb.remove(wb.active)
    
    # 按年级分组
    grades = {'七年级': [], '八年级': [], '九年级': []}
    for class_name, data in sorted(class_schedules.items()):
        grades[data['grade']].append((class_name, data['schedule']))
    
    for grade_name in ['七年级', '八年级', '九年级']:
        ws = wb.create_sheet(title=grade_name)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        for i in range(1, 6):
            ws.column_dimensions[get_column_letter(i+1)].width = 18
        
        # 标题行
        ws['A1'] = f'{grade_name}班级课表'
        ws['A1'].font = TITLE_FONT
        
        # 表头
        headers = ['节次'] + DAYS
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
        
        # 为每个班级创建课表
        row_offset = 3
        
        for class_name, schedule in grades[grade_name]:
            # 班级标题
            ws.cell(row=row_offset, column=1, value=class_name).font = Font(bold=True)
            row_offset += 1
            
            # 课表数据
            for period in PERIODS:
                ws.cell(row=row_offset, column=1, value=f'第{period}节').alignment = CENTER_ALIGN
                ws.cell(row=row_offset, column=1).border = THIN_BORDER
                
                for day_idx, day in enumerate(DAYS):
                    col = day_idx + 2
                    cell = ws.cell(row=row_offset, column=col)
                    
                    if period in schedule.get(day, {}):
                        subject, teacher = schedule[day][period]
                        cell.value = f'{subject}\n({teacher})'
                        
                        # 根据学科设置背景色
                        if subject in main_subjects:
                            cell.fill = MAIN_SUBJECT_FILL
                        elif subject in skill_subjects:
                            cell.fill = SKILL_SUBJECT_FILL
                    else:
                        cell.value = ''
                    
                    cell.alignment = CENTER_ALIGN
                    cell.border = THIN_BORDER
                
                row_offset += 1
            
            # 空行分隔
            row_offset += 1
    
    wb.save(output_path)
    print(f"班级课表已保存：{output_path}")

def create_teacher_schedule_excel(output_path):
    """创建教师课表 Excel"""
    with open('/home/admin/.openclaw/workspace/schedule/teacher_schedules.txt', 'r', encoding='utf-8') as f:
        content = f.read()
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # 解析教师课表
    teacher_schedules = {}
    blocks = content.split('【教师：')
    
    for block in blocks[1:]:
        if '】' not in block:
            continue
        
        teacher = block.split('】')[0].strip()
        schedule = {day: {} for day in DAYS}
        
        lines = block.split('\n')
        for line in lines:
            for period in PERIODS:
                if f'第{period}节' in line:
                    for day in DAYS:
                        if day in line:
                            idx = line.find(day)
                            cell = line[idx:idx+20]
                            class_match = re.search(r'(\d+) 班', cell)
                            if class_match:
                                schedule[day][period] = class_match.group(1)
        
        teacher_schedules[teacher] = schedule
    
    # 按姓氏拼音首字母分组（简化为按首字）
    ws = wb.create_sheet(title='教师课表')
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12
    for i in range(1, 6):
        ws.column_dimensions[get_column_letter(i+1)].width = 15
    
    # 表头
    headers = ['教师', '周一', '周二', '周三', '周四', '周五']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # 写入数据
    row = 2
    for teacher in sorted(teacher_schedules.keys()):
        schedule = teacher_schedules[teacher]
        
        # 检查是否有课
        has_class = any(schedule[day] for day in DAYS)
        if not has_class:
            continue
        
        # 创建教师标题行
        ws.cell(row=row, column=1, value=teacher).font = Font(bold=True)
        row += 1
        
        # 课表数据
        for period in PERIODS:
            ws.cell(row=row, column=1, value=f'第{period}节').alignment = CENTER_ALIGN
            ws.cell(row=row, column=1).border = THIN_BORDER
            
            for day_idx, day in enumerate(DAYS):
                col = day_idx + 2
                cell = ws.cell(row=row, column=col)
                
                if period in schedule.get(day, {}):
                    cell.value = f'{schedule[day][period]}班'
                else:
                    cell.value = ''
                
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER
            
            row += 1
        
        # 空行
        row += 2
    
    wb.save(output_path)
    print(f"教师课表已保存：{output_path}")

def create_summary_excel(assignments, output_path):
    """创建任课汇总表"""
    wb = Workbook()
    wb.remove(wb.active)
    
    ws = wb.create_sheet(title='任课汇总')
    
    # 表头
    headers = ['班级', '年级', '班主任', '语文', '数学', '英语', '科学', '历史与社会', 
               '体育', '音乐', '美术', '劳动', '综合实践', '班队心理', '地方课程']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    # 数据
    row = 2
    for class_name in sorted(assignments.keys()):
        data = assignments[class_name]
        grade = data.get('grade', '')
        
        ws.cell(row=row, column=1, value=class_name)
        ws.cell(row=row, column=2, value=grade)
        
        for col, subject in enumerate(headers[2:], 3):
            teacher = data.get(subject, '')
            ws.cell(row=row, column=col, value=teacher)
        
        # 设置边框
        for col in range(1, len(headers)+1):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = CENTER_ALIGN
        
        row += 1
    
    # 调整列宽
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 10
    
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['A'].width = 6
    
    wb.save(output_path)
    print(f"任课汇总已保存：{output_path}")

def create_zip_package():
    """创建压缩包"""
    import os
    
    zip_path = '/home/admin/.openclaw/workspace/schedule/课表汇总.zip'
    
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # 添加 Excel 文件
        files_to_add = [
            '/home/admin/.openclaw/workspace/schedule/班级课表.xlsx',
            '/home/admin/.openclaw/workspace/schedule/教师课表.xlsx',
            '/home/admin/.openclaw/workspace/schedule/任课汇总.xlsx',
            '/home/admin/.openclaw/workspace/schedule/REPORT.md'
        ]
        
        for file_path in files_to_add:
            if os.path.exists(file_path):
                arcname = os.path.basename(file_path)
                zipf.write(file_path, arcname)
                print(f"已添加：{arcname}")
    
    print(f"\n压缩包已创建：{zip_path}")
    return zip_path

def main():
    print("开始解析原始数据...")
    original_file = '/home/admin/.openclaw/media/inbound/61567064-d9e7-4034-aafd-86a2a4067022.xlsx'
    sheets_data, strings = parse_xlsx(original_file)
    assignments = extract_assignments(sheets_data)
    print(f"解析完成：{len(assignments)} 个班级")
    
    print("\n读取生成的课表...")
    class_schedules = read_generated_schedules()
    print(f"读取完成：{len(class_schedules)} 个班级课表")
    
    print("\n生成 Excel 文件...")
    create_class_schedule_excel(class_schedules, '/home/admin/.openclaw/workspace/schedule/班级课表.xlsx')
    create_teacher_schedule_excel('/home/admin/.openclaw/workspace/schedule/教师课表.xlsx')
    create_summary_excel(assignments, '/home/admin/.openclaw/workspace/schedule/任课汇总.xlsx')
    
    print("\n创建压缩包...")
    zip_path = create_zip_package()
    
    print(f"\n✅ 完成！压缩包位置：{zip_path}")
    return zip_path

if __name__ == '__main__':
    main()
